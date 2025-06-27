import pandas as pd
from datetime import datetime, timedelta
from solver.api import solve_for_days_v2
from app import df_to_excel
from io import BytesIO

def test_algorithm_improvement():
    print("=== 알고리즘 개선 테스트 (디폴트 설정) ===")
    
    # 원래 디폴트 설정 그대로 사용
    activities = pd.DataFrame({
        "use": [True, True, True],
        "activity": ["토론면접", "발표준비", "발표면접"],
        "mode": ["batched", "parallel", "individual"],
        "duration_min": [30, 5, 15],
        "room_type": ["토론면접실", "발표준비실", "발표면접실"],
        "min_cap": [4, 1, 1],
        "max_cap": [6, 2, 1],
    })
    
    # 원래 방 설정 그대로 사용
    room_plan = pd.DataFrame({
        "토론면접실_count": [2],
        "토론면접실_cap": [6],
        "발표준비실_count": [1],
        "발표준비실_cap": [2],
        "발표면접실_count": [2],  # 원래대로 2개
        "발표면접실_cap": [1]
    })
    
    oper_window = pd.DataFrame({
        "start_time": ["09:00"],
        "end_time": ["17:30"]
    })
    
    # 원래 선후행 제약 그대로 사용
    precedence = pd.DataFrame([
        {"predecessor": "발표준비", "successor": "발표면접", "gap_min": 5, "adjacent": True}
    ])
    
    # 테스트용 인원
    job_acts_map = pd.DataFrame({
        "code": ["JOB01"],
        "count": [6],  # 6명으로 테스트
        "토론면접": [True],
        "발표준비": [True],
        "발표면접": [True]
    })
    
    tomorrow = datetime.now().date() + timedelta(days=1)
    
    cfg_ui = {
        'activities': activities,
        'job_acts_map': job_acts_map,
        'room_plan': room_plan,
        'oper_window': oper_window,
        'precedence': precedence,
        'interview_dates': [tomorrow],
        'interview_date': tomorrow
    }
    
    print("테스트 설정 (디폴트):")
    print(f"  - 지원자: {job_acts_map['count'].sum()}명")
    print(f"  - 발표준비실: {room_plan.iloc[0]['발표준비실_count']}개 (용량: {room_plan.iloc[0]['발표준비실_cap']}명)")
    print(f"  - 발표면접실: {room_plan.iloc[0]['발표면접실_count']}개 (용량: {room_plan.iloc[0]['발표면접실_cap']}명)")
    print(f"  - 선후행 제약: 발표준비 → 발표면접 (gap: {precedence.iloc[0]['gap_min']}분, adjacent: {precedence.iloc[0]['adjacent']})")
    print("  - 🔧 알고리즘 개선: 후속 활동 예약 시스템 + 스케줄링 순서 최적화")
    
    try:
        # 스케줄링 실행
        status, result, logs, limit = solve_for_days_v2(cfg_ui, debug=False)
        print(f"\n스케줄링 결과: {status}")
        
        if result is not None and not result.empty:
            print(f"성공: {len(result)}개 스케줄 생성")
            
            # 선후행 제약 검증
            print(f"\n=== 알고리즘 개선 효과 검증 ===")
            violations = []
            successes = []
            
            # 지원자별로 검증
            for applicant_id in sorted(result['applicant_id'].unique()):
                applicant_data = result[result['applicant_id'] == applicant_id].copy()
                
                # 발표준비와 발표면접 찾기
                prep_data = applicant_data[applicant_data['activity_name'] == '발표준비']
                present_data = applicant_data[applicant_data['activity_name'] == '발표면접']
                
                if not prep_data.empty and not present_data.empty:
                    prep_end = prep_data.iloc[0]['end_time']
                    present_start = present_data.iloc[0]['start_time']
                    prep_room = prep_data.iloc[0]['room_name']
                    present_room = present_data.iloc[0]['room_name']
                    
                    # timedelta를 분으로 변환
                    if isinstance(prep_end, pd.Timedelta):
                        prep_end_min = prep_end.total_seconds() / 60
                    else:
                        prep_end_min = prep_end.hour * 60 + prep_end.minute
                    
                    if isinstance(present_start, pd.Timedelta):
                        present_start_min = present_start.total_seconds() / 60
                    else:
                        present_start_min = present_start.hour * 60 + present_start.minute
                    
                    gap_min = present_start_min - prep_end_min
                    
                    print(f"  {applicant_id}: {prep_end} ({prep_room}) → {present_start} ({present_room}) | 간격: {gap_min}분")
                    
                    # adjacent=True이면 gap이 정확히 gap_min이어야 함
                    expected_gap = precedence.iloc[0]['gap_min']
                    if precedence.iloc[0]['adjacent']:
                        if abs(gap_min - expected_gap) > 0.1:  # 소수점 오차 허용
                            violations.append(f"{applicant_id}: 연속배치 위반 (실제 간격: {gap_min}분, 예상: {expected_gap}분)")
                            print(f"    ❌ 연속배치 위반")
                        else:
                            successes.append(applicant_id)
                            print(f"    ✅ 연속배치 완벽!")
                    else:
                        if gap_min < expected_gap:
                            violations.append(f"{applicant_id}: 최소 간격 위반 (실제 간격: {gap_min}분, 최소: {expected_gap}분)")
                            print(f"    ❌ 최소 간격 위반")
                        else:
                            successes.append(applicant_id)
                            print(f"    ✅ 최소 간격 준수")
            
            # 결과 요약
            total = len(successes) + len(violations)
            success_rate = len(successes) / total * 100 if total > 0 else 0
            
            print(f"\n=== 🎯 알고리즘 개선 결과 ===")
            print(f"성공: {len(successes)}명 / 총 {total}명")
            print(f"성공률: {success_rate:.1f}%")
            
            if violations:
                print(f"\n❌ 여전히 위반된 케이스:")
                for violation in violations:
                    print(f"  • {violation}")
                    
                print(f"\n💡 추가 개선 방안:")
                print("  1. 더 정교한 시간 슬롯 예약 시스템")
                print("  2. 백트래킹 기반 재배치")
                print("  3. 방 사용 패턴 최적화")
            else:
                print(f"\n🎉 모든 선후행 제약 완벽 준수!")
            
            # Excel 파일 생성하여 확인
            print(f"\n=== Excel 파일 생성 ===")
            excel_buffer = BytesIO()
            df_to_excel(result, excel_buffer)
            
            with open("test_algorithm_improved.xlsx", "wb") as f:
                f.write(excel_buffer.getvalue())
            print(f"✅ 결과 파일 저장: test_algorithm_improved.xlsx")
            
            # 첫행 고정 확인
            import openpyxl
            wb = openpyxl.load_workbook("test_algorithm_improved.xlsx")
            print(f"  시트: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.freeze_panes:
                    print(f"  ✅ {sheet_name}: 첫행 고정 적용")
                    
        else:
            print("❌ 스케줄링 실패")
            print(f"로그: {logs}")
            
    except Exception as e:
        print(f"❌ 테스트 오류: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_algorithm_improvement() 