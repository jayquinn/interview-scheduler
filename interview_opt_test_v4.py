# %%
# ──────────────────────────────────────
import itertools, pandas as pd        # ← import 는 그대로

def _build_param_grid() -> pd.DataFrame:       # ★ 새 함수
    seed_rows = [
        dict(priority=0, scenario_id="S_SAFE", wave_len=35, max_wave=18,
             br_offset_A=4, br_offset_B=3, min_gap_min=5, tl_sec=30)
    ]

    grid = []
    for wl, mw, brA, brB, mg in itertools.product(
            [35], [18], [-2,-1,0,1,2], [-2,-1,0,1,2], [5]):
        if wl==50 and mw==16 and brA==3 and brB==2 and mg==5:
            continue
        pr = 1
        if wl > 35: pr += 1
        if mw < 14: pr += 1
        if mg > 10: pr += 1
        grid.append(dict(priority=pr, wave_len=wl, max_wave=mw,
                         br_offset_A=brA, br_offset_B=brB,
                         min_gap_min=mg, tl_sec=30))

    df = (pd.DataFrame(seed_rows + grid)
            .sort_values(["priority","wave_len","min_gap_min","max_wave"])
            .reset_index(drop=True))

    if "scenario_id" not in df.columns:
        df.insert(0, "scenario_id",
                  [f"S{str(i+1).zfill(3)}" for i in range(len(df))])
    else:
        mask = df["scenario_id"].isna() | (df["scenario_id"]=="")
        df.loc[mask, "scenario_id"] = [
            f"S{str(i+1).zfill(3)}" for i in range(mask.sum())
        ]
    return df
# ──────────────────────────────────────


# %%
# interview_opt_test_v4.py
# -*- coding: utf-8 -*-
"""
============================================================
Interview Schedule Optimiser – multi-date / multi-grid
============================================================
* parameter_grid_test_v4.csv 에 정의된 파라미터 세트 × 날짜를 순차 탐색
* 첫 SAT+하드룰 통과 해를 찾으면 그 날짜는 종료
* 결과는 schedule_wide.csv 로 누적 저장, 시도 내역은 run_log.csv 기록
"""
# ────────────────────────────────
# 0. 공통 import & 상수
# ────────────────────────────────
import sys, itertools, time
from datetime import timedelta
from pathlib import Path
from collections import defaultdict
import yaml
import pandas as pd, yaml
from pandas.api.types import is_integer_dtype
from ortools.sat.python import cp_model
from tqdm import tqdm

# 고정 파일 경로
CAND_CSV = Path("candidate_activities_input_before_test_v4_HF.csv")
GRID_CSV = Path("parameter_grid_test_v4.csv")
YAML_FILE = Path("precedence_config_test_v4_HF.yaml")

OUT_WIDE = Path("schedule_wide_test_v4_HF.csv")
OUT_LOG  = Path("log/run_log_test_v4_HF.csv")
Path("log").mkdir(exist_ok=True)          # log 폴더 보장
from types import SimpleNamespace

def load_params(row):
    """grid CSV 한 행 → 네임스페이스 객체"""
    return SimpleNamespace(
        wave_len     = int(row.wave_len),
        max_wave     = int(row.max_wave),
        br_offset_A  = int(row.br_offset_A),
        br_offset_B  = int(row.br_offset_B),
        min_gap_min  = int(row.min_gap_min),
        tl_sec       = int(row.tl_sec),
    )
# ── activities helper ─────────────────────────────────────
def get_all_activities(yaml_path: Path, df_candidates: pd.DataFrame) -> list[str]:
    """precedence YAML + CSV에서 활동명 set 추출 → 알파벳순 list"""
    prec = yaml.safe_load(open(yaml_path, encoding="utf-8"))

    acts = set(df_candidates["activity"].unique())          # ① CSV에 실제 등장
    # ② YAML – common
    for r in prec.get("common", []):
        acts.update([r["predecessor"], r["successor"]])
    # ③ YAML – by_code
    for branches in prec.get("by_code", {}).values():
        for rules in branches.values():
            for r in rules:
                acts.update([r["predecessor"], r["successor"]])
    return sorted(acts)
# ───────────────────────────────────────────────────────────

# ─────────── 사이클 검증 함수 정의 ───────────
def detect_cycle(edges):
    """
    edges: [(pred, succ), …] 리스트.
    순환이 있으면 True 반환.
    """
    # 1) 그래프 생성
    g = defaultdict(list)
    nodes = set()
    for u, v in edges:
        g[u].append(v)
        nodes.add(u)
        nodes.add(v)

    visited, onstack = set(), set()
    def dfs(u):
        visited.add(u)
        onstack.add(u)
        for w in g.get(u, []):
            if w not in visited:
                if dfs(w):
                    return True
            elif w in onstack:
                return True
        onstack.remove(u)
        return False

    # 2) 미리 추출한 노드 집합으로 순환 검사
    return any(dfs(node) for node in nodes if node not in visited)
# ───────────────────────────────────────────────
def expand_availability(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    UI 에서 받은 집계형 space_availability
      (date · *_count · *_cap · 사용여부 …)
    → solver 가 요구하는
      (date · loc · capacity_max/override) 행 단위 DF 로 변환
    """
    rows = []
    ROOM_TYPES = [
        ("발표면접실", "발표면접실_cap",   "발표면접실_count"),
        ("심층면접실", "심층면접실_cap",   "심층면접실_count"),
        ("커피챗실",   "커피챗실_cap",     "커피챗실_count"),
        ("발표준비실", "발표준비실_cap",   "발표준비실_count"),
    ]

    for _, r in df_raw.iterrows():
        if str(r.get("사용여부", "TRUE")).upper() == "FALSE":
            continue                      # 사용 안 하는 날짜면 skip
        date = pd.to_datetime(r["date"])
        for base, cap_col, cnt_col in ROOM_TYPES:
            n_room = int(r[cnt_col])
            cap    = int(r[cap_col])
            for i in range(1, n_room + 1):
                loc = f"{base}{chr(64+i)}"        # A,B,C…
                rows.append({
                    "date":           date,
                    "loc":            loc,
                    "capacity_max":   cap,        # capacity_override 로 쓰셔도 OK
                })
    return pd.DataFrame(rows)

# ────────────────────────────────
# 1. 하드-룰 검증 함수 (순서 + Wave 정렬)
# ────────────────────────────────
# --- 간결해진 verify_rules ---
# ─── PATCH-3 : 기존 verify_rules() 함수 통째로 갈아끼우기 ───
def verify_rules(wide: pd.DataFrame,
                 yaml_rules: dict,
                 params: dict,                     # ← NEW
                 wave_len: int = 30,
                 company_end = pd.to_timedelta("17:45:00")) -> bool:
     # ────────────── NEW: 파라미터로부터 오프셋 읽기 ──────────────
    # --- optional columns safeguard ---
    for a in ("인성검사", "토론면접"):
        for p in ("loc","start","end"):
            col = f"{p}_{a}"
            if col not in wide.columns:
                wide[col] = pd.NA
    br_offset_A = int(params.get("br_offset_A", 3))   # 디폴트 = 3
    br_offset_B = int(params.get("br_offset_B", 2))   # 디폴트 = 2
    default_codes = {
        c for c, b in yaml_rules.get("by_code", {}).items()
        if "default" in b and not ("A" in b or "B" in b)
    }
    """
    하드룰: ① precedence ② 토론-인성 δ-격자(0-60, 5분) ③ 17:30 이전 종료
    위반 시 바로 ❌ 로그 출력 후 False
    """
    for _, r in wide.iterrows():
        cid = r["id"]
        arr_off = 0 if str(r["loc_인성검사"]).endswith("A") else 5
        code = r["code"]
        if code in default_codes:
            branch = "default"
        else:
            branch = "A" if arr_off == 0 else "B"
        # ---------- ① precedence ----------
        def viol(p, s, g):
            if pd.isna(r[f"start_{p}"]) or pd.isna(r[f"start_{s}"]):
                return False
            if r[f"end_{p}"] + pd.Timedelta(minutes=g) > r[f"start_{s}"]:
                print(f"❌ precedence {cid}: {p}->{s}  "
                      f"end={r[f'end_{p}'].time()} "
                      f"start={r[f'start_{s}'].time()} gap={g}")
                return True
            return False

        for rule in yaml_rules.get("common", []):
            if viol(rule["predecessor"], rule["successor"], rule["min_gap_min"]):
                return False
        for code, branches in yaml_rules.get("by_code", {}).items():
            if r["code"] != code:
                continue
            if code in default_codes:
                rules_iter = branches.get("default", [])
            else:
                rules_iter = branches.get(branch, []) + branches.get("default", [])
            for rule in rules_iter:
                if viol(rule["predecessor"], rule["successor"], rule["min_gap_min"]):
                    return False

        # ---------- ② δ-격자 ----------
        if pd.notna(r["start_토론면접"]):
            raw = (r["start_토론면접"] - r["start_인성검사"]).total_seconds() / 60

            if branch == "A" or branch == "default":        # 인성 → 토론
                base = raw - br_offset_A * wave_len         #   δ = slide (0‥60)
            else:                                           # branch == "B"  토론 → 인성
                base = raw + br_offset_B * wave_len         #   (raw가 음수 ⇒ + 로 보정)

            if base % 5 or not (0 <= base <= 60):
                print(f"❌ δ-grid  {cid}: rawΔ={raw}  base={base}")
                return False
        # ---------- ③ 17:30 초과 ----------
        # wide DF 안에 실제로 있는 end_ 컬럼만 검사
        for col in [c for c in wide.columns if c.startswith("end_")]:
            if pd.isna(r[col]):
                continue
            if (r[col] - r[col].normalize()) > company_end:
                print(f"❌ overtime {cid}: {col}={r[col].time()}")
                return False


    return True

       # ← lit 도 같이
# ────────────────────────────────
# 2. build_model() –  **엔진**  (기존 코드를 함수화·경량화)
#    →  SAT & 규칙통과 : ('OK', wide_df)
#       규칙위반      : ('RULE', None)
#       UNSAT/타임아웃 : ('UNSAT', None)
#       예외          : ('ERR',  None)
# ────────────────────────────────
def build_model(the_date: pd.Timestamp,
                params: dict,
                cfg: dict      # ← CSV 묶음
               ) -> tuple[str, pd.DataFrame | None]:
    """
    반환 ('OK'|'RULE_VIOL'|'UNSAT'|'ERR', wide_df or None)
    """



    try:
        # ── 2-1. 날짜 필터 ──
        df_cand = cfg["df_raw"][cfg["df_raw"]["interview_date"] == the_date].copy()
        if df_cand.empty:
            return "NO_DATA", None
        ALL_ACTS = get_all_activities(YAML_FILE, df_cand)  
        # ── 2-2. 짧은 별칭 ──
        cfg_duration = cfg["cfg_duration"].copy()
        cfg_avail    = cfg["cfg_avail"].copy()
        # 집계형 테이블(date · *_count …)이면 행 단위(loc) 형태로 펼친다
        if "loc" not in cfg_avail.columns:
            cfg_avail = expand_availability(cfg_avail)
        cfg_map      = cfg["cfg_map"]
        cfg_oper     = cfg["cfg_oper"]
        prec_yaml = cfg["prec_yaml"]
        # ▶️ NEW ───────────────────────────────────────
        group_meta      = cfg.get("group_meta", pd.DataFrame()).copy()
        MODE            = group_meta.set_index("activity")["mode"].to_dict()
        MIN_CAP_ACT     = group_meta.set_index("activity")["min_cap"].to_dict()
        MAX_CAP_ACT     = group_meta.set_index("activity")["max_cap"].to_dict()
        # ──────────────────────────────────────────────
        # ▶️ NEW: invalid config → 즉시 예외
        for act, md in MODE.items():
            if md != "batched" and MAX_CAP_ACT.get(act, 1) > 1:
                raise ValueError(f"[CONFIG] '{act}' 는 individual 인데 max_cap > 1")
            if MIN_CAP_ACT.get(act, 1) > MAX_CAP_ACT.get(act, 1):
                raise ValueError(f"[CONFIG] '{act}'  min_cap > max_cap")
        # ───────────────────────── 모델 파라미터 ─────────────────────────
        WAVE_LEN = int(params["wave_len"])
        MAX_WAVE = int(params["max_wave"])
        tl_sec   = int(params["tl_sec"])
        MIN_GAP  = int(params["min_gap_min"])
        # ────────────────────────────────────────────────────────────────
        # ═════════════════════ 0. 상수 및 설정 ═════════════════════
        INPUT_CSV    = CAND_CSV

        DEBUG = True

        TALK_GMIN, TALK_GMAX = 3, 5
        W_MAKESPAN, W_GAP_AB, W_WAIT, W_SOFT, W_OT = 1000, 300, 200, 200, 1
        W_SLIDE = 80

        ENABLE_CAPACITY  = True
        COMPANY_END_CAL  = 17*60 + 45



        # ═════════════════════ 2. 전처리 ═════════════════════
        # 2-1 duration & act_space
        cfg_duration["duration_min"] = pd.to_numeric(
            cfg_duration["duration_min"], errors="coerce"
        ).fillna(0).astype(int)
        DUR = cfg_duration.set_index("activity")["duration_min"].to_dict()

        df_cand = df_cand.merge(
            cfg_duration[["activity","duration_min"]],
            on="activity", how="left"
        )
        if df_cand["duration_min"].isna().any():
            miss = df_cand.loc[df_cand["duration_min"].isna(), "activity"].unique()
            raise ValueError(f"[CONFIG] duration 값이 없는 활동: {miss}")

        df_cand["duration_min"] = df_cand["duration_min"].astype(int)

        ACT_SPACE = cfg_map.groupby("activity")["loc"].apply(list).to_dict()
        DEBATE_ROOMS = ACT_SPACE.get("토론면접", [])

        if DEBUG:                        # 🐞 새 디버그 출력
            print("[bm] ACT_SPACE keys:", list(ACT_SPACE.keys())[:5])

        for a in ACT_SPACE:
            ACT_SPACE[a].sort()
        def get_space(act: str):
            """ACT_SPACE 안전 래퍼 – 정의돼 있지 않으면 빈 리스트 반환"""
            return ACT_SPACE.get(act, [])
        # 2-2 capacity
        cfg_avail["capacity_effective"] = pd.to_numeric(
            cfg_avail["capacity_override"], errors="coerce"
        ).fillna(cfg_avail["capacity_max"]).astype(int)
        CAP = cfg_avail.set_index(["loc","date"])["capacity_effective"].to_dict()
        if DEBUG:
            print("[bm] CAP sample   :", list(CAP.items())[:5])
        # 2-3 operating window
        cfg_oper["start_dt"] = pd.to_datetime(
            cfg_oper["date"].dt.strftime("%Y-%m-%d") + " " + cfg_oper["start_time"]
        )
        cfg_oper["end_dt"] = pd.to_datetime(
            cfg_oper["date"].dt.strftime("%Y-%m-%d") + " " + cfg_oper["end_time"]
        )
        # ① 운영창 dict
        OPER = cfg_oper.set_index(["code", "date"])[["start_dt", "end_dt"]].to_dict("index")

        # ② 각 전형별 운영 길이·수평선
        OPER_LEN = {
            (c, d): int((v["end_dt"] - v["start_dt"]).total_seconds() // 60)
            for (c, d), v in OPER.items()
        }
        HORIZON = max(OPER_LEN.values())
        # # ─────────── YAML 로드: 기본 branch 코드 파악 ───────────
        # prec_yaml = yaml.safe_load(open(YAML_FILE, encoding="utf-8"))
        # default_codes = {
        #     c for c, b in prec_yaml.get("by_code", {}).items()
        #     if "default" in b and not ("A" in b or "B" in b)
        # }
        default_codes = {
        c for c, b in prec_yaml.get("by_code", {}).items()
        if "default" in b and not ("A" in b or "B" in b)
        }
        # 2-4 lookup dicts
        CIDS     = sorted(df_cand["id"].unique())
        CODE_MAP = df_cand.set_index("id")["code"].to_dict()

        # ═════ 2-5 팀 & 동선 초기 배정 ═════════════════════
        isA         = {cid: (i % 2 == 0) for i,cid in enumerate(CIDS)}
        init_Tfirst = {cid: (i % 2 == 1) for i,cid in enumerate(CIDS)}

        # ═════════════════════ 3. 모델 구축 ═════════════════════
        # ───────── helper (모델 내부 전용) ─────────────────────────
        def _apply_prec_constraint(cid, pred, succ, min_gap, extra_lits=None):
            """
            cid      : 지원자 id
            pred,succ: activity 이름
            min_gap  : 분
            extra_lits (list[BoolVar]) : 조건부 활성화용 추가 리터럴
            """
            if extra_lits is None:
                extra_lits = []

            for loc_p in ACT_SPACE.get(pred, []):
                if (cid, pred, loc_p) not in sel:
                    continue
                for loc_s in ACT_SPACE.get(succ, []):
                    if (cid, succ, loc_s) not in sel:
                        continue

                    lit = model.NewBoolVar(f"PRE_{cid}_{pred}->{succ}")
                    model.AddAssumption(lit)
                    ASSUMPTIONS.append(lit)
                    ASSUME_IDX[lit.Index()] = lit
                    conds = [sel[cid, pred, loc_p],
                            sel[cid, succ, loc_s],
                            lit] + extra_lits

                    model.Add(
                        start[cid, succ, loc_s] + ARR_OFF[cid] >=
                        end[cid,   pred, loc_p] + ARR_OFF[cid] + min_gap
                    ).OnlyEnforceIf(conds)
        ASSUMPTIONS = []
        ASSUME_IDX = {}
        model = cp_model.CpModel()
        start,end,sel,iv = {},{},{},{}
        cid_iv, loc_iv = defaultdict(list), defaultdict(list)

        # ----------  Branch 선택 BoolVar + Offset IntVar  ----------
        isA_lit, ARR_OFF, BR_OFFSET = {}, {}, {}

        OFFSET_A = params['br_offset_A']     # e.g. 4
        OFFSET_B = params['br_offset_B']     # e.g. 3

        for cid in CIDS:
            # 1) A-branch?  BoolVar
            lit = model.NewBoolVar(f"isA_{cid}")
            model.AddHint(lit, 1 if isA[cid] else 0)   # 예전 isA[] 값은 힌트로만
            isA_lit[cid] = lit

            # 2) ARR_OFF   0 (A) / 5 (B)
            arr = model.NewIntVar(0, 5, f"arrOff_{cid}")
            ARR_OFF[cid] = arr
            model.Add(arr == 0).OnlyEnforceIf(lit)
            model.Add(arr == 5).OnlyEnforceIf(lit.Not())

            # 3) BR_OFFSET  offset_A / offset_B
            # br = model.NewIntVar(min(OFFSET_A, OFFSET_B),
            #                     max(OFFSET_A, OFFSET_B),
            #                     f"brOff_{cid}")
            # BR_OFFSET[cid] = br
            # model.Add(br == OFFSET_A).OnlyEnforceIf(lit)
            # model.Add(br == OFFSET_B).OnlyEnforceIf(lit.Not())
            BR_OFFSET[cid] = (isA_lit[cid]       *  OFFSET_A +
                            isA_lit[cid].Not() * (-OFFSET_B))
        # -----------------------------------------------------------
        # IntervalVar 생성
        for _, row in df_cand.iterrows():
            cid, act, dur = row["id"], row["activity"], row["duration_min"]
            for loc in ACT_SPACE[act]:
                if act in ("인성검사", "발표면접"):
                    pass
                if CAP.get((loc, the_date), 0) == 0:
                    continue

                # ---- Interval & decision var ----
                s = model.NewIntVar(0, HORIZON, f"s_{cid}_{act}_{loc}")
                e = model.NewIntVar(0, HORIZON, f"e_{cid}_{act}_{loc}")
                x = model.NewBoolVar(f"x_{cid}_{act}_{loc}")
                ivar = model.NewOptionalIntervalVar(s, dur, e, x,
                                                    f"iv_{cid}_{act}_{loc}")
                start[cid, act, loc], end[cid, act, loc], sel[cid, act, loc], iv[cid, act, loc] = (
                    s, e, x, ivar
                )
                if MODE.get(act, "individual") != "parallel":   # parallel 은 NoOverlap 대상 제외
                    cid_iv[cid].append(ivar)
                if dur > 0:
                    loc_iv[loc, the_date].append(ivar)

                # ---- ① 5분 격자 강제 (발표 준비·면접) ----
                if act in ("발표준비", "발표면접"):
                    k = model.NewIntVar(0, HORIZON // 5,
                                        f"k_{cid}_{act}_{loc}")  # 5-분 그리드 인덱스

                    # start + ARR_OFF 가 5 의 배수 ⇒ 5 분 격자
                    model.Add(s + ARR_OFF[cid] == k * 5).OnlyEnforceIf(x)

            # ---- ② exactly-1 loc per activity (loc loop 바깥) ----
            model.Add(
                sum(sel[cid, act, l]                       # ← sel dict 는 위에서 채움
                    for l in ACT_SPACE[act]
                    if (cid, act, l) in sel) == 1
            )

        # 3-1b T_FIRST
        T_FIRST = {cid: model.NewBoolVar(f"Tfirst_{cid}") for cid in CIDS}
        for cid in CIDS:
            if init_Tfirst[cid]:
                model.AddHint(T_FIRST[cid],1)
        # 3-2 (선택적) 그룹 활동 Wave / δ-slide 로직 ────────
        SLIDE_UNIT = 5            # 5분 단위
        SLIDE_MAX  = 12           # 0‥60분
        # HAS_GROUP = bool(get_space("인성검사")) and bool(get_space("토론면접"))
        # 날짜별 지원자 목록에 토론면접이 실제로 존재하는지로 판단
        # AFTER  ── (build_model() 상단, MODE·MIN_CAP_ACT·MAX_CAP_ACT 만든 바로 아래) ──
        BATCH_ACTS = [a for a, m in MODE.items() if m == "batched"]
        HAS_GROUP  = df_cand["activity"].isin(BATCH_ACTS).any()

        # 모든 batched 활동에 대해 {act: cap} dict 로 보관
        GROUP_MIN = {a: MIN_CAP_ACT.get(a, 1) for a in BATCH_ACTS}
        GROUP_MAX = {a: MAX_CAP_ACT.get(a, 999) for a in BATCH_ACTS}
        # HAS_GROUP을 계산한 **뒤**에 가중치 결정
        gap_ab_wt = 0 if not HAS_GROUP else W_GAP_AB
        slide_wt  = 0 if not HAS_GROUP else W_SLIDE
        # ---------- (1) 공통 더미 변수 먼저 만들기 ----------
        # → HAS_GROUP 이 False 여도 아래쪽 코드가 안전하게 참조 가능
        y = {}                                          # wave 배정 BoolVar
        delta_unit = []                                # 전체 δ-unit(IntVar) 리스트
        delta_unit_cid = {cid: model.NewIntVar(0, 0, f"dSel_{cid}") for cid in CIDS}
        I_wave =      {cid: model.NewIntVar(0, 0, f"Iwave_{cid}")   for cid in CIDS}
        slide_penalty = 0                              # 목적함수용 기본값
        if HAS_GROUP:
            print("=== Wave capacity debug ===")
            y = {}
            for cid in CIDS:
                for room in ("인성검사실A","인성검사실B","인성검사실C"):
                    if (cid,"인성검사",room) not in sel: continue
                    for w in range(MAX_WAVE):
                        y[cid,room,w] = model.NewBoolVar(f"y_{cid}_{room}_{w}")
                        model.Add(sel[cid,"인성검사",room] == 1).OnlyEnforceIf(y[cid,room,w])
                        model.Add(start[cid,"인성검사",room] == w*WAVE_LEN).OnlyEnforceIf(y[cid,room,w])
                    model.Add(sum(y[cid,room,w] for w in range(MAX_WAVE)) == sel[cid,"인성검사",room])

            # wave 동시입실(3–5)
            for room in ("인성검사실A","인성검사실B","인성검사실C"):
                for w in range(MAX_WAVE):
                    members = [y[c,room,w] for c in CIDS if (c,room,w) in y]
                    if not members: continue
                    non_empty = model.NewBoolVar(f"nonEmpty_{room}_{w}")
                    model.Add(sum(members)>0).OnlyEnforceIf(non_empty)
                    model.Add(sum(members)==0).OnlyEnforceIf(non_empty.Not())
                    # model.Add(sum(members)>=3).OnlyEnforceIf(non_empty)
                    # model.Add(sum(members)<=5).OnlyEnforceIf(non_empty)
                    # model.Add(sum(members) >= GROUP_MIN).OnlyEnforceIf(non_empty)
                    # model.Add(sum(members) <= GROUP_MAX).OnlyEnforceIf(non_empty)
                    act = "인성검사"      # ← 이미 그 블록이 인성검사용이면 하드코딩 그대로 둬도 무방
                    model.Add(sum(members) >= GROUP_MIN[act]).OnlyEnforceIf(non_empty)
                    model.Add(sum(members) <= GROUP_MAX[act]).OnlyEnforceIf(non_empty)
            # 인성검사 종료 시각 저장
            # I_END = {}
            # for cid in CIDS:
            #     lst = [end[cid, "인성검사", l]
            #         for l in ACT_SPACE.get("인성검사", [])
            #         if (cid, "인성검사", l) in end]
            #     if lst:                      # 인성검사 있는 경우에만 저장
            #         I_END[cid] = lst[0]

            # I_wave 정의
            I_wave = {}
            for cid in CIDS:
                wvar = model.NewIntVar(0, MAX_WAVE-1, f"Iwave_{cid}")
                I_wave[cid] = wvar
                for room in ("인성검사실A","인성검사실B",'인성검사실C'):
                    key = (cid,"인성검사",room)
                    if key in start:
                        model.Add(wvar*WAVE_LEN == start[key]).OnlyEnforceIf(sel[key])
            # --- B-브랜치(토론→인성) 최소 wave 하한 ---  ← 새 코드
            if OFFSET_B > 0:                                         # 안전가드
                for cid in CIDS:
                    model.Add(I_wave[cid] >= OFFSET_B).OnlyEnforceIf(isA_lit[cid].Not())
            # ─── δ-slide용 z 변수 및 “exactly-one” 제약 (★ 방별로 계산) ──────────
            deb_z = {}                               # (cid, loc, abs_t) → BoolVar

            # 0) 토론 방 목록 & 방별 좌석 수
            DEBATE_ROOMS = ACT_SPACE.get("토론면접", [])
            ROOM_CAP = {loc: CAP[(loc, the_date)] for loc in DEBATE_ROOMS}  # ex) 5

            for cid in CIDS:
                # 이 지원자가 실제로 갖고 있는 토론 interval 키 모으기
                debate_keys = [
                    (cid, "토론면접", loc)
                    for loc in DEBATE_ROOMS
                    if (cid, "토론면접", loc) in sel
                ]
                if not debate_keys:            # 토론이 없는 지원자
                    continue

                for loc in DEBATE_ROOMS:
                    if (cid, "토론면접", loc) not in sel:
                        continue
                    for w in range(MAX_WAVE):
                        for du in range(SLIDE_MAX + 1):      # δ-unit 0‥12
                            for offset, lit_ok in [(OFFSET_A,  isA_lit[cid]),   # A-branch
                                                ( -OFFSET_B, isA_lit[cid].Not() )]:  # B
                                abs_t = (w + offset) * WAVE_LEN + du * SLIDE_UNIT
                                if abs_t < 0 or abs_t > HORIZON + 60:   # 안전가드
                                    continue

                                z = model.NewBoolVar(f"deb_{cid}_{loc}_{abs_t}")
                                deb_z[cid, loc, abs_t] = z

                                model.Add(sel[cid,"토론면접",loc] == 1).OnlyEnforceIf([z, lit_ok])
                                model.Add(start[cid,"토론면접",loc] + ARR_OFF[cid] == abs_t).OnlyEnforceIf([z, lit_ok])
                # 한 지원자당 z 하나만 1
                model.Add(
                    sum(deb_z[cid, loc, t]
                        for loc in DEBATE_ROOMS
                        for t in range(0, HORIZON + 60, SLIDE_UNIT)
                        if (cid, loc, t) in deb_z) == 1
                )

            # ★ 방·시각별 인원 3‥방수용량(cap) 제약
            for loc in DEBATE_ROOMS:
                cap = ROOM_CAP[loc]              # 대부분 5
                for abs_t in range(0, HORIZON + 60, SLIDE_UNIT):
                    members = [deb_z[cid, loc, abs_t]
                            for cid in CIDS if (cid, loc, abs_t) in deb_z]
                    if not members:
                        continue
                    non_empty = model.NewBoolVar(f"deb_nonEmpty_{loc}_{abs_t}")
                    # model.Add(sum(members) >= 3).OnlyEnforceIf(non_empty)     # 켜졌으면 ≥3
                    # model.Add(sum(members) <= 2).OnlyEnforceIf(non_empty.Not())   # 꺼졌으면 ≤2
                    # model.Add(sum(members) <= cap).OnlyEnforceIf(non_empty)\
                    act = "토론면접"
                    model.Add(sum(members) >= GROUP_MIN[act]).OnlyEnforceIf(non_empty)
                    model.Add(sum(members) <= 2).OnlyEnforceIf(non_empty.Not())       # (이 줄은 그대로)
                    model.Add(sum(members) <= cap).OnlyEnforceIf(non_empty)
            # ── δ-unit(IntVar) – wave당 하나 ─────────────────
            delta_unit = [
                model.NewIntVar(0, SLIDE_MAX, f"deltaUnit_{w:02d}")
                for w in range(MAX_WAVE)
            ]

            # ── 후보별 δ 선택 변수 & Element 제약 ──────────────
            delta_unit_cid = {}
            for cid in CIDS:
                delta_unit_cid[cid] = model.NewIntVar(
                    0, SLIDE_MAX, f"deltaSel_{cid}")
                model.AddElement(I_wave[cid],
                                delta_unit,
                                delta_unit_cid[cid])

            # ── 목적함수용 페널티 계산 ────────────────────────
            slide_penalty = sum(delta_unit) * SLIDE_UNIT
        else:
            I_wave = {cid: model.NewIntVar(0, 0, f"Iwave_{cid}") for cid in CIDS}
            delta_unit      = []
            delta_unit_cid  = {cid: model.NewIntVar(0, 0, f"deltaSel_{cid}") for cid in CIDS}
            slide_penalty   = 0
        # ───── 토론면접 start 식 수정 ───────────────────────────────────────
# === δ-slide 선택 변수는 이미 정의 ===
# delta : list[IntVar]   delta_cid : dict[cid -> IntVar]
# ─── 토론면접 start 등식 ─────────────────────────
        # ─── 토론면접 start 등식 (ARR_OFF 제거‧재배치) ─────────────────
        # for cid in CIDS:
        #     for loc in DEBATE_ROOMS:
        #         key_T = (cid, "토론면접", loc)
        #         if key_T not in sel: continue

        #         model.Add(
        #             start[key_T]
        #             ==
        #             (I_wave[cid] + BR_OFFSET[cid]) * WAVE_LEN
        #             + delta_unit_cid[cid] * SLIDE_UNIT
        #         ).OnlyEnforceIf(sel[key_T])
        for cid in CIDS:
            for loc in DEBATE_ROOMS:
                key = (cid, "토론면접", loc)
                if key not in sel: continue

                # A-branch: 토론 = 인성 + OFFSET_A
                model.Add(
                    start[key] ==
                    (I_wave[cid] + OFFSET_A) * WAVE_LEN +
                    delta_unit_cid[cid] * SLIDE_UNIT
                ).OnlyEnforceIf(isA_lit[cid])

                # B-branch: 토론 = 인성 – OFFSET_B
                model.Add(
                    start[key] ==
                    (I_wave[cid] - OFFSET_B) * WAVE_LEN +
                    delta_unit_cid[cid] * SLIDE_UNIT
                ).OnlyEnforceIf(isA_lit[cid].Not())


        # ───────── 토론실 capacity ─────────
        for loc in DEBATE_ROOMS:
            iv_list = [ iv[cid, "토론면접", loc]
                        for cid in CIDS if (cid, "토론면접", loc) in iv ]
            if iv_list:
                max_cap = CAP[(loc, the_date)]                  # CSV 값 (5)
                model.AddCumulative(iv_list, [1]*len(iv_list), max_cap)

        # ─────────── YAML 로드 (prec 정의) ───────────
        prec = prec_yaml

                # ───────── 3-2b Debug: precedence 적용 전 가능 옵션 확인 ─────────
        print("=== PRECEDENCE OPTIONS DEBUG ===")
        for cid in CIDS:
            code = CODE_MAP[cid]
            branch = "default" if code in default_codes else ("A" if isA[cid] else "B")
            # 공통(common)
            for rule in prec.get("common", []):
                p,s = rule["predecessor"], rule["successor"]
                preds = [loc for loc in ACT_SPACE.get(p,[])     if (cid,p,loc) in sel]
                succs = [loc for loc in ACT_SPACE.get(s,[])     if (cid,s,loc) in sel]
                if not preds or not succs:
                    print(f"CID={cid}  common {p}->{s}: preds={preds}  succs={succs}")
            # by_code
            for rules in prec.get("by_code", {}).get(code, {}).get(branch, []):
                p,s = rules["predecessor"], rules["successor"]
                preds = [loc for loc in ACT_SPACE.get(p,[])     if (cid,p,loc) in sel]
                succs = [loc for loc in ACT_SPACE.get(s,[])     if (cid,s,loc) in sel]
                if not preds or not succs:
                    print(f"CID={cid}  {code}-{branch} {p}->{s}: preds={preds}  succs={succs}")
        print("=== END DEBUG ===\n")
        # ───────── 3-3 precedence from YAML ─────────
        # prec = yaml.safe_load(open(YAML_FILE, encoding="utf-8"))
        prec = prec_yaml
        # --- 사이클 검증 실행 ---
        # 1) 공통 제약 쌍 수집
        common_edges = [
            (c["predecessor"], c["successor"])
            for c in prec.get("common", [])
        ]
        # 2) 코드·브랜치별로 검사
        for code, branches in prec.get("by_code", {}).items():
            for branch, rules in branches.items():
                edges = common_edges + [
                    (r["predecessor"], r["successor"])
                    for r in rules
                ]
                if detect_cycle(edges):
                    print("[UNSAT] precedence cycle detected")
                    return 'UNSAT', None
        # 3-3a) 모든 지원자에게 공통으로 적용할 순서 제약
        for c in prec.get("common", []):
            for cid in CIDS:
                _apply_prec_constraint(cid,
                                    c["predecessor"],
                                    c["successor"],
                                    c["min_gap_min"])

        # 3-3b) 코드별·branch별 특별 제약  ← 전체 교체
        # ─── 3-3b) 코드별·branch별 특별 제약 ─────────────────────────
        for cid in CIDS:
            code = CODE_MAP[cid]
            lit = isA_lit[cid]
            # ①-a 인성검사 방과 lit 연결
            for loc in ("인성검사실A","인성검사실B","인성검사실C"):
                key = (cid, "인성검사", loc)
                if key in sel:
                    if loc.endswith("A"):
                        model.Add(lit == 1).OnlyEnforceIf(sel[key])
                    else:
                        model.Add(lit == 0).OnlyEnforceIf(sel[key])



            # ①-b  인성검사가 **아예 없는** 지원자는 자유 변수
            #       (힌트를 주고 싶으면 여기서 model.AddHint(isA_lit, …) 가능)
            #       별도 제약은 주지 않는다.

            # ───── branch 가려내서 precedence 넣기 ─────
            br_A_rules = prec.get("by_code", {}).get(code, {}).get("A", [])
            br_B_rules = prec.get("by_code", {}).get(code, {}).get("B", [])
            def_rules  = prec.get("by_code", {}).get(code, {}).get("default", [])

            # A-branch rules
            for r in br_A_rules:
                _apply_prec_constraint(cid,
                                    r["predecessor"], r["successor"],
                                    r["min_gap_min"],
                                    extra_lits=[lit])

            # B-branch rules
            for r in br_B_rules:
                _apply_prec_constraint(cid,
                                    r["predecessor"], r["successor"],
                                    r["min_gap_min"],
                                    extra_lits=[lit.Not()])

            # default rules (branch 무관)
            for r in def_rules:
                _apply_prec_constraint(cid,
                                    r["predecessor"], r["successor"],
                                    r["min_gap_min"])
        # ───────────────────────────────────────────────
        # 3-4 NoOverlap & Capacity
        for ivs in cid_iv.values():
            model.AddNoOverlap(ivs)
        if ENABLE_CAPACITY:
            for (loc,date), ivs in loc_iv.items():
                model.AddCumulative(ivs, [1]*len(ivs), CAP[loc,date])

        # 3-5  회사 종료시간 제한  (기존 코드 대체)
        for (cid, act, loc), x in sel.items():
            oper_len = OPER_LEN[(CODE_MAP[cid], the_date)]      # ← 전형별 길이
            model.Add(end[cid, act, loc] + ARR_OFF[cid] <= oper_len).OnlyEnforceIf(x)

        # 3-7 Soft same-code wave
        codes = set(CODE_MAP.values())
        M_wave = {code: model.NewIntVar(0,MAX_WAVE-1,f"Mwave_{code}") for code in codes}
        diff_vars = []
        for cid in CIDS:
            d = model.NewIntVar(0,MAX_WAVE,f"diff_{cid}")
            model.AddAbsEquality(d, I_wave[cid] - M_wave[CODE_MAP[cid]])
            diff_vars.append(d)
        soft_penalty = sum(diff_vars)



        # 목적함수
        real_end = []
        for (cid,act,loc),x in sel.items():
            rend = model.NewIntVar(0,COMPANY_END_CAL+10,f"rend_{cid}_{act}")
            model.Add(rend == end[cid,act,loc] + ARR_OFF[cid]).OnlyEnforceIf(x)
            real_end.append(rend)
        makespan = model.NewIntVar(0,COMPANY_END_CAL+10,"makespan")
        model.AddMaxEquality(makespan, real_end)

        cntA = sum(x for (cid,a,l),x in sel.items() if a=="발표면접" and l.endswith("A"))
        cntB = sum(x for (cid,a,l),x in sel.items() if a=="발표면접" and l.endswith("B"))
        gapAB = model.NewIntVar(0,len(CIDS),"gapAB")
        model.AddAbsEquality(gapAB, cntA-cntB)
        model.Add(gapAB <= 2)
        OBJ = (W_MAKESPAN * makespan +
            gap_ab_wt   * gapAB +
            W_SOFT      * soft_penalty +
            slide_wt    * slide_penalty)
        model.Minimize(OBJ)

        # ═════════════════════ 4. Solver ═════════════════════
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = tl_sec        # 30 초 제한
        solver.parameters.num_search_workers  = 1
        solver.parameters.max_memory_in_mb    = 0
        solver.parameters.stop_after_first_solution = True
        solver.parameters.log_search_progress = True

        status = solver.Solve(model)
        # ───────────── precedence GAP 샘플 체크 ─────────────
        if status == cp_model.INFEASIBLE:        # UNSAT 인 경우에만 찍어 보자
            try:
                sample = []
                for cid in CIDS[:15]:            # 앞 15명만 샘플
                    for lp in ACT_SPACE['발표준비']:
                        for ls in ACT_SPACE['발표면접']:
                            if (cid,'발표준비',lp) in start and (cid,'발표면접',ls) in start:
                                # 아직 값이 없어도 Var 는 존재 → 최적값 대신 lower/upper bound 이용
                                gap_lb = start[cid,'발표면접',ls].Proto().domain[0] - \
                                        end  [cid,'발표준비',lp].Proto().domain[-1]
                                sample.append((cid, gap_lb))
                print("[bm] GAP lower-bounds (cid, min_gap_min 후보):", sample)
            except Exception as e:
                print("[bm] gap-debug failed:", e)
        # ───────────────────────────────────────────────────

        ARR_OFF_VAL = {cid: solver.Value(ARR_OFF[cid]) for cid in CIDS}
        BR_OFFSET_VAL = {cid: solver.Value(BR_OFFSET[cid]) for cid in CIDS}
        # ── INFEASIBLE 처리 ─────────────────────────────────
        if status == cp_model.INFEASIBLE:
            if DEBUG and tl_sec >= 5:
                try:
                    core = solver.SufficientAssumptionsForInfeasibility()
                    print("❌ UNSAT core size:", len(core))
                    for lit in core[:20]:
                        lv   = int(lit)                         # ← ① int() 로 캐스팅
                        sign = "¬" if lv < 0 else ""            # ← ② 이제 int 로 비교
                        var  = ASSUME_IDX.get(abs(lv))
                        name = var.Name() if var is not None else f"lit#{lv}"
                        print("   ⊠", f"{sign}{name}")



                except Exception as e:
                    print("[WARN] UNSAT-core fetch failed:", e)
            else:
                print("[INFO] UNSAT (core skipped)")
            return "UNSAT", None

        # SAT 해가 아니면 바로 종료
        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            return "UNSAT", None

        # ───────────────── 디버그 출력 (SAT 해일 때만) ─────────
        if DEBUG:
            # 4-1 Wave capacity (인성검사 그룹)
            if HAS_GROUP:
                print("=== Wave capacity debug ===")
                for w in range(MAX_WAVE):
                    a = sum(solver.Value(v)
                            for (cid, room, ww), v in y.items()
                            if room == "인성검사실A" and ww == w)
                    b = sum(solver.Value(v)
                            for (cid, room, ww), v in y.items()
                            if room == "인성검사실B" and ww == w)
                    c = sum(solver.Value(v)
                            for (cid, room, ww), v in y.items()
                            if room == "인성검사실C" and ww == w)
                    print(f"Wave{w:02d}: A={a}, B={b}, C={c}")
                print("=== End wave debug ===")

            # 4-2 토론면접 확정 start 시각
            print("=== Debug: 실제 할당된 토론면접 start times (minutes) ===")
            for cid in CIDS:
                for loc in DEBATE_ROOMS:
                    key = (cid, "토론면접", loc)
                    if key in sel and solver.Value(sel[key]):
                        real_start = solver.Value(start[key]) + ARR_OFF_VAL[cid]
                        print(f"{cid}({loc}): start+ARR_OFF = {real_start}")
            print("=============================================")

            # 4-3 raw 토론면접 start 시각
            print("=== Debug: raw 토론면접 start times (minutes) ===")
            for cid in CIDS:
                for loc in DEBATE_ROOMS:
                    key = (cid, "토론면접", loc)
                    if key in start:
                        val = solver.Value(start[key])
                        print(f"{cid}({loc}): start={val} (+ARR_OFF {ARR_OFF_VAL[cid]})")
            print("=============================================")

            # 4-4 precedence gap 확인 (샘플 5명)
            for cid in CIDS[:5]:
                print(f"--- Debug for {cid} ---")
                # ① common rules
                for idx, rule in enumerate(prec.get("common", [])):
                    pred, succ, gap = rule["predecessor"], rule["successor"], rule["min_gap_min"]
                    for loc_p in ACT_SPACE.get(pred, []):
                        if (cid, pred, loc_p) not in sel: continue
                        for loc_s in ACT_SPACE.get(succ, []):
                            if (cid, succ, loc_s) not in sel: continue
                            pe = solver.Value(end[cid, pred, loc_p])
                            ss = solver.Value(start[cid, succ, loc_s])
                            print(f"  [common#{idx}] {pred}->{succ}: end={pe}, start={ss}, gap={gap}")
                # ② by_code rules
                for code, branches in prec.get("by_code", {}).items():
                    if CODE_MAP[cid] != code: continue
                    branch = "default" if code in default_codes else ("A" if isA[cid] else "B")
                    for idx, rule in enumerate(branches.get(branch, [])):
                        pred, succ, gap = rule["predecessor"], rule["successor"], rule["min_gap_min"]
                        for loc_p in ACT_SPACE.get(pred, []):
                            if (cid, pred, loc_p) not in sel: continue
                            for loc_s in ACT_SPACE.get(succ, []):
                                if (cid, succ, loc_s) not in sel: continue
                                pe = solver.Value(end[cid, pred, loc_p])
                                ss = solver.Value(start[cid, succ, loc_s])
                                print(f"  [{code}-{branch}#{idx}] {pred}->{succ}: end={pe}, start={ss}, gap={gap}")

        # ═════════════════════ 5. 결과 출력 ═════════════════════
        rows = []
        for (cid,act,loc),x in sel.items():
            if solver.Value(x)==0: continue
            st = solver.Value(start[cid,act,loc]) + ARR_OFF_VAL[cid]
            ed = solver.Value(end  [cid,act,loc]) + ARR_OFF_VAL[cid]
            base = OPER[CODE_MAP[cid],the_date]["start_dt"]
            rows.append({
                "id":cid,
                "code":CODE_MAP[cid],
                "interview_date":the_date,
                "wave":solver.Value(I_wave[cid]),
                "activity":act,
                "loc":loc,
                "start":base + timedelta(minutes=st),
                "end":base   + timedelta(minutes=ed)
            })

        df_long = pd.DataFrame(rows)
        CORE_ACTS = ALL_ACTS      # pivot 에 쓸 최종 활동 리스트
        wide = (
            df_long[df_long["activity"].isin(CORE_ACTS)]
            .pivot_table(
                index=["id","code","interview_date"],
                columns="activity",
                values=["loc","start","end"],
                aggfunc="first"
            )
        )
        for a in ("인성검사", "토론면접"):
            for p in ("loc", "start", "end"):
                col = f"{p}_{a}"
                if col not in wide.columns:
                    wide[col] = pd.NA
        wide.columns = [f"{t}_{a}" for t,a in wide.columns]
        wide = wide.reset_index()
        order_cols = (
            ["id","code","interview_date"]
        + list(itertools.chain.from_iterable([[f"loc_{a}",f"start_{a}",f"end_{a}"] for a in CORE_ACTS]))
        )
        wide = wide.reindex(columns=order_cols, fill_value=pd.NA)
        # ────── wide DF 만들고 난 뒤 ──────
        wide["wave"] = wide["id"].map({cid: solver.Value(I_wave[cid]) for cid in CIDS})

        # === 🔽 중복 없는 카운트 컬럼 추가 ===============
        # 1) 카운트 계산
        # 1) wave cnt (인성검사 없으면 건너뜀)
        if "loc_인성검사" in wide.columns and "start_인성검사" in wide.columns:
            cnt_wave = (wide.groupby(["interview_date","wave","loc_인성검사"])["id"]
                            .transform("size"))
        else:
            cnt_wave = pd.NA

        # 2) debate cnt (토론면접 없으면 건너뜀)
        if "start_토론면접" in wide.columns:
            cnt_debate = (wide.groupby(["interview_date","start_토론면접"])["id"]
                            .transform("size"))
        else:
            cnt_debate = pd.NA
        # 2) 이미 있다면 제거
        for col in ("wave_in_cnt", "debate_in_cnt"):
            if col in wide.columns:
                wide.drop(columns=[col], inplace=True)

        # 3) 원하는 위치에 삽입
        wave_pos = wide.columns.get_loc("wave") + 1   # wave 바로 뒤
        wide.insert(wave_pos, "wave_in_cnt", cnt_wave)
        wide["debate_in_cnt"] = cnt_debate            # 맨 끝에 둠(필요하면 insert 사용)

        # =================================================

        wave_map = {cid: solver.Value(I_wave[cid]) for cid in CIDS}
        rule_ok = verify_rules(wide, prec_yaml, params, wave_len=WAVE_LEN)
                
        if rule_ok:
            return 'OK', wide
        else:
            print("❗ RULE FAIL at", cid)
            print("[RULE_VIOL] order / grid mis-match")
            return 'RULE_VIOL', None



    except Exception as e:
        import traceback, sys
        traceback.print_exc()      # 전체 콜스택을 콘솔에 출력
        raise                      # 예외를 다시 올려 solve() → Streamlit 까지 전달


# ────────────────────────────────
# 3. main –  날짜 × 파라미터 루프
# ────────────────────────────────
def main():
    # ── 0) 지원자 CSV 한 번만 읽기 ──────────────────────────────
    df_raw = (
        pd.read_csv(CAND_CSV, encoding="utf-8-sig")      # 필요하면 utf-8-sig, cp949
          .assign(activity=lambda d: d["activity"].str.split(","))
          .explode("activity")
          .assign(activity=lambda d: d["activity"].str.strip())
          .assign(
              interview_date=lambda d:
                  pd.to_datetime(d["interview_date"], errors="coerce")
          )
    )

    # ── 1) 날짜 리스트 자동 생성 ────────────────────────────────
    date_list = (
        pd.to_datetime(df_raw["interview_date"].dropna().unique())
    )
    date_list = pd.DatetimeIndex(date_list).sort_values()

    # ── 2) 공통 cfg 모음 ───────────────────────────────────────
    cfg = {
        "df_raw"       : df_raw,                              # ← 재활용
        "cfg_duration" : pd.read_csv("duration_config_test_v4_HF.csv"),
        "cfg_avail"    : pd.read_csv("space_availability_test_v4_HF.csv",
                                     parse_dates=["date"]),
        "cfg_map"      : pd.read_csv("activity_space_map_test_v4_HF.csv"),
        "cfg_oper"     : pd.read_csv("operating_config_test_v4_HF.csv",
                                     parse_dates=["date"]),
    }

    # (3) 파라미터 그리드
    grid = pd.read_csv(GRID_CSV, dtype=str).fillna('')
    if 'scenario_id' not in grid.columns:
        grid.insert(0, 'scenario_id', range(1, len(grid)+1))
    if 'tl_sec' not in grid.columns:
        grid['tl_sec'] = '30'
    for col, default in [("br_offset_A", "2"), ("br_offset_B", "1")]:
        if col not in grid.columns:
            grid[col] = default

    # (4) 실행
    all_wides, log_rows = [], []
    for d in tqdm(date_list, desc="Dates"):

        ok_found = False
        for _, p in grid.iterrows():
            params = {k: (int(v) if str(v).lstrip('-').isdigit() else v) for k,v in p.items()}
            status, wide = build_model(d, params, cfg)
            if status == "NO_DATA":
                continue      # 해당 날짜에 지원자 없음
            log_rows.append({
                "date"    : d.date(),
                "scenario": p["scenario_id"],
                "status"  : status,
                "success" : 1 if status == "OK" else 0
            })


            if status == "OK":
                all_wides.append(wide)
                ok_found = True
                break   # 다음 날짜

        if not ok_found:
            print(f"[WARN] {d.date()} – feasible 해 없음")

    # (5) 저장
    pd.DataFrame(log_rows).to_csv(OUT_LOG, mode='w', index=False,
                              encoding="utf-8-sig")

    if all_wides:
        pd.concat(all_wides).to_csv(OUT_WIDE, index=False, encoding="utf-8-sig")
        print(f"[SAVE] {OUT_WIDE}")
    print("[SAVE] run_log_test_v4.csv")


# ────────────────────────────────
# 4. 실행
# ────────────────────────────────
if __name__ == "__main__":
    # (1) 파라미터 그리드 csv 저장 – 스크립트를 직접 실행할 때만
    _build_param_grid().to_csv(
        "parameter_grid_test_v4.csv",
        index=False, encoding="utf-8-sig")
    print("📝 parameter_grid_test_v4.csv 생성 완료")

    # (2) 기존 main() 호출
    main()
# %%
# interview_opt_test_v4.py
# 0. import & 상수 ───────────────────────────────────────────
import re, itertools, yaml
from collections import Counter, defaultdict, deque
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, numbers

CSV  = Path('schedule_wide_test_v4_HF.csv')     # ← 입력
XLSX = Path('schedule_view_test_v4_HF.xlsx')    # ← 결과
YAML = Path('precedence_config_test_v4.yaml')

META        = ['id', 'code', 'interview_date']
GROUP_ACTS  = ('인성검사', '토론면접')
PALETTE     = ['E3F2FD','FFF3E0','E8F5E9','FCE4EC','E1F5FE',
               'F3E5F5','FFFDE7','E0F2F1','EFEBE9','ECEFF1']


#1. 입력 ────────────────────────────────────────────────
def load_csv(path: Path) -> pd.DataFrame:
    cols_dt = [c for c in pd.read_csv(path, nrows=0)
               if re.match(r'(start|end)_', c)]
    df = pd.read_csv(path, parse_dates=cols_dt, keep_default_na=False)
    def is_useless(col: str, s: pd.Series) -> bool:
        """
        열 전체가 비어 있거나(값이 하나도 없음) 정보량이 1 이하이면 True를 반환.
        loc_/start_/end_ 로 시작하더라도 예외 없이 검사한다.
        """
        nunique = s.replace('', pd.NA).nunique(dropna=True)   # '' → NaN 처리 후 고유값 개수
        return nunique == 0 if re.match(r'^(loc|start|end)_', col) else nunique <= 1

    # def is_useless(col: str, s: pd.Series) -> bool:
    #     # (1) 메타(id/code/…) 는 놔둔다
    #     if not re.match(r'^(loc|start|end)_', col):
    #         return s.replace('', pd.NA).nunique(dropna=True) <= 1
    #     # (2) 활동 컬럼은 절대 지우지 않는다
    #     return False
    df.drop(columns=[c for c in df if is_useless(c, df[c])], inplace=True)
    return df

#2. 활동 순서(위상정렬 – 시각 기반) ─────────────────────

PREFIXES = ('loc_', 'start_', 'end_')          # 3-관절
IS_ACT = lambda c: any(c.startswith(p) for p in PREFIXES)
JOINT_RANK = {'loc': 0, 'start': 1, 'end': 2}  # loc<start<end
def detect_variants(df: pd.DataFrame,
                    order: dict[tuple[str, str], int]) -> None:

    # ── 0) 기본 순위만 추려 두기 ───────────────────────────────
    base_rank = {act: r for (act, var), r in order.items() if var == ''}
    base_acts = set(base_rank)

    # ── ★ 보조 함수 정의 ★ ──────────────────────────────────
    def _is_okay(act: str,
                 times: dict[str, str],
                 rank: dict[str, int]) -> bool:
        """
        해당 활동(act)의 시작시간이 ‘정상 순서’ 안에 들어있으면 True.
        times  : {활동: "HH:MM"}
        rank   : {활동: 예상순위}
        """
        t_act = pd.to_datetime(times[act])
        # act 앞뒤로 와야 할 활동들의 시간
        for other, t_other in times.items():
            if other == act:
                continue
            t_other = pd.to_datetime(t_other)
            if rank[other] < rank[act] and t_other > t_act:   # 앞에 와야 할 놈이 뒤에 있다
                return False
            if rank[other] > rank[act] and t_other < t_act:   # 뒤에 와야 할 놈이 앞에 있다
                return False
        return True

    # ── 1) 행 단위 스캔 ─────────────────────────────────────
    for idx, row in df.iterrows():
        # 1-A) 시작시간 모으기 (NA 안전)
        times = {act: row[f'start_{act}']
                 for act in base_acts
                 if f'start_{act}' in df.columns
                 and pd.notna(row[f'start_{act}'])
                 and row[f'start_{act}'] != ''}
        if not times:
            continue

        # 2) 예상·실제 순서
        expect = sorted(times, key=lambda a: base_rank[a])
        actual = sorted(times, key=lambda a: pd.to_datetime(times[a]))
        if actual == expect:
            continue            # 이상 없음

        # 3) 순서를 깨뜨린 활동만 변종으로 이동
        for act in actual:
            if _is_okay(act, times, base_rank):
                continue        # 정상 위치면 skip

            # 3-A) 비어 있는 v2/v3… 열 찾기
            ver = 2
            while f'start_{act}_v{ver}' in df.columns and df.at[idx, f'start_{act}_v{ver}'] != '':
                ver += 1

            # 3-B) 이동 (loc/start/end 세트)
            for joint in ('loc', 'start', 'end'):
                base_col = f'{joint}_{act}'
                var_col  = f'{joint}_{act}_v{ver}'
                if var_col not in df.columns:
                    df[var_col] = ''      # 새 열 추가 (빈 값)
                df.at[idx, var_col], df.at[idx, base_col] = df.at[idx, base_col], ''

            break   # 한 행당 한 활동만 변종 처리 (원래 로직 유지)



def split_col(col: str):
    """loc_토론면접_v2 → ('loc','토론면접','_v2')"""
    joint, body = col.split('_', 1)
    m = re.match(r'(.+?)(_v\d+)?$', body)
    return joint, m.group(1), m.group(2) or ''
def build_graph(df: pd.DataFrame):
    from collections import Counter, defaultdict, deque

    nodes  = set()

    # ──────────────────────────────
    # ① start_* 칼럼 → (act, '') 또는 (act, 'v2') …
    # ──────────────────────────────
    for c in df.columns:
        if not c.startswith('start_'):
            continue
        m = re.match(r'start_(.+?)(?:_v(\d+))?$', c)
        if m:
            act, ver = m.groups()
            nodes.add((act, f'v{ver}' if ver else ''))

    # ──────────────────────────────
    # ② 행 단위 다수결 투표
    #    **여기도 튜플 키만 사용**
    # ──────────────────────────────
    votes = defaultdict(Counter)      # {(a,b): Counter({'AB':n,'BA':m})}

    for _, row in df.iterrows():
        times = {}
        for col in df.columns:
            if not col.startswith('start_'):
                continue
            val = row[col]
            if val == '' or pd.isna(val):
                continue

            act = re.match(r'start_(.+?)(?:_v\d+)?$', col).group(1)
            key = (act, '')                          # ★★ 항상 튜플로 ★★
            t   = pd.to_datetime(val)

            # 동일 활동이 여러 번 등장하면 가장 이른 시간만
            times[key] = min(times.get(key, t), t)

        # 투표
        for a, b in itertools.permutations(times, 2):
            votes[(a, b)]['AB' if times[a] < times[b] else 'BA'] += 1
            nodes.update([a, b])       # 혹시 빠진 노드 있을까 봐 보강

    # ──────────────────────────────
    # ③ 간선‧위상정렬 부분은 그대로,
    #    a, b 가 이제 모두 튜플이므로 추가 수정 필요 없음
    # ──────────────────────────────


    # ---------------------------------------------
    # 2) 우세 방향만 간선으로 — 가중치 = 득표차
    # ---------------------------------------------
    G, weight = defaultdict(set), {}
    for (a, b), cnt in votes.items():
        if cnt['AB'] > cnt['BA']:
            G[a].add(b); weight[(a, b)] = cnt['AB'] - cnt['BA']
        elif cnt['BA'] > cnt['AB']:
            G[b].add(a); weight[(b, a)] = cnt['BA'] - cnt['AB']
        # 완전 동점이면 간선 없음

    # ---------------------------------------------
    # 3) Condorcet 순환 깨기 (가장 약한 간선부터 제거)
    # ---------------------------------------------
    while True:
        indeg = defaultdict(int)
        for u in G:
            for v in G[u]:
                indeg[v] += 1
                indeg.setdefault(u, 0)

        # Kahn 위상정렬 시도
        q      = deque([n for n in nodes if indeg.get(n, 0) == 0])
        seen   = set()
        while q:
            u = q.popleft(); seen.add(u)
            for v in G[u]:
                indeg[v] -= 1
                if indeg[v] == 0:
                    q.append(v)

        if len(seen) == len(nodes):          # ✅ DAG 완성
            break

        # 순환에 남은 노드 & 가장 약한 간선 찾기
        cyclic_nodes   = nodes - seen
        cyclic_edges   = [(u, v) for u in cyclic_nodes for v in G[u] if v in cyclic_nodes]
        weakest_edge   = min(cyclic_edges, key=lambda e: weight.get(e, 1))
        G[weakest_edge[0]].remove(weakest_edge[1])
        # 루프가 무한히 돌 일은 없습니다: 간선은 유한, 매 회 하나씩 제거

    return nodes, G

def topo_sort(nodes, G):
    indeg = {n: 0 for n in nodes}
    for u in G:
        for v in G[u]:
            indeg[v] += 1
    q = deque([n for n,d in indeg.items() if d == 0])
    order = []
    while q:
        u = q.popleft()
        order.append(u)
        for v in G[u]:
            indeg[v] -= 1
            if indeg[v] == 0:
                q.append(v)
    if len(order) != len(nodes):
        raise ValueError('⚠️  cycle detected in activity order!')
    return {n:i for i,n in enumerate(order)}
# 변종이 기본보다 얼마나 ‘당겨/미룰’지
SHIFT = {
    '':      0,     # 기본
    '_v2': -0.4,    # 제일 먼저
    '_v3': -0.3,
    '_v4': -0.2,
    # 그 이후는 0.01씩 뒤로 미룬다 → 예: _v5 = +0.01
}
def order_key_factory(order_map):
    def key(col: str):
        # 0) 메타·카운트 열 빠른 리턴 ----------------------------
        if col in META or '_' not in col:
            return (-1, 1, 0, col)   # (가중치 맞추기용 dummy)

        joint, base, var = split_col(col)

        if (base, '') not in order_map:     # in_cnt 같은 열
            return (-1, 1, 0, col)

        base_rank = order_map[(base, '')]

        # ── 1) 변종 rank 계산 ----------------------------------
        if var:                               # ex) _v2
            n = int(var[2:])                  # 2
            act_rank = base_rank - n          # 5-2=3
            variant_flag = 0                  # ★ 변종은 0 (우선)
        else:
            act_rank = base_rank
            variant_flag = 1                  # 본판은 1 (뒤)

        # ── 2) 최종 key ---------------------------------------
        # (rank, variant_flag, joint_rank, col)
        return (act_rank,
                variant_flag,
                JOINT_RANK[joint],
                col)
    return key


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    nodes, G   = build_graph(df)
    order_map  = topo_sort(nodes, G)
    new_cols   = sorted(df.columns, key=order_key_factory(order_map))
    return df.loc[:, new_cols]

# ------------------------------------------------------------
def prepare_schedule(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    solver 가 뱉은 wide-DF를
    (1) 변종(_v2 …) 정리 → (2) 열 재배열 → (3) wave 컬럼/정렬까지
    마친 최종 테이블로 변환한다.
    """
    df = df_raw.copy()

    # A. 기본 선후관계 → order_map
    nodes, G = build_graph(df)
    order_map = topo_sort(nodes, G)

    # B. 변종 열 이동
    detect_variants(df, order_map)

    # C. wave 등 그룹 보조 컬럼
    add_group_cols(df)

    # E. 열 순서 정리
    new_cols = sorted(df.columns, key=order_key_factory(order_map))
    df = df.loc[:, META + [c for c in new_cols if c not in META]]

    # F. 행 정렬 (첫 start → date → wave → code)
    start_cols = [c for c in df if c.startswith('start_')]
    df['_sort_key'] = (df[start_cols]
                       .apply(pd.to_datetime, errors='coerce')
                       .min(axis=1))
    sort_cols = ['_sort_key', 'interview_date']
    if 'wave' in df.columns:
        sort_cols.append('wave')
    sort_cols.append('code')
    return (df
            .sort_values(sort_cols)
            .drop(columns='_sort_key')
            .reset_index(drop=True))
# ------------------------------------------------------------


#3. 집단활동/wave 보조 칼럼 ─────f────────────────────────
def add_group_cols(df: pd.DataFrame) -> bool:
    has_group = any(f'start_{a}' in df for a in GROUP_ACTS)
    if not has_group: return False
    if 'wave' not in df:
        df.insert(df.columns.get_loc('id') + 1, 'wave', pd.NA)
    df['wave_in_cnt'] = (df.groupby(['interview_date', 'wave'])['id']
                           .transform('size'))
    if 'start_토론면접' in df:
        df['debate_in_cnt'] = (df.groupby(['interview_date', 'start_토론면접'])['id']
                                 .transform('size'))
    return True


#4. 엑셀로 저장 ─────────────────────────────────────────
def df_to_excel(df: pd.DataFrame, by_wave: bool, stream=None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'
    df = df.copy()

    # ── wave 컬럼: dtype 상관없이 숫자로 맞춤 ───────────────────
    if 'wave' in df.columns:
        df['wave'] = (pd.to_numeric(df['wave'], errors='coerce')  # 숫자 아닌 건 NaN
                        .fillna(-1)                               # NaN → -1
                        .astype(int))                             # int 로 통일
    # ── ⬇️ 이 줄을 바로 **아래**에 삽입 ⬇️
    use_wave_color = (by_wave                      # 외부에서 켜졌고
                    and 'wave' in df.columns     # wave 컬럼이 있으며
                    and (df['wave'] >= 0).any()) # 실제로 ≥0 값이 하나라도

    # 나머지는 전부 None 으로
    df = df.astype(object).where(pd.notna(df), None)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    act_color = {}
    time_cols = [c for c in df if c.startswith(('start_', 'end_'))]
    wave_col_idx = (df.columns.get_loc('wave') + 1               # 1-based
                    if 'wave' in df.columns else None)
    for j, col in enumerate(df.columns, 1):
        m = re.match(r'(loc|start|end)_(.+?)', col)
        if m:
            act = m.group(2)
            act_color.setdefault(act, PALETTE[len(act_color) % len(PALETTE)])
            hdr = act_color[act]
            ws.cell(1, j).fill = PatternFill('solid', fgColor=hdr)
            for i in range(2, ws.max_row + 1):
                w = ws.cell(i, wave_col_idx).value if wave_col_idx else None
                color = (PALETTE[int(w) % len(PALETTE)]
                        if (use_wave_color and w is not None) else hdr)
                ws.cell(i, j).fill = PatternFill('solid', fgColor=color)
        if col in time_cols:
            for i in range(2, ws.max_row + 1):
                ws.cell(i, j).number_format = 'hh:mm'

    wb.save(stream or XLSX)
    print('✅ saved:', XLSX)


#5. 엔트리 포인트 ───────────────────────────────────────────
def export_schedule_view() -> None:
    df = load_csv(CSV)

    # (A) 기본 활동 선후관계 구하기
    nodes, G  = build_graph(df)
    order_map = topo_sort(nodes, G)

    # (B) 변종(_v2 …) 이동
    detect_variants(df, order_map)

    # (C) 그룹(시트)용 컬럼 확보 ― wave 가 없으면 여기서 만들어 둠
    by_wave = add_group_cols(df)

    # (D) 칼럼 재정렬
    new_cols = sorted(df.columns, key=order_key_factory(order_map))
    df = df.loc[:, META + [c for c in new_cols if c not in META]]

    # ────────────────────────────────────────────────────────
    # (E) ★ 첫 시작시각 → interview_date → wave → code 순 정렬 ★
    start_cols = [c for c in df.columns if c.startswith('start_')]   # ← 모든 start_ 컬럼

    df['_sort_key'] = (
        df[start_cols]
        .apply(pd.to_datetime, errors='coerce')   # '' → NaT
        .min(axis=1)                              # 행별 가장 빠른 시각
    )

    sort_cols = ['_sort_key', 'interview_date']
    if 'wave' in df.columns:
        sort_cols.append('wave')
    sort_cols.append('code')

    df = (df
        .sort_values(sort_cols)
        .drop(columns='_sort_key')
        .reset_index(drop=True))
    # ────────────────────────────────────────────────────────

    # (F) 엑셀 출력
    df_to_excel(df, by_wave)

#6. 실행 스위치
if __name__ == '__main__':
    export_schedule_view()
