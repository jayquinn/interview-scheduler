# pages/6_Simulator.py
# "며칠이 필요?" — 기존 입력값을 그대로 이용해 최소 소요 날짜 추정
import streamlit as st
import pandas as pd
from io import BytesIO
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from datetime import datetime

import core
from solver.solver import solve_for_days, load_param_grid

st.set_page_config(
    page_title="운영일정추정",
    layout="wide"
)

if 'final_schedule' not in st.session_state:
    st.session_state['final_schedule'] = None
if 'last_solve_logs' not in st.session_state:
    st.session_state['last_solve_logs'] = ""
if 'solver_status' not in st.session_state:
    st.session_state['solver_status'] = "미실행"

def df_to_excel(df: pd.DataFrame, stream=None) -> None:
    """ DataFrame을 엑셀 파일 스트림으로 만듭니다. """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'
    df = df.copy()

    PALETTE = ['E3F2FD','FFF3E0','E8F5E9','FCE4EC','E1F5FE', 'F3E5F5','FFFDE7','E0F2F1','EFEBE9','ECEFF1']

    # 날짜별로 색을 다르게 하기 위해 날짜 목록 추출
    unique_dates = df['interview_date'].dt.date.unique()
    date_color_map = {date: PALETTE[i % len(PALETTE)] for i, date in enumerate(unique_dates)}

    df = df.astype(object).where(pd.notna(df), None)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill('solid', fgColor='D9D9D9')
    for cell in ws[1]:
        cell.fill = header_fill

    # 날짜 열 찾기
    date_col_idx = -1
    for j, col_name in enumerate(df.columns, 1):
        if col_name == 'interview_date':
            date_col_idx = j
            break

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        if date_col_idx != -1:
            date_val = row[date_col_idx - 1].value
            if date_val and hasattr(date_val, 'date'):
                row_color = date_color_map.get(date_val.date())
                if row_color:
                    row_fill = PatternFill('solid', fgColor=row_color)
                    for cell in row:
                        cell.fill = row_fill
    
    # 시간 형식 지정
    for j, col_name in enumerate(df.columns, 1):
         if 'start' in col_name or 'end' in col_name:
             for i in range(2, ws.max_row + 1):
                ws.cell(i, j).number_format = 'hh:mm'


    wb.save(stream or "recommended_schedule.xlsx")

def reset_run_state():
    st.session_state['final_schedule'] = None
    st.session_state['last_solve_logs'] = ""
    st.session_state['solver_status'] = "미실행"

with st.sidebar:
    st.markdown("## 파라미터")
    debug_mode = st.checkbox("🐞 디버그 모드", value=False)
    param_grid = load_param_grid()

    if not param_grid.empty:
        scenario_options = param_grid['scenario_id'].tolist()
        selected_scenario_id = st.selectbox(
            "시나리오 선택",
            options=scenario_options,
            index=0,
            help="파라미터 시나리오를 선택합니다."
        )
        params = param_grid[param_grid['scenario_id'] == selected_scenario_id].iloc[0].to_dict()
    else:
        st.warning("파라미터 파일을 찾을 수 없습니다.")
        params = {}

st.header("⑥ 운영일정 추정 (시뮬레이터)")
st.markdown("""
- `② 직무별 면접활동`, `③ 운영공간설정`, `④ 운영시간설정` 등 이전 단계에서 설정한 **템플릿**을 기반으로, 모든 지원자를 배정하는 데 필요한 **최소 운영일**을 계산합니다.
- 버튼을 누르면 시뮬레이션이 시작됩니다.
""")

# ─────────────────────────────────────────────
# 0. 세션 값 로드 & 기본 검증
# ─────────────────────────────────────────────
acts_df   = st.session_state.get("activities",           pd.DataFrame())
job_df    = st.session_state.get("job_acts_map",         pd.DataFrame())
room_plan = st.session_state.get("room_plan",            pd.DataFrame())
oper_df   = st.session_state.get("oper_window",          pd.DataFrame())
prec_df   = st.session_state.get("precedence",           pd.DataFrame())

# ① 활동
if acts_df.empty or not (acts_df["use"] == True).any():
    st.error("① Activities 페이지에서 'use=True' 활동을 하나 이상 지정하세요.")
    st.stop()
acts_df = acts_df.query("use == True").reset_index(drop=True)
act_list = acts_df["activity"].tolist()

# ② 직무 + 인원수
if job_df.empty or (job_df["count"].sum() == 0):
    st.error("② Job ↔ Activities 페이지에서 인원수를 1 명 이상 입력하세요.")
    st.stop()
if job_df["code"].duplicated().any():
    st.error("Job code 중복이 있습니다. 수정 후 다시 실행해 주세요.")
    st.stop()
# ── NEW: job_df <-> act_list 열 동기화 ────────────────────────
# (A) act_list 에 있지만 job_df 에 없는 열 → False 로 추가
for a in act_list:
    if a not in job_df.columns:
        job_df[a] = False

# (B) act_list 에서 빠진(=더 이상 use=True 가 아닌) 열은 제거
keep_cols = ["code", "count"] + act_list
job_df = job_df[[c for c in job_df.columns if c in keep_cols]]
# ────────────────────────────────────────────────────────────

# ③ Room Plan – "첫째 날" 값을 템플릿으로 사용
if room_plan.empty:
    st.error("③ Room Plan 페이지에서 방 정보를 먼저 입력하세요.")
    st.stop()
room_tpl = room_plan.iloc[0]   # 첫 행(날짜)에 입력된 방 세트
room_types = [
    re.sub(r"_count$", "", col)           # 뒤의 '_count' 만 제거
    for col in room_tpl.index
    if col.endswith("_count")
]

# ④ 운영시간 – "기본 시작/종료(템플릿)" 우선, 없으면 08:55~17:45
if not oper_df.empty and {"start_time","end_time"} <= set(oper_df.columns):
    common_start = str(oper_df.iloc[0]["start_time"])[:5]   # HH:MM
    common_end   = str(oper_df.iloc[0]["end_time"])[:5]
else:
    # ⑤ Operating Window 페이지에서 저장해 둔 '메모'(def_*_mem) 사용
    t_s = st.session_state.get("def_start_mem", datetime.time(8,55))
    t_e = st.session_state.get("def_end_mem",   datetime.time(17,45))
    common_start = t_s.strftime("%H:%M")
    common_end   = t_e.strftime("%H:%M")

st.success("✅ 입력 데이터 검증 통과 – Estimate 버튼을 누르세요!")

# ─────────────────────────────────────────────
# 1. Estimate!
# ─────────────────────────────────────────────
if st.button("운영일정추정 시작", type="primary", use_container_width=True, on_click=reset_run_state):
    with st.spinner("최적의 운영 일정을 계산 중입니다..."):
        cfg = core.build_config(st.session_state)
        
        status, final_wide, logs = solve_for_days(cfg, params, debug=debug_mode)
        
        st.session_state['last_solve_logs'] = logs
        st.session_state['solver_status'] = status

        if status == "OK" and final_wide is not None and not final_wide.empty:
            st.session_state['final_schedule'] = final_wide
            st.balloons()
        else:
            st.session_state['final_schedule'] = None

# ────────────────────────────────
# 3. 화면 표시
# ────────────────────────────────
st.markdown("---")
status = st.session_state.get('solver_status', '미실행')
st.info(f"Solver Status: `{status}`")

df = st.session_state.get('final_schedule')

if df is not None and not df.empty:
    total_days = df['interview_date'].nunique()
    st.success(f"**시뮬레이션 성공!** 모든 지원자를 배정하는 데 **총 {total_days}일**이 소요됩니다.")
    
    st.markdown("### 🗓️ 추천 스케줄 미리보기")
    st.dataframe(df)

    output = BytesIO()
    df_to_excel(df, stream=output)
    
    st.download_button(
        label="📥 추천 스케줄 다운로드 (Excel)",
        data=output.getvalue(),
        file_name="recommended_schedule.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if st.session_state.get('last_solve_logs'):
    with st.expander("상세 시뮬레이션 로그 보기"):
        st.code(st.session_state['last_solve_logs'])

