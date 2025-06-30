#!/usr/bin/env python3
"""
엑셀 결과 파일 분석 스크립트
사용자가 지적한 문제들을 확인합니다.
"""

import pandas as pd
import openpyxl

def analyze_excel_result():
    excel_file = "test_schedule_result.xlsx"
    
    try:
        # 엑셀 파일의 시트 목록 확인
        wb = openpyxl.load_workbook(excel_file)
        print('📋 엑셀 시트 목록:')
        for sheet in wb.sheetnames:
            print(f'  - {sheet}')
        print()

        # Schedule 시트 확인 
        try:
            schedule_df = pd.read_excel(excel_file, sheet_name='Schedule')
            print('📊 Schedule 시트:')
            print(f'  - 행 수: {len(schedule_df)}')
            print(f'  - 컬럼: {list(schedule_df.columns)}')
            print()
            print('첫 5행:')
            print(schedule_df.head())
            print()
            
            # group_size 컬럼 확인
            if 'group_size' in schedule_df.columns:
                print('🔍 group_size 컬럼 분석:')
                print(f'  - 유니크 값: {schedule_df["group_size"].unique()}')
                print(f'  - 값 분포:')
                print(schedule_df["group_size"].value_counts())
                print()
                
                # 날짜/시간/공간별 group_size 분석
                print('📊 날짜/시간/공간별 group_size 분석:')
                group_analysis = schedule_df.groupby(['interview_date', 'start_time', 'room_name']).agg({
                    'group_size': 'first',
                    'applicant_id': 'count'
                }).rename(columns={'applicant_id': 'actual_count'})
                
                print(group_analysis.head(10))
                
                # group_size와 실제 카운트 불일치 확인
                mismatches = group_analysis[group_analysis['group_size'] != group_analysis['actual_count']]
                if len(mismatches) > 0:
                    print(f"\n❌ group_size 불일치 발견: {len(mismatches)}건")
                    print(mismatches)
                else:
                    print("\n✅ group_size가 실제 카운트와 일치")
                    
            else:
                print('❌ group_size 컬럼이 없습니다.')
                
        except Exception as e:
            print(f'❌ Schedule 시트 읽기 실패: {e}')

        print("\n" + "="*60)

        # TS_ 시트들 확인
        ts_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith('TS_')]
        print(f'📋 TS_ 시트 목록 ({len(ts_sheets)}개):')
        for sheet in ts_sheets:
            print(f'  - {sheet}')

        # 각 TS_ 시트에서 중복 ID 확인
        all_applicants = {}  # {applicant_id: [sheet_names]}
        
        for sheet_name in ts_sheets:
            try:
                ts_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                print(f'\n📊 {sheet_name} 시트:')
                print(f'  - 행 수: {len(ts_df)}')
                print(f'  - 컬럼: {list(ts_df.columns)}')
                
                # ID 컬럼 찾기
                id_col = None
                for col in ['applicant_id', 'id', 'ID']:
                    if col in ts_df.columns:
                        id_col = col
                        break
                
                if id_col:
                    unique_ids = ts_df[id_col].unique()
                    print(f'  - 유니크 ID 수: {len(unique_ids)}')
                    
                    # 전체 applicant 딕셔너리에 추가
                    for app_id in unique_ids:
                        if pd.notna(app_id):  # NaN 제외
                            if app_id not in all_applicants:
                                all_applicants[app_id] = []
                            all_applicants[app_id].append(sheet_name)
                    
                    # 시트 내 중복 확인
                    duplicates = ts_df[id_col].value_counts()
                    duplicates = duplicates[duplicates > 1]
                    if len(duplicates) > 0:
                        print(f'  ❌ 시트 내 중복 ID: {len(duplicates)}개')
                        for id_val, count in duplicates.items():
                            print(f'    - {id_val}: {count}번')
                    else:
                        print(f'  ✅ 시트 내 중복 ID 없음')
                else:
                    print(f'  ❌ ID 컬럼을 찾을 수 없음')
                    
            except Exception as e:
                print(f'❌ {sheet_name} 시트 읽기 실패: {e}')

        # 시트 간 중복 ID 확인 (중요!)
        print(f'\n🔍 시트 간 중복 ID 분석:')
        cross_duplicates = {app_id: sheets for app_id, sheets in all_applicants.items() if len(sheets) > 1}
        
        if cross_duplicates:
            print(f'❌ 시트 간 중복 ID 발견: {len(cross_duplicates)}개')
            for app_id, sheets in cross_duplicates.items():
                print(f'  - {app_id}: {", ".join(sheets)}')
                
            # 토론면접실과 발표면접실 중복 확인 (사용자 지적 사항)
            discussion_presentation_duplicates = []
            for app_id, sheets in cross_duplicates.items():
                has_discussion = any('토론면접실' in sheet for sheet in sheets)
                has_presentation = any('발표면접실' in sheet for sheet in sheets)
                if has_discussion and has_presentation:
                    discussion_presentation_duplicates.append((app_id, sheets))
            
            if discussion_presentation_duplicates:
                print(f'\n⚠️ 토론면접실과 발표면접실 중복 (사용자 지적 사항): {len(discussion_presentation_duplicates)}개')
                for app_id, sheets in discussion_presentation_duplicates:
                    print(f'  - {app_id}: {", ".join(sheets)}')
            else:
                print(f'\n✅ 토론면접실과 발표면접실 중복 없음')
        else:
            print('✅ 시트 간 중복 ID 없음')

    except Exception as e:
        print(f'❌ 엑셀 파일 분석 실패: {e}')
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel_result() 