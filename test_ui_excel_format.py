#!/usr/bin/env python3
"""
UI 엑셀 형식으로 결과 저장 테스트
이전 테스트에서 성공한 CP-SAT 결과를 app.py의 df_to_excel 함수를 사용해서 저장
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time
from io import BytesIO
import os
import sys

# 필요한 모듈 임포트
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    from solver.api import solve_for_days_v2
    from app import df_to_excel
    print("✅ 모듈 임포트 성공")
except ImportError as e:
    print(f"❌ 모듈 임포트 실패: {e}")
    sys.exit(1)

def create_full_ui_defaults():
    """app.py UI 전체 디폴트 설정 생성"""
    
    # 1. 활동 설정
    activities = pd.DataFrame([
        {
            "activity": "토론면접",
            "mode": "batched",
            "duration_min": 30,
            "room_type": "토론면접실",
            "min_cap": 6,
            "max_cap": 6,
            "use": True
        },
        {
            "activity": "발표준비",
            "mode": "parallel",
            "duration_min": 5,
            "room_type": "발표준비실",
            "min_cap": 2,
            "max_cap": 2,
            "use": True
        },
        {
            "activity": "발표면접",
            "mode": "individual",
            "duration_min": 15,
            "room_type": "발표면접실",
            "min_cap": 1,
            "max_cap": 1,
            "use": True
        }
    ])
    
    # 2. 방 설정
    room_plan = pd.DataFrame([{
        "토론면접실_count": 2,
        "토론면접실_cap": 6,
        "발표준비실_count": 1,
        "발표준비실_cap": 2,
        "발표면접실_count": 2,
        "발표면접실_cap": 1
    }])
    
    # 3. 직무별 활동 매핑
    job_acts_map = pd.DataFrame([
        {"job_code": "JOB01", "activity": "토론면접", "order": 1},
        {"job_code": "JOB01", "activity": "발표준비", "order": 2},
        {"job_code": "JOB01", "activity": "발표면접", "order": 3},
        {"job_code": "JOB02", "activity": "토론면접", "order": 1},
        {"job_code": "JOB02", "activity": "발표준비", "order": 2},
        {"job_code": "JOB02", "activity": "발표면접", "order": 3},
        {"job_code": "JOB03", "activity": "토론면접", "order": 1},
        {"job_code": "JOB03", "activity": "발표준비", "order": 2},
        {"job_code": "JOB03", "activity": "발표면접", "order": 3},
        {"job_code": "JOB04", "activity": "토론면접", "order": 1},
        {"job_code": "JOB04", "activity": "발표준비", "order": 2},
        {"job_code": "JOB04", "activity": "발표면접", "order": 3},
        {"job_code": "JOB05", "activity": "토론면접", "order": 1},
        {"job_code": "JOB05", "activity": "발표준비", "order": 2},
        {"job_code": "JOB05", "activity": "발표면접", "order": 3},
        {"job_code": "JOB06", "activity": "토론면접", "order": 1},
        {"job_code": "JOB06", "activity": "발표준비", "order": 2},
        {"job_code": "JOB06", "activity": "발표면접", "order": 3},
        {"job_code": "JOB07", "activity": "토론면접", "order": 1},
        {"job_code": "JOB07", "activity": "발표준비", "order": 2},
        {"job_code": "JOB07", "activity": "발표면접", "order": 3},
        {"job_code": "JOB08", "activity": "토론면접", "order": 1},
        {"job_code": "JOB08", "activity": "발표준비", "order": 2},
        {"job_code": "JOB08", "activity": "발표면접", "order": 3},
        {"job_code": "JOB09", "activity": "토론면접", "order": 1},
        {"job_code": "JOB09", "activity": "발표준비", "order": 2},
        {"job_code": "JOB09", "activity": "발표면접", "order": 3},
        {"job_code": "JOB10", "activity": "토론면접", "order": 1},
        {"job_code": "JOB10", "activity": "발표준비", "order": 2},
        {"job_code": "JOB10", "activity": "발표면접", "order": 3},
        {"job_code": "JOB11", "activity": "토론면접", "order": 1},
        {"job_code": "JOB11", "activity": "발표준비", "order": 2},
        {"job_code": "JOB11", "activity": "발표면접", "order": 3}
    ])
    
    # 4. 시간 설정
    time_settings = {
        "start_time": time(9, 0),  # 09:00
        "end_time": time(17, 30),  # 17:30
        "break_duration": 0,
        "lunch_start": time(12, 0),
        "lunch_end": time(13, 0)
    }
    
    # 5. 선후행 설정
    precedence = pd.DataFrame([
        {"job_code": "JOB01", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB02", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB03", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB04", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB05", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB06", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB07", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB08", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB09", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB10", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0},
        {"job_code": "JOB11", "predecessor": "발표준비", "successor": "발표면접", "gap_min": 0}
    ])
    
    # 6. 멀티 날짜 계획 (4일간, 총 137명)
    base_date = datetime(2024, 7, 1)
    date_plans = {
        "day1": {
            "enabled": True,
            "date": base_date,
            "jobs": [
                {"code": "JOB01", "count": 23},
                {"code": "JOB02", "count": 23}
            ]
        },
        "day2": {
            "enabled": True,
            "date": base_date + timedelta(days=1),
            "jobs": [
                {"code": "JOB03", "count": 20},
                {"code": "JOB04", "count": 20}
            ]
        },
        "day3": {
            "enabled": True,
            "date": base_date + timedelta(days=2),
            "jobs": [
                {"code": "JOB05", "count": 12},
                {"code": "JOB06", "count": 15},
                {"code": "JOB07", "count": 6}
            ]
        },
        "day4": {
            "enabled": True,
            "date": base_date + timedelta(days=3),
            "jobs": [
                {"code": "JOB08", "count": 6},
                {"code": "JOB09", "count": 6},
                {"code": "JOB10", "count": 3},
                {"code": "JOB11", "count": 3}
            ]
        }
    }
    
    return {
        "activities": activities,
        "room_plan": room_plan,
        "job_acts_map": job_acts_map,
        "time_settings": time_settings,
        "precedence": precedence,
        "multidate_plans": date_plans
    }

def analyze_stay_time(df):
    """체류시간 분석"""
    if df is None or df.empty:
        return None
    
    stay_times = []
    
    # 지원자별 체류시간 계산
    for applicant_id in df['applicant_id'].unique():
        applicant_schedule = df[df['applicant_id'] == applicant_id]
        
        if applicant_schedule.empty:
            continue
        
        # 날짜별로 처리
        for date in applicant_schedule['interview_date'].unique():
            date_schedule = applicant_schedule[applicant_schedule['interview_date'] == date]
            
            if len(date_schedule) == 0:
                continue
            
            # 시작과 끝 시간 계산
            start_time = date_schedule['start_time'].min()
            end_time = date_schedule['end_time'].max()
            
            # timedelta를 시간으로 변환
            if hasattr(start_time, 'total_seconds'):
                stay_duration = (end_time - start_time).total_seconds() / 3600
            else:
                # datetime 객체인 경우
                stay_duration = (end_time - start_time).total_seconds() / 3600
            
            stay_times.append({
                'applicant_id': applicant_id,
                'date': date,
                'stay_time_hours': stay_duration
            })
    
    return pd.DataFrame(stay_times)

def test_ui_excel_format():
    """UI 엑셀 형식으로 결과 저장 테스트"""
    print("🧪 UI 엑셀 형식 결과 저장 테스트 시작")
    print("=" * 80)
    
    # 1. 전체 UI 디폴트 설정 생성
    cfg_ui = create_full_ui_defaults()
    total_applicants = sum(
        sum(job["count"] for job in day["jobs"]) 
        for day in cfg_ui["multidate_plans"].values()
    )
    
    print(f"📊 테스트 설정:")
    print(f"  - 총 {len(cfg_ui['multidate_plans'])}일간 멀티데이트")
    print(f"  - 총 {total_applicants}명 지원자")
    print(f"  - 활동: {len(cfg_ui['activities'])}개")
    print(f"  - 방 타입: {cfg_ui['room_plan'].shape[1]//2}개")
    
    # 2. CP-SAT 스케줄링 실행
    print(f"\n🚀 CP-SAT 스케줄링 실행...")
    status, result_df, logs, limit = solve_for_days_v2(cfg_ui, debug=False)
    
    print(f"스케줄링 결과: {status}")
    if status == "SUCCESS":
        print(f"✅ 성공: {len(result_df)}개 스케줄 항목")
        print(f"   지원자 수: {result_df['applicant_id'].nunique()}명")
        print(f"   성공률: {result_df['applicant_id'].nunique()/total_applicants*100:.1f}%")
        
        # 3. 체류시간 분석
        print(f"\n📊 체류시간 분석...")
        stay_df = analyze_stay_time(result_df)
        if stay_df is not None:
            avg_stay = stay_df['stay_time_hours'].mean()
            max_stay = stay_df['stay_time_hours'].max()
            print(f"   평균 체류시간: {avg_stay:.1f}시간")
            print(f"   최대 체류시간: {max_stay:.1f}시간")
        
        # 4. UI 형식 엑셀 파일 생성
        print(f"\n📄 UI 형식 엑셀 파일 생성...")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ui_format_cpsat_result_{timestamp}.xlsx"
        
        try:
            # app.py의 df_to_excel 함수 사용
            df_to_excel(result_df, filename)
            print(f"✅ UI 형식 엑셀 파일 생성 성공: {filename}")
            
            # 파일 크기 확인
            if os.path.exists(filename):
                size = os.path.getsize(filename)
                print(f"   파일 크기: {size:,} bytes")
                
                # 엑셀 파일 시트 확인
                import openpyxl
                wb = openpyxl.load_workbook(filename)
                print(f"   시트 목록: {wb.sheetnames}")
                
                # 각 시트의 행 수 확인
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"   - {sheet_name}: {ws.max_row}행 × {ws.max_column}열")
                    
                    # 첫행 고정 확인
                    if ws.freeze_panes:
                        print(f"     (첫행 고정 적용)")
                
                wb.close()
            
            return True, filename
            
        except Exception as e:
            print(f"❌ 엑셀 파일 생성 실패: {e}")
            import traceback
            traceback.print_exc()
            return False, None
    
    else:
        print(f"❌ 스케줄링 실패")
        print(f"로그: {logs}")
        return False, None

def main():
    """메인 실행"""
    print("=" * 80)
    print("🎯 UI 엑셀 형식으로 CP-SAT 결과 저장")
    print("📋 app.py의 df_to_excel 함수를 사용하여 UI 형식 엑셀 파일 생성")
    print("=" * 80)
    
    success, filename = test_ui_excel_format()
    
    if success:
        print(f"\n🎉 UI 형식 엑셀 파일 생성 성공!")
        print(f"💾 저장된 파일: {filename}")
        print(f"   - Schedule 시트: 기본 스케줄 정보 (조 정보, 색상 코딩)")
        print(f"   - TS_MMDD 시트들: 날짜별 타임슬롯 매트릭스")
    else:
        print(f"\n❌ UI 형식 엑셀 파일 생성 실패")
    
    print("=" * 80)

if __name__ == "__main__":
    main() 