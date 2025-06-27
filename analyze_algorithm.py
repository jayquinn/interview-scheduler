import pandas as pd
import openpyxl

def analyze_algorithm_improvement():
    print("=== 알고리즘 개선 결과 분석 ===")
    
    try:
        # Excel 파일 읽기
        wb = openpyxl.load_workbook("test_algorithm_improved.xlsx")
        print(f"시트 목록: {wb.sheetnames}")
        
        # 첫행 고정 확인
        print(f"\n=== 첫행 고정 확인 ===")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            freeze_panes = ws.freeze_panes
            if freeze_panes:
                print(f"  ✅ {sheet_name}: 첫행 고정 적용됨 ({freeze_panes})")
            else:
                print(f"  ❌ {sheet_name}: 첫행 고정 없음")
        
        # Schedule 시트 데이터 분석
        print(f"\n=== Schedule 시트 분석 ===")
        schedule_df = pd.read_excel("test_algorithm_improved.xlsx", sheet_name="Schedule")
        print(f"총 {len(schedule_df)}개 스케줄")
        
        # 지원자별 선후행 제약 검증
        print(f"\n=== 🔧 알고리즘 개선 효과 분석 ===")
        violations = []
        successes = []
        
        for applicant_id in sorted(schedule_df['applicant_id'].unique()):
            applicant_data = schedule_df[schedule_df['applicant_id'] == applicant_id].copy()
            
            # 발표준비와 발표면접 찾기
            prep_data = applicant_data[applicant_data['activity_name'] == '발표준비']
            present_data = applicant_data[applicant_data['activity_name'] == '발표면접']
            
            if not prep_data.empty and not present_data.empty:
                prep_start = prep_data.iloc[0]['start_time']
                prep_end = prep_data.iloc[0]['end_time']
                present_start = present_data.iloc[0]['start_time']
                present_end = present_data.iloc[0]['end_time']
                prep_room = prep_data.iloc[0]['room_name']
                present_room = present_data.iloc[0]['room_name']
                
                print(f"\n{applicant_id}:")
                print(f"  발표준비: {prep_start} ~ {prep_end} ({prep_room})")
                print(f"  발표면접: {present_start} ~ {present_end} ({present_room})")
                
                # 시간을 분으로 변환
                if hasattr(prep_end, 'hour'):
                    prep_end_min = prep_end.hour * 60 + prep_end.minute
                elif hasattr(prep_end, 'total_seconds'):
                    prep_end_min = prep_end.total_seconds() / 60
                else:
                    prep_end_min = 0
                
                if hasattr(present_start, 'hour'):
                    present_start_min = present_start.hour * 60 + present_start.minute
                elif hasattr(present_start, 'total_seconds'):
                    present_start_min = present_start.total_seconds() / 60
                else:
                    present_start_min = 0
                
                gap_min = present_start_min - prep_end_min
                print(f"  간격: {gap_min}분")
                
                # 디폴트 규칙: gap_min=5, adjacent=True 
                if abs(gap_min - 5) > 0.1:  # 소수점 오차 허용
                    violations.append(f"{applicant_id}: 연속배치 위반 (실제 간격: {gap_min}분, 예상: 5분)")
                    print(f"  ❌ 연속배치 위반")
                else:
                    successes.append(applicant_id)
                    print(f"  ✅ 연속배치 완벽!")
        
        # 성능 비교
        total = len(successes) + len(violations)
        success_rate = len(successes) / total * 100 if total > 0 else 0
        
        print(f"\n=== 🎯 성능 비교 ===")
        print(f"기존 알고리즘: 33.3% 성공률 (6명 중 2명)")
        print(f"개선 알고리즘: {success_rate:.1f}% 성공률 ({len(successes)}명 중 {total}명)")
        
        if success_rate > 33.3:
            improvement = success_rate - 33.3
            print(f"🚀 개선 효과: +{improvement:.1f}% 향상!")
        else:
            print(f"⚠️ 개선 효과 없음")
        
        if violations:
            print(f"\n❌ 여전히 위반된 케이스:")
            for violation in violations:
                print(f"  • {violation}")
                
            print(f"\n🔍 위반 원인 분석:")
            # 위반 케이스의 패턴 분석
            violation_patterns = {}
            for applicant_id in [v.split(":")[0] for v in violations]:
                applicant_data = schedule_df[schedule_df['applicant_id'] == applicant_id]
                토론_data = applicant_data[applicant_data['activity_name'] == '토론면접']
                if not 토론_data.empty:
                    토론_end = 토론_data.iloc[0]['end_time']
                    if hasattr(토론_end, 'hour'):
                        토론_end_min = 토론_end.hour * 60 + 토론_end.minute
                    else:
                        토론_end_min = 토론_end.total_seconds() / 60
                    
                    prep_data = applicant_data[applicant_data['activity_name'] == '발표준비']
                    if not prep_data.empty:
                        prep_start = prep_data.iloc[0]['start_time']
                        if hasattr(prep_start, 'hour'):
                            prep_start_min = prep_start.hour * 60 + prep_start.minute
                        else:
                            prep_start_min = prep_start.total_seconds() / 60
                        
                        토론_prep_gap = prep_start_min - 토론_end_min
                        print(f"  • {applicant_id}: 토론면접 종료 후 {토론_prep_gap}분 → 발표준비")
        else:
            print(f"\n🎉 모든 선후행 제약 완벽 준수!")
        
        # 활동별 통계
        print(f"\n=== 활동별 통계 ===")
        for activity in schedule_df['activity_name'].unique():
            activity_data = schedule_df[schedule_df['activity_name'] == activity]
            print(f"{activity}: {len(activity_data)}명")
            if not activity_data.empty:
                start_times = sorted(activity_data['start_time'].unique())
                print(f"  시작 시간: {start_times}")
                rooms = sorted(activity_data['room_name'].unique())
                print(f"  사용 방: {rooms}")
        
        # TS_ 시트 확인
        ts_sheets = [name for name in wb.sheetnames if name.startswith('TS_')]
        if ts_sheets:
            print(f"\n=== TS_ 시트 확인 ===")
            for ts_name in ts_sheets:
                ts_ws = wb[ts_name]
                print(f"{ts_name}: {ts_ws.max_row}행 x {ts_ws.max_column}열")
        
    except Exception as e:
        print(f"❌ 분석 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_algorithm_improvement() 