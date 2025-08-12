"""
🔧 Simple Core Module
복잡성을 줄인 간결한 핵심 모듈

기존 core.py의 복잡한 구조를 단순화하여
SimpleInterviewScheduler와 연동하는 간결한 인터페이스 제공
"""

import pandas as pd
from datetime import datetime, time, timedelta
from typing import Dict, List, Tuple, Any, Optional
from io import BytesIO

from simple_scheduler import (
    SimpleInterviewScheduler, 
    Activity, Room, Applicant, PrecedenceRule,
    convert_to_dataframe, validate_schedule
)

# =============================================================================
# 🔧 설정 변환 함수들
# =============================================================================

def build_config(state: dict) -> dict:
    """
    Streamlit 세션 상태를 스케줄러 설정으로 변환
    
    기존 build_config()의 복잡한 로직을 단순화
    """
    empty = lambda: pd.DataFrame()
    
    cfg = {
        "activities": state.get("activities", empty()),
        "job_acts_map": state.get("job_acts_map", empty()),
        "room_plan": state.get("room_plan", empty()),
        "oper_window": state.get("oper_window", empty()),
        "precedence": state.get("precedence", empty()),
        "interview_dates": state.get("interview_dates", []),
        "multidate_plans": state.get("multidate_plans", {}),
    }
    
    return cfg

def convert_ui_to_scheduler_data(cfg: dict) -> Tuple[List[Activity], List[Room], List[Applicant], List[PrecedenceRule], Tuple[time, time]]:
    """
    UI 설정을 스케줄러 데이터로 변환
    """
    
    # 1. 활동 데이터 변환
    activities = []
    if not cfg["activities"].empty:
        for _, row in cfg["activities"].iterrows():
            if row.get("use", True):  # 사용 여부 확인
                activity = Activity(
                    name=row["activity"],
                    mode=row.get("mode", "individual"),
                    duration_min=row.get("duration_min", 10),
                    room_type=row.get("room_type", ""),
                    min_cap=row.get("min_cap", 1),
                    max_cap=row.get("max_cap", 1)
                )
                activities.append(activity)
    
    # 2. 방 데이터 변환
    rooms = []
    room_plan = cfg.get("room_plan", pd.DataFrame())
    if not room_plan.empty:
        for _, row in room_plan.iterrows():
            # 방 타입별로 방 생성
            for col in row.index:
                if col.endswith("_count"):
                    room_type = col.replace("_count", "")
                    count = int(row[col])
                    cap_col = f"{room_type}_cap"
                    capacity = int(row.get(cap_col, 1))
                    
                    for i in range(count):
                        room_name = f"{room_type}{i+1}" if count > 1 else room_type
                        room = Room(
                            name=room_name,
                            room_type=room_type,
                            capacity=capacity,
                            date=datetime.now()  # 날짜는 나중에 설정
                        )
                        rooms.append(room)
    
    # 3. 지원자 데이터 변환
    applicants = []
    job_acts_map = cfg.get("job_acts_map", pd.DataFrame())
    if not job_acts_map.empty:
        for _, row in job_acts_map.iterrows():
            job_code = row["code"]
            count = int(row.get("count", 1))
            
            # 해당 직무의 활동 목록
            required_activities = []
            for col in row.index:
                if col not in ["code", "count"] and row.get(col, False):
                    required_activities.append(col)
            
            # 지원자별로 생성
            for i in range(count):
                applicant_id = f"{job_code}_{i+1:03d}"
                applicant = Applicant(
                    id=applicant_id,
                    job_code=job_code,
                    required_activities=required_activities,
                    date=datetime.now()  # 날짜는 나중에 설정
                )
                applicants.append(applicant)
    
    # 4. 선후행 제약 변환
    precedence_rules = []
    precedence_df = cfg.get("precedence", pd.DataFrame())
    if not precedence_df.empty:
        for _, row in precedence_df.iterrows():
            rule = PrecedenceRule(
                predecessor=row["predecessor"],
                successor=row["successor"],
                gap_min=int(row.get("gap_min", 0)),
                is_adjacent=bool(row.get("adjacent", False))
            )
            precedence_rules.append(rule)
    
    # 5. 운영 시간 변환
    oper_window = cfg.get("oper_window", pd.DataFrame())
    if not oper_window.empty:
        start_time_str = oper_window.iloc[0].get("start_time", "09:00")
        end_time_str = oper_window.iloc[0].get("end_time", "17:30")
        
        start_time = time.fromisoformat(start_time_str)
        end_time = time.fromisoformat(end_time_str)
    else:
        start_time = time(9, 0)
        end_time = time(17, 30)
    
    return activities, rooms, applicants, precedence_rules, (start_time, end_time)

# =============================================================================
# 🚀 스케줄러 실행 함수
# =============================================================================

def run_simple_scheduler(cfg: dict, params: dict = None) -> Tuple[str, pd.DataFrame, str]:
    """
    단순화된 스케줄러 실행
    
    기존 run_solver()의 복잡한 로직을 단순화
    """
    
    try:
        # 1. 설정 검증
        validation_result = validate_config(cfg)
        if not validation_result[0]:
            return "VALIDATION_ERROR", pd.DataFrame(), validation_result[1]
        
        # 2. UI 데이터를 스케줄러 데이터로 변환
        activities, rooms, applicants, precedence_rules, operating_hours = convert_ui_to_scheduler_data(cfg)
        
        if not activities:
            return "ERROR", pd.DataFrame(), "활동이 정의되지 않았습니다"
        
        if not applicants:
            return "ERROR", pd.DataFrame(), "지원자가 정의되지 않았습니다"
        
        # 3. 스케줄러 실행
        scheduler = SimpleInterviewScheduler()
        status, results, logs = scheduler.schedule(
            applicants=applicants,
            activities=activities,
            rooms=rooms,
            precedence_rules=precedence_rules,
            operating_hours=operating_hours,
            params=params or {}
        )
        
        if status == "SUCCESS" and results:
            # 4. 결과를 DataFrame으로 변환
            df = convert_to_dataframe(results)
            
            # 5. 결과 검증
            is_valid, errors = validate_schedule(results)
            if not is_valid:
                logs += f"\n경고: {', '.join(errors)}"
            
            return "SUCCESS", df, logs
        else:
            return status, pd.DataFrame(), logs
            
    except Exception as e:
        error_msg = f"스케줄러 실행 중 오류: {str(e)}"
        return "ERROR", pd.DataFrame(), error_msg

# =============================================================================
# 📊 Excel 출력 함수 (단순화)
# =============================================================================

def to_excel_simple(df: pd.DataFrame) -> bytes:
    """
    단순화된 Excel 출력
    
    기존 to_excel()의 복잡한 시각화를 단순화
    """
    
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Schedule'
        
        # 데이터프레임을 워크시트에 추가
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 스타일링
        header_fill = PatternFill('solid', fgColor='D9D9D9')
        for cell in ws[1]:
            cell.fill = header_fill
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        raise Exception(f"Excel 생성 중 오류: {str(e)}")

# =============================================================================
# 🔍 설정 검증 함수
# =============================================================================

def validate_config(cfg: dict) -> Tuple[bool, str]:
    """
    설정 검증
    
    기존 복잡한 검증 로직을 단순화
    """
    
    errors = []
    
    # 1. 활동 검증
    activities = cfg.get("activities", pd.DataFrame())
    if activities.empty:
        errors.append("활동이 정의되지 않았습니다")
    elif "use" in activities.columns and not (activities["use"] == True).any():
        errors.append("사용 가능한 활동이 없습니다")
    
    # 2. 지원자 검증
    job_acts_map = cfg.get("job_acts_map", pd.DataFrame())
    if job_acts_map.empty:
        errors.append("지원자 정보가 정의되지 않았습니다")
    elif job_acts_map.get("count", 0).sum() == 0:
        errors.append("지원자 수가 0명입니다")
    
    # 3. 방 검증
    room_plan = cfg.get("room_plan", pd.DataFrame())
    if room_plan.empty:
        errors.append("방 정보가 정의되지 않았습니다")
    
    # 4. 운영 시간 검증
    oper_window = cfg.get("oper_window", pd.DataFrame())
    if oper_window.empty:
        errors.append("운영 시간이 정의되지 않았습니다")
    
    if errors:
        return False, "; ".join(errors)
    
    return True, "설정이 유효합니다"

# =============================================================================
# 📈 간단한 통계 함수
# =============================================================================

def calculate_simple_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """
    간단한 통계 계산
    
    복잡한 체류시간 분석을 단순화
    """
    
    if df.empty:
        return {}
    
    stats = {
        "total_applicants": df['applicant_id'].nunique(),
        "total_activities": len(df),
        "unique_activities": df['activity_name'].nunique(),
        "unique_rooms": df['room_name'].nunique(),
        "date_range": f"{df['interview_date'].min()} ~ {df['interview_date'].max()}"
    }
    
    # 체류시간 계산 (간단한 버전)
    try:
        applicant_times = {}
        for _, row in df.iterrows():
            applicant_id = row['applicant_id']
            start_time = row['start_time']
            end_time = row['end_time']
            
            if applicant_id not in applicant_times:
                applicant_times[applicant_id] = {'start': start_time, 'end': end_time}
            else:
                applicant_times[applicant_id]['start'] = min(
                    applicant_times[applicant_id]['start'], start_time
                )
                applicant_times[applicant_id]['end'] = max(
                    applicant_times[applicant_id]['end'], end_time
                )
        
        stay_times = []
        for times in applicant_times.values():
            duration = times['end'] - times['start']
            stay_hours = duration.total_seconds() / 3600
            stay_times.append(stay_hours)
        
        if stay_times:
            stats.update({
                "avg_stay_hours": round(sum(stay_times) / len(stay_times), 2),
                "max_stay_hours": round(max(stay_times), 2),
                "min_stay_hours": round(min(stay_times), 2)
            })
    
    except Exception:
        pass  # 체류시간 계산 실패 시 무시
    
    return stats 