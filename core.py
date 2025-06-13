# core.py
"""
면접 스케줄링 시스템의 핵심 유틸리티 모듈
"""

from io import BytesIO
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# OR-Tools 래퍼
from solver.solver import solve, load_param_grid


def build_config(state: dict) -> dict:
    """
    Streamlit 세션 상태를 Solver 설정 딕셔너리로 변환한다.
    
    Args:
        state: Streamlit 세션 상태 딕셔너리
        
    Returns:
        dict: Solver 설정 딕셔너리
    """
    empty = lambda: pd.DataFrame()  # 빈 DataFrame 생성 헬퍼

    cfg = {
        "activities": state.get("activities", empty()),
        "job_acts_map": state.get("job_acts_map", empty()),
        "room_plan": state.get("room_plan", empty()),
        "oper_window": state.get("oper_window", empty()),
        "precedence": state.get("precedence", empty()),
        "candidates": state.get("candidates", empty()),
        "candidates_exp": state.get("candidates_exp", empty()),
        # 집단면접 관련 설정 추가
        "job_similarity_groups": state.get("job_similarity_groups", []),
        "prefer_job_separation": state.get("prefer_job_separation", True),
        "use_similarity_groups": state.get("use_similarity_groups", False),
    }
    cfg["group_meta"] = cfg["activities"].copy()
    return cfg


def run_solver(cfg: dict, params: dict | None = None, *, debug=False):
    """
    Solver 실행 래퍼 함수
    
    Args:
        cfg: Solver 설정 딕셔너리
        params: 솔버 파라미터
        debug: 디버그 모드 여부
        
    Returns:
        tuple: (status, wide_df, logs)
    """
    status, wide, logs = solve(cfg, params=params, debug=debug)
    return status, wide, logs

def df_to_excel_internal(df: pd.DataFrame, stream=None) -> None:
    """
    DataFrame을 Excel 파일로 변환하는 내부 함수
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'
    df = df.copy()
    
    PALETTE = ['E3F2FD', 'FFF3E0', 'E8F5E9', 'FCE4EC', 'E1F5FE', 'F3E5F5', 'FFFDE7', 'E0F2F1', 'EFEBE9', 'ECEFF1']
    
    # 날짜별로 색상 지정
    if 'interview_date' in df.columns and not df.empty:
        unique_dates = pd.to_datetime(df['interview_date']).dt.date.unique()
        date_color_map = {date: PALETTE[i % len(PALETTE)] for i, date in enumerate(unique_dates)}
    else:
        date_color_map = {}
    
    df = df.astype(object).where(pd.notna(df), None)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    header_fill = PatternFill('solid', fgColor='D9D9D9')
    for cell in ws[1]:
        cell.fill = header_fill
    
    # 날짜 열 찾기
    date_col_idx = -1
    for j, col_name in enumerate(df.columns, 1):
        if col_name == 'interview_date':
            date_col_idx = j
            break
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        if date_col_idx != -1 and date_color_map:
            date_val = row[date_col_idx - 1].value
            if date_val and hasattr(date_val, 'date'):
                row_color = date_color_map.get(date_val.date())
                if row_color:
                    row_fill = PatternFill('solid', fgColor=row_color)
                    for cell in row:
                        cell.fill = row_fill
    
    # 시간 형식 지정
    for j, col_name in enumerate(df.columns, 1):
        if 'start' in col_name or 'end' in col_name:
            for i in range(2, ws.max_row + 1):
                ws.cell(i, j).number_format = 'hh:mm'
    
    wb.save(stream or "schedule.xlsx")


def to_excel(wide_df: pd.DataFrame) -> bytes:
    """
    DataFrame을 Excel 파일(bytes)로 변환한다.
    
    Args:
        wide_df: 스케줄 데이터프레임
        
    Returns:
        bytes: Excel 파일 바이너리 데이터
    """
    buf = BytesIO()
    df_to_excel_internal(wide_df, buf)
    buf.seek(0)
    return buf.getvalue()
