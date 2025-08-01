# pages/1_면접운영스케줄링.py
import streamlit as st
import pandas as pd
import re
from datetime import time, datetime, timedelta, date
from st_aggrid import (
    AgGrid,
    GridOptionsBuilder,
    GridUpdateMode,
    DataReturnMode,
)
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import core
from solver.solver import solve_for_days
from solver.api import solve_for_days_v2, solve_for_days_two_phase, get_scheduler_comparison
from solver.types import ProgressInfo
from test_internal_analysis import run_multi_date_scheduling

# 진행 상황 콜백 함수
def progress_callback(info: ProgressInfo):
    """실시간 진행 상황 업데이트"""
    st.session_state['progress_info'] = info
    st.session_state['current_stage'] = info.stage
    st.session_state['progress_value'] = info.progress
    st.session_state['stage_details'] = info.details

st.set_page_config(
    page_title="면접운영스케줄링",
    layout="wide"
)

# 사이드바에서 app 페이지 숨기기
st.markdown("""
<style>
    /* 사이드바에서 첫 번째 항목(app) 숨기기 */
    .css-1d391kg .css-1rs6os .css-17eq0hr:first-child {
        display: none !important;
    }
    
    /* 다른 방법으로 app 항목 숨기기 */
    [data-testid="stSidebar"] [data-testid="stSidebarNav"] ul li:first-child {
        display: none !important;
    }
    
    /* 또 다른 방법 */
    .css-1rs6os .css-17eq0hr:first-child {
        display: none !important;
    }
    
    /* 사이드바 네비게이션에서 첫 번째 링크 숨기기 */
    .css-1rs6os a[href="/"]:first-child {
        display: none !important;
    }
    
    /* 최신 Streamlit 버전용 */
    [data-testid="stSidebarNav"] > ul > li:first-child {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

# 페이지 상단 앵커 포인트
st.markdown('<div id="top"></div>', unsafe_allow_html=True)

st.title("🎯 면접운영스케줄링")
st.markdown("""
**올인원 면접 스케줄링 솔루션**  
이 페이지에서 면접 스케줄링에 필요한 모든 설정과 일정 생성을 한 번에 진행할 수 있습니다.
운영일정 추정을 먼저 확인한 후, 필요한 설정들을 순차적으로 진행해보세요.
""")

# 세션 상태 초기화
def init_session_states():
    # 기본 활동 템플릿 (사용자 제공 데이터)
    default_activities = pd.DataFrame({
        "use": [True, True, True],
        "activity": ["토론면접", "발표준비", "발표면접"],
        "mode": ["batched", "parallel", "individual"],
        "duration_min": [30, 5, 15],
        "room_type": ["토론면접실", "발표준비실", "발표면접실"],
        "min_cap": [4, 1, 1],
        "max_cap": [6, 2, 1],
    })
    st.session_state.setdefault("activities", default_activities)
    
    # 스마트 직무 매핑 (사용자 제공 데이터 - 하위 호환성용)
    if "job_acts_map" not in st.session_state:
        act_list = default_activities.query("use == True")["activity"].tolist()
        job_data = {"code": ["JOB01"], "count": [20]}  # 기본값
        for act in act_list:
            job_data[act] = True
        st.session_state["job_acts_map"] = pd.DataFrame(job_data)
    
    # 기본 선후행 제약 (사용자 제공 데이터)
    default_precedence = pd.DataFrame([
        {"predecessor": "발표준비", "successor": "발표면접", "gap_min": 0, "adjacent": True}  # 발표준비 → 발표면접 (연속배치, 0분 간격)
    ])
    st.session_state.setdefault("precedence", default_precedence)
    
    # 기본 운영 시간 (사용자 제공 데이터: 09:00 ~ 17:30)
    st.session_state.setdefault("oper_start_time", time(9, 0))
    st.session_state.setdefault("oper_end_time", time(17, 30))
    
    # 스마트 방 템플릿 (사용자 제공 데이터)
    if "room_template" not in st.session_state:
        room_template = {
            "토론면접실": {"count": 2, "cap": 6},
            "발표준비실": {"count": 1, "cap": 2},  # 원래 디폴트: 1개, 2명
            "발표면접실": {"count": 2, "cap": 1}   # 원래 디폴트: 2개, 1명
        }
        st.session_state["room_template"] = room_template
    
    # 활동 데이터에서 방 타입 동기화
    acts_df = st.session_state.get("activities", pd.DataFrame())
    if not acts_df.empty and "room_type" in acts_df.columns:
        room_types = acts_df[acts_df["use"] == True]["room_type"].unique()
        room_template = st.session_state.get("room_template", {})
        
        # 새로운 방 타입 추가
        for room_type in room_types:
            if room_type and room_type not in room_template:
                room_template[room_type] = {"count": 1, "cap": 6}  # 기본값
        
        # 사용하지 않는 방 타입 제거
        used_room_types = set(room_types)
        room_template = {k: v for k, v in room_template.items() if k in used_room_types}
        
        st.session_state["room_template"] = room_template
    
    # 스마트 운영 공간 계획 (room_template 기반으로 자동 생성)
    if "room_plan" not in st.session_state:
        room_template = st.session_state.get("room_template", {})
        if room_template:
            final_plan_dict = {}
            for rt, values in room_template.items():
                final_plan_dict[f"{rt}_count"] = values['count']
                final_plan_dict[f"{rt}_cap"] = values['cap']
            st.session_state["room_plan"] = pd.DataFrame([final_plan_dict])
        else:
            st.session_state["room_plan"] = pd.DataFrame()
    
    # 스마트 운영 시간 (자동 생성)
    if "oper_window" not in st.session_state:
        t_start = st.session_state["oper_start_time"]
        t_end = st.session_state["oper_end_time"]
        oper_window_dict = {
            "start_time": t_start.strftime("%H:%M"),
            "end_time": t_end.strftime("%H:%M")
        }
        st.session_state["oper_window"] = pd.DataFrame([oper_window_dict])
    
    # 시뮬레이션 결과
    st.session_state.setdefault('final_schedule', None)
    st.session_state.setdefault('last_solve_logs', "")
    st.session_state.setdefault('solver_status', "미실행")
    st.session_state.setdefault('daily_limit', 0)
    
    # 집단면접 설정 초기화 (사용자 제공 데이터)
    st.session_state.setdefault('group_min_size', 4)
    st.session_state.setdefault('group_max_size', 6)
    st.session_state.setdefault('global_gap_min', 5)
    st.session_state.setdefault('max_stay_hours', 5)  # 5시간으로 변경
    
    # 멀티 날짜 계획 초기화 (사용자 제공 데이터)
    if "multidate_plans" not in st.session_state:
        from datetime import date, timedelta
        
        # 현재 날짜 기준으로 동적 생성
        today = date.today()
        
        multidate_plans = {
            today.strftime('%Y-%m-%d'): {
                "date": today,
                "enabled": True,
                "jobs": [
                    {"code": "JOB01", "count": 23},
                    {"code": "JOB02", "count": 23}
                ]
            },
            (today + timedelta(days=1)).strftime('%Y-%m-%d'): {
                "date": today + timedelta(days=1),
                "enabled": True,
                "jobs": [
                    {"code": "JOB03", "count": 20},
                    {"code": "JOB04", "count": 20}
                ]
            },
            (today + timedelta(days=2)).strftime('%Y-%m-%d'): {
                "date": today + timedelta(days=2),
                "enabled": True,
                "jobs": [
                    {"code": "JOB05", "count": 12},
                    {"code": "JOB06", "count": 15},
                    {"code": "JOB07", "count": 6}
                ]
            },
            (today + timedelta(days=3)).strftime('%Y-%m-%d'): {
                "date": today + timedelta(days=3),
                "enabled": True,
                "jobs": [
                    {"code": "JOB08", "count": 6},
                    {"code": "JOB09", "count": 6},
                    {"code": "JOB10", "count": 3},
                    {"code": "JOB11", "count": 3}
                ]
            }
        }
        st.session_state["multidate_plans"] = multidate_plans
    
    # 진행 상황 표시를 위한 세션 상태 초기화
    st.session_state.setdefault('progress_info', None)
    st.session_state.setdefault('current_stage', '준비중')
    st.session_state.setdefault('progress_value', 0.0)
    st.session_state.setdefault('stage_details', {})

init_session_states()

# =============================================================================
# 섹션 0: 운영일정 추정 (메인 섹션)
# =============================================================================
st.header("🚀 운영일정 추정")
st.markdown("현재 설정을 바탕으로 최적의 운영일정을 추정합니다.")

# 스케줄러 선택 - 성능 정보 추가
st.subheader("🔧 스케줄러 선택")

col1, col2 = st.columns([2, 1])

with col1:
    scheduler_choice = st.selectbox(
        "사용할 스케줄러를 선택하세요:",
        ["계층적 스케줄러 v2 (권장) - 2단계 하드 제약 포함", "OR-Tools 스케줄러 (기존)", "3단계 스케줄러 (새로 추가)"],
        help="계층적 v2는 대규모 처리에 최적화되어 있으며, 2단계 하드 제약 스케줄링을 기본으로 포함합니다. 3단계는 새로 추가된 스케줄러입니다."
    )

with col2:
    st.info("🚀 **성능 정보**\n\n"
           "**계층적 v2 (2단계 포함):**\n"
           "• 처리량: ~6,000명/초\n"
           "• 500명: ~0.1초\n"
           "• Batched 활동 지원\n"
           "• 메모리 효율적\n"
           "• 2단계 하드 제약 자동 적용\n\n"
           "**OR-Tools:**\n"
           "• 처리량: ~100명/초\n"
           "• 500명: ~5초\n"
           "• Individual만 지원\n"
           "• 최적해 보장\n\n"
           "**3단계 스케줄러:**\n"
           "• 처리량: ~1,000명/초\n"
           "• 500명: ~1초\n"
           "• 3단계 하드 제약 스케줄링 지원")

# 고급 옵션 (계층적 v2 선택시)
if "계층적" in scheduler_choice:
    with st.expander("⚙️ 고급 성능 옵션"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            enable_parallel = st.checkbox("병렬 처리 활성화", value=True, 
                                        help="멀티코어 CPU 활용")
            enable_caching = st.checkbox("캐싱 활성화", value=True,
                                       help="반복 계산 최적화")
        
        with col2:
            max_workers = st.number_input("최대 워커 수", min_value=1, max_value=16, 
                                        value=4, help="병렬 처리 워커 수")
            chunk_threshold = st.number_input("청킹 임계값", min_value=50, max_value=500,
                                            value=100, help="이 수 이상일 때 청킹 적용")
        
        with col3:
            memory_cleanup = st.number_input("메모리 정리 간격", min_value=10, max_value=100,
                                           value=50, help="N명마다 메모리 정리")
            
            # 성능 예측 (지원자 수 및 날짜 수 계산)
            multidate_plans = st.session_state.get("multidate_plans", {})
            
            if multidate_plans:
                # 새로운 멀티 날짜 계획 방식
                total_applicants = 0
                active_dates = []
                
                for date_key, plan in multidate_plans.items():
                    if plan.get("enabled", True):
                        active_dates.append(plan["date"])
                        for job in plan.get("jobs", []):
                            total_applicants += job.get("count", 0)
                
                num_dates = len(active_dates)
            else:
                # 기존 방식
                job_df = st.session_state.get("job_acts_map", pd.DataFrame())
                total_applicants = job_df["count"].sum() if not job_df.empty and "count" in job_df.columns else 0
                selected_dates = st.session_state.get("interview_dates", [])
                num_dates = len(selected_dates) if selected_dates else 1
                active_dates = selected_dates
            
            if total_applicants > 0:
                estimated_time = total_applicants / 6000  # 평균 처리량 기준 (날짜별 병렬 처리)
                st.success(f"📊 **예상 처리 시간**: {estimated_time:.3f}초 "
                         f"(총 {total_applicants}명)")
                
                # 멀티 날짜 정보 표시
                if num_dates > 1:
                    date_list = [d.strftime('%m/%d') for d in active_dates]
                    st.info(f"📅 **멀티 날짜 모드**: {num_dates}일간 ({', '.join(date_list)})")
            else:
                st.info("💡 지원자 수를 설정하면 예상 처리 시간을 확인할 수 있습니다.")

# 스케줄러 비교 정보 표시
comparison = get_scheduler_comparison()

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📊 OR-Tools 스케줄러 (기존)")
    legacy_info = comparison["legacy"]
    st.markdown(f"**설명:** {legacy_info['description']}")
    st.markdown("**장점:**")
    for pro in legacy_info['pros']:
        st.markdown(f"  ✅ {pro}")
    st.markdown("**단점:**")  
    for con in legacy_info['cons']:
        st.markdown(f"  ❌ {con}")
    st.markdown(f"**적합한 경우:** {legacy_info['suitable_for']}")

with col2:
    st.markdown("### 🚀 계층적 스케줄러 v2 (신규)")
    new_info = comparison["new"]
    st.markdown(f"**설명:** {new_info['description']}")
    st.markdown("**장점:**")
    for pro in new_info['pros']:
        st.markdown(f"  ✅ {pro}")
    st.markdown("**단점:**")
    for con in new_info['cons']:
        st.markdown(f"  ❌ {con}")
    st.markdown(f"**적합한 경우:** {new_info['suitable_for']}")

# 스케줄러 선택
use_new_scheduler = st.radio(
    "사용할 스케줄러를 선택하세요:",
    options=[True, False],
    format_func=lambda x: "🚀 계층적 스케줄러 v2 (권장)" if x else "📊 OR-Tools 스케줄러 (기존)",
    index=0,  # 기본값은 새로운 스케줄러
    help="대규모 처리나 Batched 활동이 필요한 경우 v2를 권장합니다."
)

st.session_state["use_new_scheduler"] = use_new_scheduler

if use_new_scheduler:
    st.success("🚀 **계층적 스케줄러 v2**를 선택했습니다. 빠른 처리와 Batched 활동을 지원합니다.")
else:
    st.info("📊 **OR-Tools 스케줄러**를 선택했습니다. 강력한 제약 해결 기능을 제공합니다.")

st.markdown("---")

# 첫 방문자를 위한 안내
if st.session_state.get('solver_status', '미실행') == '미실행':
    st.info("👋 **처음 방문하셨나요?** 바로 아래 '운영일정추정 시작' 버튼을 눌러보세요! 기본 설정으로 데모를 체험할 수 있습니다.")
    st.markdown("💡 **팁:** 추정 후 아래 섹션들에서 세부 설정을 조정하여 더 정확한 결과를 얻을 수 있습니다.")

# 헬퍼 함수들 (df_to_excel보다 먼저 정의)
def _convert_integrated_to_dual_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    🚀 이중 스케줄 표시: 통합된 활동을 원래 활동들로 분리하여 표시
    
    알고리즘은 통합 방식으로 처리하되, 사용자에게는 분리된 형태로 보여줍니다.
    이를 통해 공간 정보를 보존하고 운영 투명성을 확보합니다.
    """
    if df.empty:
        return df
    
    # 활동명 컬럼 찾기
    activity_col = None
    for col in ['activity_name', 'activity']:
        if col in df.columns:
            activity_col = col
            break
    
    if not activity_col:
        return df  # 활동 컬럼이 없으면 원본 반환
    
    # 통합된 활동 찾기 ('+' 포함된 활동명) - regex=False로 설정하여 문자 그대로 검색
    integrated_activities = df[df[activity_col].str.contains('+', na=False, regex=False)]
    
    if integrated_activities.empty:
        return df  # 통합된 활동이 없으면 원본 반환
    
    print(f"🔧 이중 스케줄 변환: {len(integrated_activities)}개 통합 활동 발견")
    
    # 분리된 스케줄 생성
    dual_schedule = []
    
    for _, row in df.iterrows():
        activity_name = str(row.get(activity_col, ''))
        
        if '+' in activity_name:
            # 통합된 활동을 분리
            parts = activity_name.split('+')
            if len(parts) == 2:
                pred_activity, succ_activity = parts[0].strip(), parts[1].strip()
                
                # 선행 활동 (발표준비)
                prep_row = row.copy()
                prep_row[activity_col] = pred_activity
                prep_row['room_name'] = _infer_prep_room(pred_activity, row)
                prep_row['duration_min'] = _get_activity_duration(pred_activity, default=5)
                prep_row['activity_stage'] = 1
                prep_row['original_integrated'] = activity_name
                
                # 후행 활동 (발표면접)
                interview_row = row.copy()
                interview_row[activity_col] = succ_activity
                # 방 정보는 원본 유지 (발표면접실)
                interview_row['duration_min'] = _get_activity_duration(succ_activity, default=15)
                interview_row['activity_stage'] = 2
                interview_row['original_integrated'] = activity_name
                
                # 시간 조정 (발표준비 → 발표면접 순서)
                _adjust_stage_times(prep_row, interview_row)
                
                dual_schedule.extend([prep_row, interview_row])
                
                print(f"  분리: {activity_name} → {pred_activity} + {succ_activity}")
            else:
                # 복잡한 통합 활동은 원본 유지
                dual_schedule.append(row)
        else:
            # 일반 활동은 그대로 유지
            dual_schedule.append(row)
    
    result_df = pd.DataFrame(dual_schedule)
    
    # 인덱스 재설정
    result_df = result_df.reset_index(drop=True)
    
    print(f"✅ 이중 스케줄 변환 완료: {len(df)} → {len(result_df)}개 항목")
    
    return result_df


def _infer_prep_room(prep_activity: str, original_row) -> str:
    """발표준비실 정보 추론"""
    # 활동명에서 방 타입 추론
    if '발표준비' in prep_activity:
        return '발표준비실'
    elif '면접준비' in prep_activity:
        return '면접준비실'
    elif '준비' in prep_activity:
        return f"{prep_activity}실"
    else:
        # 기본값으로 발표준비실 사용
        return '발표준비실'


def _get_activity_duration(activity_name: str, default: int = 10) -> int:
    """활동별 기본 소요시간 반환"""
    duration_map = {
        '발표준비': 5,
        '면접준비': 5,
        '발표면접': 15,
        '개별면접': 15,
        '토론면접': 30,
        '그룹면접': 30,
        '인성검사': 20,
        '적성검사': 30,
        '커피챗': 10
    }
    
    # 정확한 매칭 먼저 시도
    if activity_name in duration_map:
        return duration_map[activity_name]
    
    # 부분 매칭 시도
    for key, duration in duration_map.items():
        if key in activity_name:
            return duration
    
    return default


def _adjust_stage_times(prep_row, interview_row):
    """단계별 시간 조정"""
    # 시작 시간 컬럼 찾기
    start_col = None
    end_col = None
    
    for col in ['start_time', 'start']:
        if col in prep_row.index:
            start_col = col
            break
    
    for col in ['end_time', 'end']:
        if col in prep_row.index:
            end_col = col
            break
    
    if not start_col:
        return  # 시간 정보가 없으면 조정하지 않음
    
    # 원본 시작 시간
    original_start = prep_row[start_col]
    
    # 발표준비: 원본 시간부터 5분
    prep_duration = prep_row.get('duration_min', 5)
    
    # 발표면접: 발표준비 종료 직후 시작
    interview_start = _add_minutes_to_time(original_start, prep_duration)
    interview_row[start_col] = interview_start
    
    # 종료 시간도 조정 (있는 경우)
    if end_col:
        prep_end = _add_minutes_to_time(original_start, prep_duration)
        prep_row[end_col] = prep_end
        
        interview_duration = interview_row.get('duration_min', 15)
        interview_end = _add_minutes_to_time(interview_start, interview_duration)
        interview_row[end_col] = interview_end


def _add_minutes_to_time(time_val, minutes: int):
    """시간에 분을 추가"""
    try:
        if pd.isna(time_val):
            return time_val
        
        if isinstance(time_val, pd.Timedelta):
            return time_val + pd.Timedelta(minutes=minutes)
        elif hasattr(time_val, 'hour'):  # datetime-like
            return time_val + pd.Timedelta(minutes=minutes)
        else:
            # 문자열 등 기타 형식
            return time_val
    except Exception:
        return time_val  # 변환 실패 시 원본 반환


# Excel 출력 함수 (타임슬롯 기능 통합)
def df_to_excel(df: pd.DataFrame, stream=None) -> None:
    # 🚀 이중 스케줄 표시: 통합된 활동을 분리하여 공간 정보 보존
    df = _convert_integrated_to_dual_display(df)
    
    wb = Workbook()
    
    # 기본 팔레트
    PALETTE = ['E3F2FD', 'FFF3E0', 'E8F5E9', 'FCE4EC', 'E1F5FE', 'F3E5F5', 'FFFDE7', 'E0F2F1', 'EFEBE9', 'ECEFF1']
    
    # ===== 1) 기본 스케줄 시트 =====
    ws1 = wb.active
    ws1.title = 'Schedule'
    df_copy = df.copy()
    
    # 조 정보 추가 (batched 활동이 있는 경우)
    if 'activity_name' in df_copy.columns:
        # 새로운 구조에서 조 정보 계산
        df_copy['group_number'] = ''
        df_copy['group_size'] = ''
        
        # batched 활동별로 그룹 번호 부여
        group_counter = {}
        for idx, row in df_copy.iterrows():
            activity = row['activity_name']
            room = row.get('room_name', '')
            start_time = row.get('start_time', '')
            interview_date = row.get('interview_date', '')
            
            # 같은 날짜, 같은 활동, 같은 방, 같은 시간 = 같은 그룹
            group_key = f"{interview_date}_{activity}_{room}_{start_time}"
            
            if group_key not in group_counter:
                # 같은 날짜, 같은 활동의 기존 그룹 수 계산
                existing_groups = len([k for k in group_counter.keys() if k.startswith(f"{interview_date}_{activity}_")]) + 1
                group_counter[group_key] = {
                    'number': existing_groups,
                    'members': []
                }
            
            group_counter[group_key]['members'].append(idx)
            
            # 그룹 번호와 크기 설정 (패딩 적용)
            group_num = group_counter[group_key]['number']
            padded_num = f"{group_num:02d}" if group_num < 10 else str(group_num)
            df_copy.at[idx, 'group_number'] = f"{activity}-{padded_num}"
            df_copy.at[idx, 'group_size'] = len(group_counter[group_key]['members'])
        
        # 모든 그룹의 크기를 최종 크기로 업데이트
        for group_key, group_info in group_counter.items():
            final_size = len(group_info['members'])
            for member_idx in group_info['members']:
                df_copy.at[member_idx, 'group_size'] = final_size
    
    # 날짜별로 색상 지정
    unique_dates = df_copy['interview_date'].dt.date.unique()
    date_color_map = {date: PALETTE[i % len(PALETTE)] for i, date in enumerate(unique_dates)}
    
    df_copy = df_copy.astype(object).where(pd.notna(df_copy), None)
    for r in dataframe_to_rows(df_copy, index=False, header=True):
        ws1.append(r)
    
    header_fill = PatternFill('solid', fgColor='D9D9D9')
    for cell in ws1[1]:
        cell.fill = header_fill
    
    # 날짜 열 찾기
    date_col_idx = -1
    for j, col_name in enumerate(df_copy.columns, 1):
        if col_name == 'interview_date':
            date_col_idx = j
            break
    
    for i, row in enumerate(ws1.iter_rows(min_row=2, max_row=ws1.max_row), 2):
        if date_col_idx != -1:
            date_val = row[date_col_idx - 1].value
            if date_val and hasattr(date_val, 'date'):
                row_color = date_color_map.get(date_val.date())
                if row_color:
                    row_fill = PatternFill('solid', fgColor=row_color)
                    for cell in row:
                        cell.fill = row_fill
    
    # 시간 형식 지정
    for j, col_name in enumerate(df_copy.columns, 1):
        if 'start' in col_name or 'end' in col_name:
            for i in range(2, ws1.max_row + 1):
                ws1.cell(i, j).number_format = 'hh:mm'
    
    # Schedule 시트 첫행 고정
    ws1.freeze_panes = 'A2'
    
    # ===== 2) 데이터 분석 시트들 추가 =====
    
    # 3단계 스케줄링 결과가 있는 경우 우선 처리
    if hasattr(st.session_state, 'three_phase_reports') and st.session_state.three_phase_reports:
        three_phase_reports = st.session_state['three_phase_reports']
        
        # 3단계 결과 분석 시트 추가
        if 'phase3' in three_phase_reports and three_phase_reports['phase3']['df'] is not None:
            ws_phase3 = wb.create_sheet('3단계_스케줄링_결과')
            phase3_df = three_phase_reports['phase3']['df']
            
            # 3단계 결과를 UI 형식으로 변환
            phase3_display = phase3_df.copy()
            if 'interview_date' in phase3_display.columns:
                phase3_display['interview_date'] = pd.to_datetime(phase3_display['interview_date']).dt.strftime('%Y-%m-%d')
            
            for r in dataframe_to_rows(phase3_display, index=False, header=True):
                ws_phase3.append(r)
            
            # 헤더 스타일링
            for cell in ws_phase3[1]:
                cell.fill = PatternFill('solid', fgColor='E8F5E9')
                cell.font = Font(bold=True)
        
        # 3단계 체류시간 분석 시트 추가
        if 'phase3' in three_phase_reports and three_phase_reports['phase3']['df'] is not None:
            ws_phase3_analysis = wb.create_sheet('3단계_체류시간_분석')
            
            # 3단계 결과에서 체류시간 계산
            phase3_df = three_phase_reports['phase3']['df']
            phase3_df['interview_date'] = pd.to_datetime(phase3_df['interview_date'])
            
            stay_time_data = []
            for date_str in phase3_df['interview_date'].dt.strftime('%Y-%m-%d').unique():
                date_df = phase3_df[phase3_df['interview_date'].dt.strftime('%Y-%m-%d') == date_str]
                for applicant_id in date_df['applicant_id'].unique():
                    applicant_df = date_df[date_df['applicant_id'] == applicant_id]
                    start_time = applicant_df['start_time'].min()
                    end_time = applicant_df['end_time'].max()
                    stay_hours = (end_time - start_time).total_seconds() / 3600
                    stay_time_data.append({
                        '날짜': date_str,
                        '응시자ID': applicant_id,
                        '체류시간(시간)': round(stay_hours, 2)
                    })
            
            if stay_time_data:
                stay_time_df = pd.DataFrame(stay_time_data)
                for r in dataframe_to_rows(stay_time_df, index=False, header=True):
                    ws_phase3_analysis.append(r)
                
                # 헤더 스타일링
                for cell in ws_phase3_analysis[1]:
                    cell.fill = PatternFill('solid', fgColor='E8F5E9')
                    cell.font = Font(bold=True)
    
    # 하드 제약 분석 시트 추가 (2단계 스케줄링 결과가 있는 경우)
    elif hasattr(st.session_state, 'two_phase_reports') and st.session_state.two_phase_reports:
        reports = st.session_state.two_phase_reports
        
        # 제약 분석 리포트
        if 'constraint_analysis' in reports and not reports['constraint_analysis'].empty:
            ws_constraint = wb.create_sheet('Hard_Constraint_Analysis')
            constraint_df = reports['constraint_analysis']
            
            for r in dataframe_to_rows(constraint_df, index=False, header=True):
                ws_constraint.append(r)
            
            # 헤더 스타일링
            for cell in ws_constraint[1]:
                cell.fill = PatternFill('solid', fgColor='FFE6CC')
                cell.font = Font(bold=True)
        
        # 제약 위반 리포트
        if 'constraint_violations' in reports and not reports['constraint_violations'].empty:
            ws_violations = wb.create_sheet('Constraint_Violations')
            violations_df = reports['constraint_violations']
            
            for r in dataframe_to_rows(violations_df, index=False, header=True):
                ws_violations.append(r)
            
            # 헤더 스타일링
            for cell in ws_violations[1]:
                cell.fill = PatternFill('solid', fgColor='FFCCCC')
                cell.font = Font(bold=True)
        
        # 단계별 비교 리포트
        if 'phase_comparison' in reports and not reports['phase_comparison'].empty:
            ws_comparison = wb.create_sheet('Phase_Comparison')
            comparison_df = reports['phase_comparison']
            
            for r in dataframe_to_rows(comparison_df, index=False, header=True):
                ws_comparison.append(r)
            
            # 헤더 스타일링
            for cell in ws_comparison[1]:
                cell.fill = PatternFill('solid', fgColor='CCE6FF')
                cell.font = Font(bold=True)
    
    # 체류시간 통계 계산 함수 (내부 정의)
    def calculate_stay_duration_stats_internal(schedule_df):
        """각 지원자의 체류시간을 계산하고 통계를 반환"""
        stats_data = []
        
        # 컬럼명 매핑 (실제 데이터에 맞게 조정)
        id_col = None
        for col in ['applicant_id', 'id', 'candidate_id']:
            if col in schedule_df.columns:
                id_col = col
                break
        
        job_col = None
        for col in ['job_code', 'code']:
            if col in schedule_df.columns:
                job_col = col
                break
        
        date_col = None
        for col in ['interview_date', 'date']:
            if col in schedule_df.columns:
                date_col = col
                break
        
        if not id_col or not job_col or not date_col:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        # 지원자별 체류시간 계산
        for candidate_id in schedule_df[id_col].unique():
            candidate_data = schedule_df[schedule_df[id_col] == candidate_id]
            
            # 더미 데이터 제외
            if str(candidate_id).startswith('dummy'):
                continue
            
            if len(candidate_data) == 0:
                continue
            
            # 시간 파싱
            all_start_times = []
            all_end_times = []
            
            for _, row in candidate_data.iterrows():
                try:
                    start_time = row['start_time']
                    end_time = row['end_time']
                    
                    # timedelta 처리
                    if isinstance(start_time, pd.Timedelta):
                        all_start_times.append(start_time)
                    elif isinstance(start_time, str):
                        start_time = pd.to_datetime(start_time, format='%H:%M:%S').time()
                        start_td = timedelta(hours=start_time.hour, minutes=start_time.minute, seconds=start_time.second)
                        all_start_times.append(start_td)
                    
                    if isinstance(end_time, pd.Timedelta):
                        all_end_times.append(end_time)
                    elif isinstance(end_time, str):
                        end_time = pd.to_datetime(end_time, format='%H:%M:%S').time()
                        end_td = timedelta(hours=end_time.hour, minutes=end_time.minute, seconds=end_time.second)
                        all_end_times.append(end_td)
                    
                except Exception as e:
                    continue
            
            if all_start_times and all_end_times:
                # 전체 체류시간 = 첫 번째 활동 시작 ~ 마지막 활동 종료
                total_start = min(all_start_times)
                total_end = max(all_end_times)
                stay_duration_hours = (total_end - total_start).total_seconds() / 3600
                
                # 직무 코드 및 날짜 (첫 번째 행에서 가져오기)
                job_code = candidate_data.iloc[0].get(job_col, 'Unknown')
                interview_date = candidate_data.iloc[0].get(date_col, 'Unknown')
                
                stats_data.append({
                    'candidate_id': candidate_id,
                    'job_code': job_code,
                    'interview_date': interview_date,
                    'stay_duration_hours': stay_duration_hours,
                    'start_time': total_start,
                    'end_time': total_end
                })
        
        if not stats_data:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        stats_df = pd.DataFrame(stats_data)
        
        # 직무별 통계 계산
        job_stats = []
        for job_code, job_data in stats_df.groupby('job_code'):
            durations = job_data['stay_duration_hours']
            job_stats.append({
                'job_code': job_code,
                'count': len(job_data),
                'min_duration': durations.min(),
                'max_duration': durations.max(),
                'avg_duration': durations.mean(),
                'median_duration': durations.median()
            })
        
        # 날짜별 통계 계산
        date_stats = []
        for date, date_data in stats_df.groupby('interview_date'):
            durations = date_data['stay_duration_hours']
            max_stay_candidate = date_data.loc[date_data['stay_duration_hours'].idxmax()]
            
            date_stats.append({
                'interview_date': date,
                'count': len(date_data),
                'min_duration': durations.min(),
                'max_duration': durations.max(),
                'avg_duration': durations.mean(),
                'max_stay_candidate': max_stay_candidate['candidate_id'],
                'max_stay_job': max_stay_candidate['job_code']
            })
        
        return pd.DataFrame(job_stats), stats_df, pd.DataFrame(date_stats)
    
    # 체류시간 통계 계산
    job_stats_df, individual_stats_df, date_stats_df = calculate_stay_duration_stats_internal(df)
    
    # 날짜별 상세 통계 시트
    if not individual_stats_df.empty:
        ws_stats = wb.create_sheet('StayTime_Analysis')
        
        # 날짜별 통계 데이터 작성
        stats_data = []
        for date, date_data in individual_stats_df.groupby('interview_date'):
            durations = date_data['stay_duration_hours']
            
            # 기본 통계
            stats_row = {
                '날짜': date,
                '응시자수': len(date_data),
                '평균체류시간(시간)': round(durations.mean(), 2),
                '중간값체류시간(시간)': round(durations.median(), 2),
                '최소체류시간(시간)': round(durations.min(), 2),
                '최대체류시간(시간)': round(durations.max(), 2),
                '표준편차(시간)': round(durations.std(), 2),
                '최소체류자ID': date_data.loc[durations.idxmin(), 'candidate_id'],
                '최소체류자직무': date_data.loc[durations.idxmin(), 'job_code'],
                '최대체류자ID': date_data.loc[durations.idxmax(), 'candidate_id'],
                '최대체류자직무': date_data.loc[durations.idxmax(), 'job_code']
            }
            stats_data.append(stats_row)
        
        # 통계 데이터프레임 생성
        stats_df = pd.DataFrame(stats_data)
        
        # 엑셀에 작성
        for r in dataframe_to_rows(stats_df, index=False, header=True):
            ws_stats.append(r)
        
        # 헤더 스타일 적용
        header_fill = PatternFill('solid', fgColor='D9D9D9')
        for cell in ws_stats[1]:
            cell.fill = header_fill
        
        # 컬럼 너비 조정
        for column in ws_stats.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_stats.column_dimensions[column_letter].width = adjusted_width
    
    # 개별 지원자 체류시간 시트
    if not individual_stats_df.empty:
        ws_individual = wb.create_sheet('Individual_StayTime')
        
        # 개별 데이터 정리
        individual_data = individual_stats_df.copy()
        individual_data['체류시간(시간)'] = individual_data['stay_duration_hours'].round(2)
        
        # 시간 표시 함수
        def format_timedelta(td):
            if pd.isna(td):
                return ''
            if isinstance(td, pd.Timedelta):
                total_seconds = td.total_seconds()
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                return f"{hours:02d}:{minutes:02d}"
            return str(td)
        
        individual_data['시작시간'] = individual_data['start_time'].apply(format_timedelta)
        individual_data['종료시간'] = individual_data['end_time'].apply(format_timedelta)
        
        # 컬럼 선택 및 한글화
        display_columns = ['candidate_id', 'job_code', 'interview_date', '체류시간(시간)', '시작시간', '종료시간']
        individual_display = individual_data[display_columns].copy()
        individual_display.columns = ['지원자ID', '직무코드', '면접일자', '체류시간(시간)', '시작시간', '종료시간']
        
        # 정렬 (날짜별, 체류시간별)
        individual_display = individual_display.sort_values(['면접일자', '체류시간(시간)'], ascending=[True, False])
        
        # 엑셀에 작성
        for r in dataframe_to_rows(individual_display, index=False, header=True):
            ws_individual.append(r)
        
        # 헤더 스타일 적용
        header_fill = PatternFill('solid', fgColor='D9D9D9')
        for cell in ws_individual[1]:
            cell.fill = header_fill
        
        # 컬럼 너비 조정
        for column in ws_individual.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws_individual.column_dimensions[column_letter].width = adjusted_width
    
    # 직무별 통계 시트
    if not job_stats_df.empty:
        ws_job = wb.create_sheet('Job_Statistics')
        
        # 직무별 데이터 정리
        job_display = job_stats_df.copy()
        job_display['최소시간(시간)'] = job_display['min_duration'].round(2)
        job_display['최대시간(시간)'] = job_display['max_duration'].round(2)
        job_display['평균시간(시간)'] = job_display['avg_duration'].round(2)
        job_display['중간값시간(시간)'] = job_display['median_duration'].round(2)
        
        # 컬럼 선택 및 한글화
        display_columns = ['job_code', 'count', '최소시간(시간)', '최대시간(시간)', '평균시간(시간)', '중간값시간(시간)']
        job_display = job_display[display_columns].copy()
        job_display.columns = ['직무코드', '인원수', '최소시간(시간)', '최대시간(시간)', '평균시간(시간)', '중간값시간(시간)']
        
        # 엑셀에 작성
        for r in dataframe_to_rows(job_display, index=False, header=True):
            ws_job.append(r)
        
        # 헤더 스타일 적용
        header_fill = PatternFill('solid', fgColor='D9D9D9')
        for cell in ws_job[1]:
            cell.fill = header_fill
        
        # 컬럼 너비 조정
        for column in ws_job.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_job.column_dimensions[column_letter].width = adjusted_width
    
    # 방별 사용률 분석 시트
    if not df.empty and 'room_name' in df.columns:
        ws_room = wb.create_sheet('Room_Utilization')
        
        # 방별 사용 통계 계산
        room_stats = []
        for room_name in df['room_name'].unique():
            room_data = df[df['room_name'] == room_name]
            
            # 방별 사용 시간 계산
            total_usage_minutes = 0
            for _, row in room_data.iterrows():
                start_time = row['start_time']
                end_time = row['end_time']
                
                if pd.notna(start_time) and pd.notna(end_time):
                    if isinstance(start_time, pd.Timedelta) and isinstance(end_time, pd.Timedelta):
                        duration = (end_time - start_time).total_seconds() / 60
                        total_usage_minutes += duration
            
            # 운영 시간 (8시간 = 480분)
            operating_minutes = 480
            utilization_rate = (total_usage_minutes / operating_minutes) * 100 if operating_minutes > 0 else 0
            
            room_stats.append({
                '방이름': room_name,
                '사용횟수': len(room_data),
                '총사용시간(분)': round(total_usage_minutes, 1),
                '총사용시간(시간)': round(total_usage_minutes / 60, 2),
                '사용률(%)': round(utilization_rate, 1),
                '평균사용시간(분)': round(total_usage_minutes / len(room_data), 1) if len(room_data) > 0 else 0
            })
        
        # 사용률 기준으로 정렬
        room_stats_df = pd.DataFrame(room_stats)
        room_stats_df = room_stats_df.sort_values('사용률(%)', ascending=False)
        
        # 엑셀에 작성
        for r in dataframe_to_rows(room_stats_df, index=False, header=True):
            ws_room.append(r)
        
        # 헤더 스타일 적용
        header_fill = PatternFill('solid', fgColor='D9D9D9')
        for cell in ws_room[1]:
            cell.fill = header_fill
        
        # 컬럼 너비 조정
        for column in ws_room.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws_room.column_dimensions[column_letter].width = adjusted_width
    
    # ===== 4) 타임슬롯 시트들 추가 =====
    def _color_picker():
        """활동명 → 고정 색상 매핑"""
        mapping = {}
        def _pick(act: str) -> str:
            if act not in mapping:
                mapping[act] = PALETTE[len(mapping) % len(PALETTE)]
            return mapping[act]
        return _pick
    
    def _build_timeslot_sheet(ws, df_day: pd.DataFrame, pick_color):
        """단일 날짜 스케줄 → 타임슬롯 매트릭스 (참고 코드 기반)"""
        # 새로운 구조인지 확인
        if 'room_name' in df_day.columns and 'activity_name' in df_day.columns:
            # 새로운 구조를 기존 구조 형태로 변환
            df_converted = _convert_new_to_legacy_format(df_day)
            if df_converted is not None:
                _build_timeslot_sheet_legacy(ws, df_converted, pick_color)
            return
        else:
            # 기존 구조 그대로 처리
            _build_timeslot_sheet_legacy(ws, df_day, pick_color)
            return
    
    def _convert_new_to_legacy_format(df_new):
        """새로운 구조(room_name, activity_name)를 기존 구조(loc_*, start_*, end_*)로 변환"""
        try:
            # 지원자별로 그룹화하여 wide format으로 변환
            result_rows = []
            
            for applicant_id, applicant_data in df_new.groupby('applicant_id'):
                row_data = {'id': applicant_id}
                
                # 각 활동별로 컬럼 생성
                for _, activity_row in applicant_data.iterrows():
                    activity = activity_row['activity_name']
                    room = activity_row['room_name']
                    start_time = activity_row['start_time']
                    end_time = activity_row['end_time']
                    
                    # timedelta를 datetime으로 변환
                    if isinstance(start_time, pd.Timedelta):
                        # 오늘 날짜를 기준으로 datetime 생성
                        base_date = pd.Timestamp.today().normalize()
                        start_time = base_date + start_time
                    
                    if isinstance(end_time, pd.Timedelta):
                        base_date = pd.Timestamp.today().normalize()
                        end_time = base_date + end_time
                    
                    # 활동명을 컬럼 접미사로 사용
                    row_data[f'loc_{activity}'] = room
                    row_data[f'start_{activity}'] = start_time
                    row_data[f'end_{activity}'] = end_time
                
                result_rows.append(row_data)
            
            if not result_rows:
                return None
                
            return pd.DataFrame(result_rows)
            
        except Exception as e:
            print(f"변환 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _build_timeslot_sheet_legacy(ws, df_day: pd.DataFrame, pick_color):
        """기존 구조용 타임슬롯 시트 생성 (참고 코드 기반)"""
        import re
        
        # 컬럼 찾기
        loc_cols = [c for c in df_day.columns if c.startswith('loc_')]
        start_cols = [c for c in df_day.columns if c.startswith('start_')]
        end_cols = [c for c in df_day.columns if c.startswith('end_')]
        
        # 공간 목록
        locs = sorted(set(df_day[loc_cols].stack().dropna().unique()))
        if not locs:
            return
        
        # 시간 범위 계산
        t_min, t_max = None, None
        for col in start_cols + end_cols:
            # 컬럼의 데이터 타입 확인 및 적절한 변환
            col_data = df_day[col].dropna()
            if col_data.empty:
                continue
            
            # 각 값에 대해 적절한 변환 수행
            converted_times = []
            for val in col_data:
                if pd.isna(val):
                    continue
                try:
                    if isinstance(val, pd.Timedelta):
                        # timedelta를 오늘 날짜 기준 datetime으로 변환
                        base_date = pd.Timestamp.today().normalize()
                        converted_time = base_date + val
                    else:
                        # 일반적인 datetime 변환
                        converted_time = pd.to_datetime(val, errors='coerce')
                    
                    if not pd.isna(converted_time):
                        converted_times.append(converted_time)
                except Exception:
                    continue
            
            if converted_times:
                col_min = min(converted_times)
                col_max = max(converted_times)
                t_min = col_min if t_min is None else min(t_min, col_min)
                t_max = col_max if t_max is None else max(t_max, col_max)
        
        if t_min is None or t_max is None:
            return
        
        TIME_STEP_MIN = 5
        t_min = t_min.floor(f'{TIME_STEP_MIN}min')
        t_max = (t_max.ceil(f'{TIME_STEP_MIN}min') + timedelta(minutes=TIME_STEP_MIN))
        times = pd.date_range(t_min, t_max, freq=f'{TIME_STEP_MIN}min')
        
        # 헤더
        ws.cell(1, 1, 'Time')
        for j, loc in enumerate(locs, start=2):
            cell = ws.cell(1, j, loc)
            cell.alignment = Alignment(horizontal='center')
        
        for i, t in enumerate(times, start=2):
            ws.cell(i, 1, t.strftime('%H:%M'))
            ws.cell(i, 1).alignment = Alignment(horizontal='right')
        
        # 셀 채우기 (참고 코드 스타일)
        for _, row in df_day.iterrows():
            for st_col in start_cols:
                suffix = st_col[len('start_'):]
                end_col = f'end_{suffix}'
                loc_col = f'loc_{suffix}'
                if end_col not in df_day.columns or loc_col not in df_day.columns:
                    continue
                st = row[st_col]
                ed = row[end_col]
                loc = row[loc_col]
                if pd.isna(st) or pd.isna(ed) or loc in ('', None):
                    continue
                    
                # timedelta 처리 개선
                try:
                    if isinstance(st, pd.Timedelta):
                        base_date = pd.Timestamp.today().normalize()
                        st_dt = base_date + st
                    else:
                        st_dt = pd.to_datetime(st, errors='coerce')
                    
                    if isinstance(ed, pd.Timedelta):
                        base_date = pd.Timestamp.today().normalize()
                        ed_dt = base_date + ed
                    else:
                        ed_dt = pd.to_datetime(ed, errors='coerce')
                    
                    if pd.isna(st_dt) or pd.isna(ed_dt):
                        continue
                except Exception:
                    continue
                    
                # 활동명에서 버전 제거 (참고 코드 스타일)
                base_act = re.sub(r'_v\d+$', '', suffix)
                color = pick_color(base_act)
                if loc not in locs:
                    continue
                col_idx = locs.index(loc) + 2
                cur = st_dt.floor(f'{TIME_STEP_MIN}min')
                while cur < ed_dt:
                    if cur < t_min or cur > t_max:
                        cur += timedelta(minutes=TIME_STEP_MIN)
                        continue
                    row_idx = times.get_loc(cur) + 2
                    cell = ws.cell(row_idx, col_idx)
                    # 값 누적: 이미 다른 사람이 있으면 줄 바꿈 후 추가 (참고 코드 스타일)
                    if cell.value in (None, ''):
                        cell.value = str(row['id'])
                        # 첫 입력 시에만 색상 채우기
                        cell.fill = PatternFill('solid', fgColor=color)
                    else:
                        existing_ids = str(cell.value)
                        if str(row['id']) not in existing_ids.split('\n'):
                            cell.value = existing_ids + '\n' + str(row['id'])
                    # 항상 wrap_text 적용
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cur += timedelta(minutes=TIME_STEP_MIN)
        
        # 열 너비·행 높이 자동 조정 (참고 코드 스타일)
        for j, loc in enumerate(locs, start=2):
            max_len = len(str(loc))
            for i in range(2, ws.max_row + 1):
                val = ws.cell(i, j).value
                if val is None:
                    continue
                for part in str(val).split('\n'):
                    max_len = max(max_len, len(part))
            col_letter = get_column_letter(j)
            # 대략 문자 1개당 1.2 단위, 최소 10, 최대 30
            ws.column_dimensions[col_letter].width = max(10, min(1.2 * max_len, 30))
        
        # 행 높이: 줄 개수에 비례해서 높이 늘리기
        default_ht = 15  # 기본 높이(approx)
        for i in range(2, ws.max_row + 1):
            max_lines = 1
            for j in range(2, ws.max_column + 1):
                val = ws.cell(i, j).value
                if val is None:
                    continue
                lines = str(val).count('\n') + 1
                max_lines = max(max_lines, lines)
            ws.row_dimensions[i].height = default_ht * max_lines
        
        # TS_ 시트 첫행 고정
        ws.freeze_panes = 'A2'
    
    # 날짜별 타임슬롯 시트 생성
    pick_color = _color_picker()
    for the_date, df_day in df.groupby("interview_date"):
        ws_name = f"TS_{pd.to_datetime(the_date).strftime('%m%d')}"
        ws_ts = wb.create_sheet(ws_name)
        _build_timeslot_sheet(ws_ts, df_day.copy(), pick_color)
    
    wb.save(stream or "interview_schedule.xlsx")

def reset_run_state():
    st.session_state['final_schedule'] = None
    st.session_state['last_solve_logs'] = ""
    st.session_state['solver_status'] = "미실행"
    st.session_state['daily_limit'] = 0
    st.session_state['two_phase_reports'] = {}
    st.session_state['three_phase_reports'] = None

# 기본 파라미터 설정 (하드코딩)
params = {
    "min_gap_min": st.session_state.get('global_gap_min', 5),
    "time_limit_sec": 120,
    "max_stay_hours": st.session_state.get('max_stay_hours', 8)
}

# 활동 데이터 가져오기 및 필터링
acts_df = st.session_state.get("activities", pd.DataFrame())
if not acts_df.empty and "use" in acts_df.columns:
    # use=True인 활동만 필터링
    acts_df = acts_df[acts_df["use"] == True].copy()
    st.session_state["activities"] = acts_df  # 필터링된 데이터로 업데이트

# batched 모드가 있는지 확인
has_batched = any(acts_df["mode"] == "batched") if not acts_df.empty and "mode" in acts_df.columns else False

# 집단면접 설정이 있으면 params에 추가
if has_batched:
    params["group_min_size"] = st.session_state.get('group_min_size', 4)
    params["group_max_size"] = st.session_state.get('group_max_size', 6)

# 데이터 검증
job_df = st.session_state.get("job_acts_map", pd.DataFrame())
room_plan = st.session_state.get("room_plan", pd.DataFrame())
oper_df = st.session_state.get("oper_window", pd.DataFrame())
prec_df = st.session_state.get("precedence", pd.DataFrame())

# 검증 결과 표시
validation_errors = []

if acts_df.empty or not (acts_df["use"] == True).any():
    validation_errors.append("활동을 하나 이상 정의하고 'use=True'로 설정해주세요.")

if job_df.empty:
    validation_errors.append("직무 코드를 추가해주세요.")
elif "count" not in job_df.columns:
    validation_errors.append("직무 매핑 데이터에 'count' 컬럼이 없습니다.")
elif job_df["count"].sum() == 0:
    validation_errors.append("직무 코드를 추가하고 인원수를 1명 이상 입력해주세요.")

if job_df["code"].duplicated().any():
    validation_errors.append("중복된 직무 코드가 있습니다.")

# batched 모드가 있으면 그룹 크기 일관성 검증
if has_batched:
    batched_activities = acts_df[acts_df["mode"] == "batched"]
    min_caps = batched_activities["min_cap"].unique()
    max_caps = batched_activities["max_cap"].unique()
    
    if len(min_caps) > 1 or len(max_caps) > 1:
        validation_errors.append("모든 batched 활동의 그룹 크기(min_cap, max_cap)는 동일해야 합니다.")
    
    # 방 용량 vs 그룹 크기 검증
    for _, act in batched_activities.iterrows():
        room_type = act["room_type"]
        max_cap = act["max_cap"]
        room_cap_col = f"{room_type}_cap"
        
        if room_cap_col in room_plan.columns:
            room_cap = room_plan[room_cap_col].iloc[0] if not room_plan.empty else 0
            if room_cap < max_cap:
                validation_errors.append(f"{room_type}의 용량({room_cap})이 {act['activity']}의 최대 그룹 크기({max_cap})보다 작습니다.")

# room_plan 검증: room_template이 있으면 유효한 것으로 간주
room_template = st.session_state.get("room_template", {})
if room_plan.empty and not room_template:
    validation_errors.append("운영 공간을 설정해주세요.")

if validation_errors:
    st.error("다음 항목을 설정해주세요:")
    for error in validation_errors:
        st.error(f"• {error}")
    st.info("⬇️ 아래 섹션들에서 필요한 설정을 완료한 후 다시 추정해보세요.")
else:
    st.success("✅ 입력 데이터 검증 통과 – 운영일정추정을 시작할 수 있습니다!")

# 스케줄링 옵션 선택
st.markdown("### 🎯 스케줄링 옵션")

st.info("💡 **계층적 스케줄러 v2**를 선택하면 자동으로 2단계 하드 제약 스케줄링이 적용됩니다.\n"
        "1단계: 초기 스케줄링 → 2단계: 90% 분위수 기반 하드 제약 적용 → 3단계: 최적화된 재스케줄링")

# 운영일정 추정 실행
if st.button("🚀 운영일정추정 시작", type="primary", use_container_width=True, on_click=reset_run_state):
    if not validation_errors:
        # 🔍 디버깅: 현재 세션 상태를 파일로 저장
        try:
            import json
            from datetime import datetime
            
            debug_data = {
                "timestamp": datetime.now().isoformat(),
                "activities": st.session_state.get("activities", pd.DataFrame()).to_dict() if not st.session_state.get("activities", pd.DataFrame()).empty else None,
                "multidate_plans": st.session_state.get("multidate_plans", {}),
                "room_template": st.session_state.get("room_template", {}),
                "job_acts_map": st.session_state.get("job_acts_map", pd.DataFrame()).to_dict() if not st.session_state.get("job_acts_map", pd.DataFrame()).empty else None,
                "oper_start_time": str(st.session_state.get("oper_start_time", "09:00")),
                "oper_end_time": str(st.session_state.get("oper_end_time", "17:30")),
                "global_gap_min": st.session_state.get("global_gap_min", 5),
                "max_stay_hours": st.session_state.get("max_stay_hours", 8)
            }
            
            with open("ui_session_debug.json", "w", encoding="utf-8") as f:
                json.dump(debug_data, f, ensure_ascii=False, indent=2, default=str)
            
            st.info("🔍 디버깅: 세션 상태가 ui_session_debug.json에 저장되었습니다.")
            
        except Exception as e:
            st.warning(f"⚠️ 디버깅 저장 실패: {e}")
        
        with st.spinner("최적의 운영 일정을 계산 중입니다..."):
            try:
                cfg = core.build_config(st.session_state)
                
                # 진행 상황 표시를 위한 UI 구성
                progress_container = st.container()
                
                with progress_container:
                    st.markdown("### 🚀 스케줄링 진행 상황")
                    
                    # Progress Bar
                    progress_bar = st.progress(0.0)
                    status_text = st.empty()
                    details_text = st.empty()
                    
                    # 실시간 로그 표시
                    log_container = st.container()
                    with log_container:
                        st.markdown("#### 📋 실시간 로그")
                        log_area = st.empty()
                
                # 진행 상황 업데이트 함수
                def update_progress():
                    info = st.session_state.get('progress_info')
                    if info:
                        # Progress bar 업데이트
                        progress_bar.progress(info.progress)
                        
                        # 상태 텍스트 업데이트
                        stage_emoji = {
                            "Level1": "🔧",
                            "Level2": "📊", 
                            "Level3": "👥",
                            "Backtrack": "🔄",
                            "Complete": "✅",
                            "Error": "❌"
                        }
                        emoji = stage_emoji.get(info.stage, "⚡")
                        status_text.markdown(f"**{emoji} {info.stage}**: {info.message}")
                        
                        # 상세 정보 표시
                        if info.details:
                            details = []
                            for key, value in info.details.items():
                                if key == "time":
                                    details.append(f"소요시간: {value:.1f}초")
                                elif key == "groups":
                                    details.append(f"그룹 수: {value}개")
                                elif key == "dummies":
                                    details.append(f"더미: {value}명")
                                elif key == "schedule_count":
                                    details.append(f"스케줄: {value}개")
                                elif key == "backtrack_count":
                                    details.append(f"백트래킹: {value}회")
                                else:
                                    details.append(f"{key}: {value}")
                            
                            if details:
                                details_text.info(" | ".join(details))
                        
                        # 로그 업데이트
                        log_area.text(f"[{info.timestamp.strftime('%H:%M:%S')}] {info.message}")
                
                # UI 수정사항 즉시 반영 (방법 1) - 강화된 버전
                # 현재 세션 상태에서 UI 수정사항을 강제로 적용
                acts_df = st.session_state.get("activities", pd.DataFrame())
                if not acts_df.empty and "room_type" in acts_df.columns:
                    # 1. 활동 데이터 동기화 (use=True인 활동만)
                    valid_acts_df = acts_df[acts_df["use"] == True].copy()
                    st.session_state["activities"] = valid_acts_df
                    
                    # 2. 방 타입 동기화
                    room_types = valid_acts_df["room_type"].unique()
                    room_template = st.session_state.get("room_template", {})
                    
                    # 새로운 방 타입 추가
                    for room_type in room_types:
                        if room_type and room_type not in room_template:
                            room_template[room_type] = {"count": 1, "cap": 6}  # 기본값
                    
                    # 사용하지 않는 방 타입 제거
                    used_room_types = set(room_types)
                    room_template = {k: v for k, v in room_template.items() if k in used_room_types}
                    st.session_state["room_template"] = room_template
                    
                    # 3. room_plan 동기화
                    room_plan_dict = {}
                    for rt, values in room_template.items():
                        room_plan_dict[f"{rt}_count"] = values['count']
                        room_plan_dict[f"{rt}_cap"] = values['cap']
                    
                    st.session_state["room_plan"] = pd.DataFrame([room_plan_dict])
                    
                    # 4. 직무 매핑 동기화 (활동명 변경 반영)
                    job_acts_map = st.session_state.get("job_acts_map", pd.DataFrame())
                    if not job_acts_map.empty:
                        act_list = valid_acts_df["activity"].tolist()
                        
                        # 기존 활동명과 새로운 활동명 매핑
                        old_acts = [col for col in job_acts_map.columns if col not in ["code", "count"]]
                        new_acts = act_list
                        
                        # 새로운 활동 추가
                        for act in new_acts:
                            if act not in job_acts_map.columns:
                                job_acts_map[act] = True
                        
                        # 사용하지 않는 활동 제거
                        unused_acts = [act for act in old_acts if act not in new_acts]
                        if unused_acts:
                            job_acts_map = job_acts_map.drop(columns=unused_acts)
                        
                        st.session_state["job_acts_map"] = job_acts_map
                    
                    # 5. 디버그 정보 출력 (개발용)
                    st.info(f"🔧 UI 수정사항 적용 완료:")
                    st.info(f"   - 활동: {valid_acts_df['activity'].tolist()}")
                    st.info(f"   - 방 타입: {list(room_template.keys())}")
                    st.info(f"   - 직무 매핑: {list(job_acts_map.columns) if not job_acts_map.empty else '없음'}")
                
                # 스케줄링 실행 전 최종 상태 확인 및 보호
                # UI 수정사항이 스케줄링에 반영되도록 강제 적용
                final_acts_df = st.session_state.get("activities", pd.DataFrame())
                if not final_acts_df.empty:
                    # 최종 활동 데이터를 스케줄링용으로 준비
                    st.session_state["activities"] = final_acts_df
                    
                    # 최종 방 템플릿 확인
                    final_room_types = final_acts_df[final_acts_df["use"] == True]["room_type"].unique()
                    final_room_template = {}
                    for rt in final_room_types:
                        if rt:
                            final_room_template[rt] = {"count": 1, "cap": 6}
                    
                    # 기존 설정과 병합
                    existing_template = st.session_state.get("room_template", {})
                    for rt, values in existing_template.items():
                        if rt in final_room_types:
                            final_room_template[rt] = values
                    
                    st.session_state["room_template"] = final_room_template
                    
                    # 최종 room_plan 생성
                    final_room_plan = {}
                    for rt, values in final_room_template.items():
                        final_room_plan[f"{rt}_count"] = values['count']
                        final_room_plan[f"{rt}_cap"] = values['cap']
                    
                    st.session_state["room_plan"] = pd.DataFrame([final_room_plan])
                
                # 스케줄링 모드에 따라 실행
                use_new_scheduler = "계층적" in scheduler_choice
                use_three_phase = "3단계" in scheduler_choice
                
                if use_new_scheduler and not use_three_phase:
                    # 계층적 스케줄러 v2 선택 시 자동으로 2단계 스케줄링 적용
                    st.info("🚀 계층적 스케줄러 v2로 2단계 하드 제약 스케줄링을 실행합니다...")
                    
                    # 2단계 하드 제약 스케줄링 실행
                    status, final_wide, logs, limit, reports = solve_for_days_two_phase(
                        cfg, params, debug=False, 
                        progress_callback=lambda info: (
                            progress_callback(info),
                            update_progress()
                        ),
                        percentile=90.0  # 기본값 90%
                    )
                    
                    # 2단계 스케줄링 결과 저장
                    st.session_state['two_phase_reports'] = reports
                    st.session_state['three_phase_reports'] = None
                    
                    # 상태 변환
                    if status == "SUCCESS":
                        status = "OK"
                    elif status in ["PARTIAL", "FAILED"]:
                        status = "FAILED"
                        
                elif use_three_phase:
                    # 3단계 스케줄링 실행
                    st.info("🚀 3단계 스케줄링 시스템을 실행합니다...")
                    st.info("1단계: 기본 스케줄링 → 2단계: 90% 백분위수 → 3단계: 2단계 결과의 90% 재조정")
                    
                    # 🔧 UI 데이터 구조 자동 변환
                    st.info("🔧 UI 데이터 구조를 스케줄러 형식으로 변환 중...")
                    
                    try:
                        # 활동 데이터 구조 변환
                        if 'activities' in st.session_state and st.session_state['activities'] is not None:
                            activities_df = st.session_state['activities']
                            
                            # 딕셔너리 형태인 경우 DataFrame으로 변환
                            if isinstance(activities_df, dict):
                                activities_df = pd.DataFrame(activities_df)
                            
                            # 컬럼명 매핑
                            column_mapping = {
                                'activity': 'activity',
                                'duration_min': 'duration',
                                'min_cap': 'group_size',
                                'max_cap': 'max_group_size',
                                'room_type': 'room_type',
                                'mode': 'mode',
                                'use': 'use'
                            }
                            
                            # 존재하는 컬럼만 매핑
                            existing_columns = {k: v for k, v in column_mapping.items() if k in activities_df.columns}
                            activities_df = activities_df.rename(columns=existing_columns)
                            
                            # use 컬럼을 boolean으로 변환
                            if 'use' in activities_df.columns:
                                activities_df['use'] = activities_df['use'].astype(bool)
                            
                            st.session_state['activities'] = activities_df
                            st.success(f"✅ 활동 데이터 변환 완료: {activities_df['activity'].tolist()}")
                        
                        # 직무 매핑 데이터 구조 변환
                        if 'job_acts_map' in st.session_state and st.session_state['job_acts_map'] is not None:
                            job_acts_df = st.session_state['job_acts_map']
                            
                            # 딕셔너리 형태인 경우 DataFrame으로 변환
                            if isinstance(job_acts_df, dict):
                                job_acts_df = pd.DataFrame(job_acts_df)
                            
                            # 컬럼명 매핑
                            job_column_mapping = {
                                'code': 'job_code',
                                'count': 'applicant_count'
                            }
                            
                            # 존재하는 컬럼만 매핑
                            existing_job_columns = {k: v for k, v in job_column_mapping.items() if k in job_acts_df.columns}
                            job_acts_df = job_acts_df.rename(columns=existing_job_columns)
                            
                            st.session_state['job_acts_map'] = job_acts_df
                            st.success(f"✅ 직무 매핑 변환 완료: {len(job_acts_df)}개 직무")
                        
                        st.success("✅ 모든 데이터 구조 변환 완료!")
                        
                    except Exception as e:
                        st.error(f"❌ 데이터 구조 변환 실패: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
                        status = "FAILED"
                        final_wide = None
                        logs = f"데이터 구조 변환 실패: {str(e)}"
                        limit = 0
                        st.session_state['three_phase_reports'] = None
                        st.session_state['two_phase_reports'] = None
                        st.session_state['final_schedule'] = None
                    
                    # 3단계 스케줄링 실행 (내부 테스트에서 구현한 로직 사용)
                    try:
                        st.info("🚀 3단계 스케줄링 시작...")
                        
                        # UI 설정을 매개변수로 전달
                        ui_config = {
                            "multidate_plans": st.session_state.get("multidate_plans", {}),
                            "activities": st.session_state.get("activities", pd.DataFrame()),
                            "room_template": st.session_state.get("room_template", {}),
                            "room_plan": st.session_state.get("room_plan", pd.DataFrame()),
                            "precedence": st.session_state.get("precedence", pd.DataFrame()),
                            "oper_start_time": st.session_state.get("oper_start_time", "09:00"),
                            "oper_end_time": st.session_state.get("oper_end_time", "17:30"),
                            "global_gap_min": st.session_state.get("global_gap_min", 5),
                            "max_stay_hours": st.session_state.get("max_stay_hours", 8),
                            "group_min_size": st.session_state.get("group_min_size", 4),
                            "group_max_size": st.session_state.get("group_max_size", 6)
                        }
                        
                        results = run_multi_date_scheduling(ui_config)
                        
                        if results and results['phase3']['status'] == "SUCCESS":
                            final_wide = results['phase3']['df']
                            status = "OK"
                            logs = "3단계 스케줄링 성공"
                            limit = len(final_wide) if not final_wide.empty else 0
                            
                            # 3단계 결과 저장
                            st.session_state['three_phase_reports'] = {
                                'phase1': results['phase1'],
                                'phase2': results['phase2'], 
                                'phase3': results['phase3']
                            }
                            st.session_state['two_phase_reports'] = None
                            
                            # 3단계 결과를 final_schedule에도 저장 (중요!)
                            st.session_state['final_schedule'] = final_wide
                            
                            st.success(f"🎉 3단계 스케줄링 성공! {len(final_wide)}개 스케줄 생성")
                            
                        else:
                            status = "FAILED"
                            final_wide = None
                            logs = "3단계 스케줄링 실패"
                            if results:
                                logs += f"\n상세 정보:\n"
                                logs += f"- 3단계 상태: {results['phase3']['status']}\n"
                                logs += f"- 3단계 로그: {results['phase3']['logs'][:500]}...\n"
                                if 'phase1' in results:
                                    logs += f"- 1단계 상태: {results['phase1']['status']}\n"
                                if 'phase2' in results:
                                    logs += f"- 2단계 상태: {results['phase2']['status']}\n"
                            limit = 0
                            st.session_state['three_phase_reports'] = None
                            st.session_state['two_phase_reports'] = None
                            st.session_state['final_schedule'] = None
                            
                            st.error("❌ 3단계 스케줄링 실패")
                            st.code(logs)
                            
                            
                            # 추가 디버깅 정보
                            st.error("🔍 추가 디버깅 정보:")
                            st.write(f"- UI 설정 전달 여부: {ui_config is not None}")
                            st.write(f"- 활동 데이터 존재: {not ui_config['activities'].empty if ui_config else False}")
                            st.write(f"- 멀티데이트 계획 존재: {len(ui_config['multidate_plans']) if ui_config else 0}")
                            st.write(f"- 방 템플릿 존재: {len(ui_config['room_template']) if ui_config else 0}")
                            
                            # 세션 상태 확인
                            st.error("🔍 세션 상태 확인:")
                            st.write(f"- 세션 활동 데이터: {st.session_state.get('activities') is not None}")
                            st.write(f"- 세션 멀티데이트 계획: {st.session_state.get('multidate_plans') is not None}")
                            st.write(f"- 세션 방 템플릿: {st.session_state.get('room_template') is not None}")
                    except Exception as e:
                        status = "FAILED"
                        final_wide = None
                        logs = f"3단계 스케줄링 실행 중 오류: {str(e)}"
                        import traceback
                        logs += f"\n상세 오류:\n{traceback.format_exc()}"
                        limit = 0
                        st.session_state['three_phase_reports'] = None
                        st.session_state['two_phase_reports'] = None
                        st.session_state['final_schedule'] = None
                        
                        st.error("❌ 3단계 스케줄링 실행 중 오류")
                        st.code(logs)
                elif "OR-Tools" in scheduler_choice:
                    # OR-Tools 스케줄러 선택 시
                    st.info("📊 OR-Tools 스케줄러로 실행 중...")
                    if has_batched:
                        st.warning("⚠️ OR-Tools 스케줄러는 Batched 활동을 완전히 지원하지 않을 수 있습니다.")
                    status, final_wide, logs, limit = solve_for_days(cfg, params, debug=False)
                    st.session_state['two_phase_reports'] = None
                    st.session_state['three_phase_reports'] = None
                else:
                    status = "FAILED"
                    final_wide = None
                    logs = "스케줄러 선택 오류"
                    limit = 0
                    st.session_state['three_phase_reports'] = None
                    st.session_state['two_phase_reports'] = None
                    st.session_state['final_schedule'] = None
                
                st.session_state['last_solve_logs'] = logs
                st.session_state['solver_status'] = status
                st.session_state['daily_limit'] = limit
                
                if status == "OK" and final_wide is not None and not final_wide.empty:
                    st.session_state['final_schedule'] = final_wide
                    st.balloons()
                else:
                    st.session_state['final_schedule'] = None
            except Exception as e:
                st.error(f"계산 중 오류가 발생했습니다: {str(e)}")
                st.session_state['solver_status'] = "ERROR"

# 결과 표시
st.markdown("---")
status = st.session_state.get('solver_status', '미실행')
daily_limit = st.session_state.get('daily_limit', 0)

col1, col2 = st.columns(2)
with col1:
    st.info(f"Solver Status: `{status}`")
with col2:
    if daily_limit > 0:
        st.info(f"계산된 일일 최대 처리 인원: **{daily_limit}명**")

if "솔버 시간 초과" in st.session_state.get('last_solve_logs', ''):
    st.warning("⚠️ 연산 시간이 2분(120초)을 초과하여, 현재까지 찾은 최적의 스케줄을 반환했습니다. 결과는 최상이 아닐 수 있습니다.")

# 결과 출력
final_schedule = st.session_state.get('final_schedule')
if final_schedule is not None and not final_schedule.empty:
    st.success("🎉 운영일정 추정이 완료되었습니다!")
    
    # 3단계 스케줄링 결과 표시 (우선순위)
    if st.session_state.get('three_phase_reports'):
        st.subheader("🔧 3단계 하드 제약 스케줄링 결과")
        
        three_phase_reports = st.session_state['three_phase_reports']
        
        # 3단계 결과 요약
        col1, col2, col3 = st.columns(3)
        with col1:
            phase1_count = len(three_phase_reports['phase1']['df']) if three_phase_reports['phase1']['df'] is not None else 0
            st.info(f"📊 **1단계 스케줄**: {phase1_count}개")
        with col2:
            phase2_count = len(three_phase_reports['phase2']['df']) if three_phase_reports['phase2']['df'] is not None else 0
            st.info(f"📊 **2단계 스케줄**: {phase2_count}개")
        with col3:
            phase3_count = len(three_phase_reports['phase3']['df']) if three_phase_reports['phase3']['df'] is not None else 0
            st.info(f"📊 **3단계 스케줄**: {phase3_count}개")
        
        # 체류시간 개선 효과 표시
        if (three_phase_reports['phase1']['df'] is not None and 
            three_phase_reports['phase2']['df'] is not None and 
            three_phase_reports['phase3']['df'] is not None):
            
            # 각 단계별 최대 체류시간 계산
            def calculate_max_stay_time(df):
                if df.empty:
                    return 0
                df_temp = df.copy()
                df_temp['interview_date'] = pd.to_datetime(df_temp['interview_date'])
                max_stay = 0
                for date_str in df_temp['interview_date'].dt.strftime('%Y-%m-%d').unique():
                    date_df = df_temp[df_temp['interview_date'].dt.strftime('%Y-%m-%d') == date_str]
                    for applicant_id in date_df['applicant_id'].unique():
                        applicant_df = date_df[date_df['applicant_id'] == applicant_id]
                        start_time = applicant_df['start_time'].min()
                        end_time = applicant_df['end_time'].max()
                        stay_hours = (end_time - start_time).total_seconds() / 3600
                        max_stay = max(max_stay, stay_hours)
                return max_stay
            
            phase1_max = calculate_max_stay_time(three_phase_reports['phase1']['df'])
            phase2_max = calculate_max_stay_time(three_phase_reports['phase2']['df'])
            phase3_max = calculate_max_stay_time(three_phase_reports['phase3']['df'])
            
            st.markdown("**📈 3단계 체류시간 개선 효과**")
            col1, col2, col3 = st.columns(3)
            with col1:
                improvement_2 = phase1_max - phase2_max
                improvement_2_pct = (improvement_2 / phase1_max * 100) if phase1_max > 0 else 0
                st.success(f"**1단계 → 2단계**: {improvement_2:.2f}시간 ({improvement_2_pct:.1f}%)")
            with col2:
                improvement_3 = phase1_max - phase3_max
                improvement_3_pct = (improvement_3 / phase1_max * 100) if phase1_max > 0 else 0
                st.success(f"**1단계 → 3단계**: {improvement_3:.2f}시간 ({improvement_3_pct:.1f}%)")
            with col3:
                additional_improvement = phase2_max - phase3_max
                additional_pct = (additional_improvement / phase1_max * 100) if phase1_max > 0 else 0
                st.success(f"**2단계 → 3단계**: {additional_improvement:.2f}시간 ({additional_pct:.1f}%)")
            
            # 상세 비교 테이블
            comparison_data = {
                '단계': ['1단계 (기본)', '2단계 (90% 백분위수)', '3단계 (2단계 90% 재조정)'],
                '최대 체류시간': [f"{phase1_max:.2f}시간", f"{phase2_max:.2f}시간", f"{phase3_max:.2f}시간"],
                '개선 효과': ['-', f"{improvement_2:.2f}시간", f"{improvement_3:.2f}시간"],
                '개선률': ['-', f"{improvement_2_pct:.1f}%", f"{improvement_3_pct:.1f}%"]
            }
            comparison_df = pd.DataFrame(comparison_data)
            st.dataframe(comparison_df, use_container_width=True, hide_index=True)
        
        st.success("✅ 3단계 스케줄링이 성공적으로 완료되었습니다!")
    
    # 2단계 스케줄링 결과 표시 (3단계가 없을 때만)
    elif "계층적" in scheduler_choice and st.session_state.get('two_phase_reports'):
        st.subheader("🔧 2단계 하드 제약 스케줄링 결과")
        
        reports = st.session_state['two_phase_reports']
        
        # 하드 제약 분석 결과 표시
        if 'constraint_analysis' in reports and not reports['constraint_analysis'].empty:
            constraint_df = reports['constraint_analysis']
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.info(f"📊 **분석된 날짜**: {len(constraint_df)}일")
            with col2:
                total_candidates = constraint_df['applicant_count'].sum()
                st.info(f"👥 **총 지원자**: {total_candidates}명")
            with col3:
                avg_constraint = constraint_df['hard_constraint_hours'].mean()
                st.info(f"⏰ **평균 하드 제약**: {avg_constraint:.1f}시간")
            
            # 하드 제약 분석 테이블
            st.markdown("**📋 날짜별 하드 제약 분석**")
            display_constraint = constraint_df.copy()
            display_constraint['hard_constraint_hours'] = display_constraint['hard_constraint_hours'].round(2)
            if 'percentile' in display_constraint.columns:
                display_constraint['percentile'] = display_constraint['percentile'].round(1)
            else:
                display_constraint['percentile'] = 90.0
            # 컬럼 자동 한글화 및 선택
            col_map = {
                'interview_date': '날짜',
                'applicant_count': '지원자수',
                'mean_stay_hours': '평균체류시간(h)',
                'max_stay_hours': '최대체류시간(h)',
                'percentile': '분위수(%)',
                'hard_constraint_hours': '하드제약(h)',
                'exceed_count': '위반자수',
                'exceed_rate': '위반율(%)'
            }
            display_cols = [c for c in col_map if c in display_constraint.columns]
            display_constraint = display_constraint[display_cols].rename(columns=col_map)
            st.dataframe(display_constraint, use_container_width=True, hide_index=True)
        
        # 제약 위반 분석 결과 표시
        if 'constraint_violations' in reports and not reports['constraint_violations'].empty:
            violations_df = reports['constraint_violations']
            st.markdown("**⚠️ 제약 위반 분석**")
            st.dataframe(violations_df, use_container_width=True, hide_index=True)
        else:
            st.success("✅ 모든 지원자가 하드 제약 내에서 성공적으로 스케줄링되었습니다!")
        
        # 단계별 비교 결과 표시
        if 'phase_comparison' in reports and not reports['phase_comparison'].empty:
            comparison_df = reports['phase_comparison']
            st.markdown("**📈 1단계 vs 2단계 비교**")
            st.dataframe(comparison_df, use_container_width=True, hide_index=True)
    
    # 요약 정보
    total_candidates = len(final_schedule)
    total_days = final_schedule['interview_date'].nunique()
    selected_dates = st.session_state.get("interview_dates", [])
    
    if len(selected_dates) > 1:
        date_range_str = f"{selected_dates[0].strftime('%m/%d')} ~ {selected_dates[-1].strftime('%m/%d')}"
        st.info(f"총 {total_candidates}명의 지원자를 {total_days}일에 걸쳐 면접 진행 ({date_range_str})")
    else:
        st.info(f"총 {total_candidates}명의 지원자를 {total_days}일에 걸쳐 면접 진행")
    
    # 체류시간 분석 추가
    st.subheader("⏱️ 체류시간 분석")
    
    def calculate_stay_duration_stats(schedule_df):
        """각 지원자의 체류시간을 계산하고 통계를 반환"""
        stats_data = []
        
        # 컬럼명 매핑 (실제 데이터에 맞게 조정)
        id_col = None
        for col in ['applicant_id', 'id', 'candidate_id']:
            if col in schedule_df.columns:
                id_col = col
                break
        
        job_col = None
        for col in ['job_code', 'code']:
            if col in schedule_df.columns:
                job_col = col
                break
        
        date_col = None
        for col in ['interview_date', 'date']:
            if col in schedule_df.columns:
                date_col = col
                break
        
        if not id_col or not job_col or not date_col:
            st.error(f"필요한 컬럼을 찾을 수 없습니다. 현재 컬럼: {list(schedule_df.columns)}")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        # 지원자별 체류시간 계산
        for candidate_id in schedule_df[id_col].unique():
            candidate_data = schedule_df[schedule_df[id_col] == candidate_id]
            
            # 더미 데이터 제외
            if str(candidate_id).startswith('dummy'):
                continue
            
            if len(candidate_data) == 0:
                continue
            
            # 시간 파싱
            all_start_times = []
            all_end_times = []
            
            for _, row in candidate_data.iterrows():
                try:
                    start_time = row['start_time']
                    end_time = row['end_time']
                    
                    # 시간 형식 변환
                    if isinstance(start_time, str):
                        start_time = pd.to_datetime(start_time, format='%H:%M:%S').time()
                    if isinstance(end_time, str):
                        end_time = pd.to_datetime(end_time, format='%H:%M:%S').time()
                    
                    # timedelta로 변환
                    start_td = timedelta(hours=start_time.hour, minutes=start_time.minute, seconds=start_time.second)
                    end_td = timedelta(hours=end_time.hour, minutes=end_time.minute, seconds=end_time.second)
                    
                    all_start_times.append(start_td)
                    all_end_times.append(end_td)
                    
                except Exception as e:
                    continue
            
            if all_start_times and all_end_times:
                # 전체 체류시간 = 첫 번째 활동 시작 ~ 마지막 활동 종료
                total_start = min(all_start_times)
                total_end = max(all_end_times)
                stay_duration_hours = (total_end - total_start).total_seconds() / 3600
                
                # 직무 코드 및 날짜 (첫 번째 행에서 가져오기)
                job_code = candidate_data.iloc[0].get(job_col, 'Unknown')
                interview_date = candidate_data.iloc[0].get(date_col, 'Unknown')
                
                stats_data.append({
                    'candidate_id': candidate_id,
                    'job_code': job_code,
                    'interview_date': interview_date,
                    'stay_duration_hours': stay_duration_hours,
                    'start_time': total_start,
                    'end_time': total_end
                })
        
        if not stats_data:
            st.warning("체류시간을 계산할 수 있는 유효한 데이터가 없습니다.")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        stats_df = pd.DataFrame(stats_data)
        
        # 직무별 통계 계산
        job_stats = []
        for job_code, job_data in stats_df.groupby('job_code'):
            durations = job_data['stay_duration_hours']
            job_stats.append({
                'job_code': job_code,
                'count': len(job_data),
                'min_duration': durations.min(),
                'max_duration': durations.max(),
                'avg_duration': durations.mean(),
                'median_duration': durations.median()
            })
        
        # 날짜별 통계 계산
        date_stats = []
        for date, date_data in stats_df.groupby('interview_date'):
            durations = date_data['stay_duration_hours']
            max_stay_candidate = date_data.loc[date_data['stay_duration_hours'].idxmax()]
            
            date_stats.append({
                'interview_date': date,
                'count': len(date_data),
                'min_duration': durations.min(),
                'max_duration': durations.max(),
                'avg_duration': durations.mean(),
                'max_stay_candidate': max_stay_candidate['candidate_id'],
                'max_stay_job': max_stay_candidate['job_code']
            })
        
        return pd.DataFrame(job_stats), stats_df, pd.DataFrame(date_stats)
    
    # 3단계 결과가 있으면 3단계 결과를 사용, 없으면 기본 결과 사용
    analysis_df = None
    if st.session_state.get('three_phase_reports'):
        # 3단계 결과 사용
        three_phase_reports = st.session_state['three_phase_reports']
        if three_phase_reports['phase3']['df'] is not None:
            analysis_df = three_phase_reports['phase3']['df']
            st.info("📊 **3단계 스케줄링 결과**를 기준으로 체류시간을 분석합니다.")
    else:
        # 기본 결과 사용
        analysis_df = final_schedule
        st.info("📊 **기본 스케줄링 결과**를 기준으로 체류시간을 분석합니다.")
    
    if analysis_df is None or analysis_df.empty:
        st.warning("⚠️ 분석할 스케줄 데이터가 없습니다.")
        st.stop()
    
    try:
        job_stats_df, individual_stats_df, date_stats_df = calculate_stay_duration_stats(analysis_df)
        
        # 디버깅 정보 출력
        st.write(f"🔍 **디버깅 정보**:")
        st.write(f"- job_stats_df 크기: {len(job_stats_df) if not job_stats_df.empty else 0}")
        st.write(f"- individual_stats_df 크기: {len(individual_stats_df) if not individual_stats_df.empty else 0}")
        st.write(f"- date_stats_df 크기: {len(date_stats_df) if not date_stats_df.empty else 0}")
        
        # 날짜별 통계 먼저 표시
        if not date_stats_df.empty:
            st.markdown("**📅 날짜별 체류시간 통계**")
            
            # 표시용 데이터프레임 생성
            display_date_stats = date_stats_df.copy()
            display_date_stats['min_duration'] = display_date_stats['min_duration'].round(1)
            display_date_stats['max_duration'] = display_date_stats['max_duration'].round(1)
            display_date_stats['avg_duration'] = display_date_stats['avg_duration'].round(1)
            
            # 최대 체류시간 지원자 정보 포함
            display_date_stats['max_info'] = display_date_stats.apply(
                lambda row: f"{row['max_stay_candidate']} ({row['max_stay_job']})", axis=1
            )
            
            # 컬럼 선택 및 한글화
            display_columns = ['interview_date', 'count', 'min_duration', 'max_duration', 'avg_duration', 'max_info']
            display_date_stats = display_date_stats[display_columns]
            display_date_stats.columns = ['면접일자', '응시자수', '최소시간(h)', '최대시간(h)', '평균시간(h)', '최대체류자(직무)']
            
            st.dataframe(display_date_stats, use_container_width=True)
            
            # 전체 요약 지표
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                total_candidates = date_stats_df['count'].sum()
                st.metric("전체 응시자 수", f"{total_candidates}명")
            with col2:
                overall_min = date_stats_df['min_duration'].min()
                st.metric("전체 최소 체류시간", f"{overall_min:.1f}시간")
            with col3:
                overall_max = date_stats_df['max_duration'].max()
                st.metric("전체 최대 체류시간", f"{overall_max:.1f}시간")
            with col4:
                overall_avg = (date_stats_df['avg_duration'] * date_stats_df['count']).sum() / date_stats_df['count'].sum()
                st.metric("전체 평균 체류시간", f"{overall_avg:.1f}시간")
            
            # 최대 체류시간 지원자 강조
            max_candidate_row = date_stats_df.loc[date_stats_df['max_duration'].idxmax()]
            st.info(f"🔥 **최대 체류시간**: {max_candidate_row['max_stay_candidate']} ({max_candidate_row['max_stay_job']}) - "
                   f"{max_candidate_row['max_duration']:.1f}시간 ({max_candidate_row['interview_date']})")
        else:
            st.warning("⚠️ 날짜별 체류시간 데이터가 없습니다.")
        
        # 직무별 통계 표시
        if not job_stats_df.empty:
            st.markdown("**👥 직무별 체류시간 통계**")
            
            # 표시용 데이터프레임 생성
            display_job_stats = job_stats_df.copy()
            display_job_stats['min_duration'] = display_job_stats['min_duration'].round(1)
            display_job_stats['max_duration'] = display_job_stats['max_duration'].round(1)
            display_job_stats['avg_duration'] = display_job_stats['avg_duration'].round(1)
            display_job_stats['median_duration'] = display_job_stats['median_duration'].round(1)
            
            # 컬럼명 한글화
            display_job_stats.columns = ['직무코드', '인원수', '최소시간(h)', '최대시간(h)', '평균시간(h)', '중간값(h)']
            
            st.dataframe(display_job_stats, use_container_width=True)
            
            # 체류시간 제한 확인
            max_stay_hours = params.get('max_stay_hours', 8)
            if not date_stats_df.empty and date_stats_df['max_duration'].max() > max_stay_hours:
                st.warning(f"⚠️ 일부 지원자의 체류시간이 설정된 제한({max_stay_hours}시간)을 초과했습니다.")
            
            # Level 4 후처리 조정 효과 표시
            if params.get('enable_level4_optimization', False):
                st.success("✅ Level 4 후처리 조정이 적용되었습니다.")
                
                # 동적 임계값 계산 (Level 4 로직과 동일)
                if not individual_stats_df.empty:
                    stay_times = individual_stats_df['stay_duration_hours'].tolist()
                    if stay_times:
                        import statistics
                        mean_stay = statistics.mean(stay_times)
                        std_dev = statistics.stdev(stay_times) if len(stay_times) > 1 else 0
                        sorted_times = sorted(stay_times, reverse=True)
                        percentile_30_index = int(len(sorted_times) * 0.3)
                        percentile_30_value = sorted_times[min(percentile_30_index, len(sorted_times) - 1)]
                        
                        statistical_threshold = mean_stay + 0.5 * std_dev  # 더 공격적 (기존 1.0 → 0.5)
                        dynamic_threshold = max(3.0, min(statistical_threshold, percentile_30_value))
                        
                        problem_cases = len([t for t in stay_times if t >= dynamic_threshold])
                        
                        st.info(f"📊 **Level 4 동적 임계값 분석**: 평균 {mean_stay:.1f}h, 표준편차 {std_dev:.1f}h, "
                               f"상위30% {percentile_30_value:.1f}h, 통계적임계값 {statistical_threshold:.1f}h → "
                               f"**최종 동적임계값 {dynamic_threshold:.1f}h** (문제케이스 {problem_cases}개)")
            else:
                st.info("ℹ️ Level 4 후처리 조정이 비활성화되어 있습니다.")
        else:
            st.warning("⚠️ 직무별 체류시간 데이터가 없습니다.")
    
    except Exception as e:
        st.error(f"체류시간 분석 중 오류 발생: {str(e)}")
    
    # 스케줄 표시 옵션
    col1, col2 = st.columns([3, 1])
    with col1:
        # 날짜별 요약 정보
        date_summary = final_schedule.groupby('interview_date').size().reset_index(name='인원수')
        date_summary['interview_date'] = pd.to_datetime(date_summary['interview_date']).dt.strftime('%Y-%m-%d')
        date_summary.columns = ['날짜', '인원수']
        
        st.markdown("**📅 날짜별 면접 인원**")
        st.dataframe(date_summary, use_container_width=True, hide_index=True)
    
    with col2:
        st.markdown("**💾 결과 다운로드**")
        excel_buffer = BytesIO()
        df_to_excel(final_schedule, excel_buffer)
        excel_buffer.seek(0)
        
        st.download_button(
            label="📥 Excel 다운로드",
            data=excel_buffer,
            file_name=f"interview_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    # 상세 스케줄 표시
    with st.expander("📋 상세 스케줄 보기", expanded=False):
        # 날짜별로 탭 생성
        dates = sorted(final_schedule['interview_date'].unique())
        tabs = st.tabs([pd.to_datetime(d).strftime('%Y-%m-%d') for d in dates])
        
        for i, (tab, date) in enumerate(zip(tabs, dates)):
            with tab:
                day_schedule = final_schedule[final_schedule['interview_date'] == date].copy()
                
                # 시간 컬럼들을 더 읽기 쉽게 표시
                display_cols = []
                
                # ID 컬럼 찾기
                for id_col in ['applicant_id', 'id', 'candidate_id']:
                    if id_col in day_schedule.columns:
                        display_cols.append(id_col)
                        break
                
                # 직무 코드 컬럼 찾기
                for job_col in ['job_code', 'code']:
                    if job_col in day_schedule.columns:
                        display_cols.append(job_col)
                        break
                
                # 기타 중요 컬럼들 추가
                for col in day_schedule.columns:
                    if col not in display_cols and (
                        col.startswith(('start_', 'end_', 'loc_')) or 
                        col in ['activity_name', 'room_name', 'duration_min', 'group_number', 'group_size']
                    ):
                        display_cols.append(col)
                
                # 실제 존재하는 컬럼만 필터링
                display_cols = [col for col in display_cols if col in day_schedule.columns]
                
                if display_cols:
                    day_schedule = day_schedule[display_cols]
                
                # 인덱스 숨기고 표시
                st.dataframe(day_schedule, use_container_width=True, hide_index=True)
                
                # batched 모드가 있으면 그룹 정보도 표시
                if has_batched:
                    # TODO: 그룹 정보 표시 구현
                    pass

st.divider()

# =============================================================================
# 섹션 1: 면접활동 정의
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("1️⃣ 면접활동 정의")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_activities", help="활동 정의 AG-Grid가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "activities" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["activities"] = 0
        st.session_state["section_refresh_counter"]["activities"] += 1
        st.rerun()

st.markdown("면접 프로세스에 포함될 활동들을 정의합니다. 각 활동의 모드, 소요시간, 필요한 공간 등을 설정하세요.")

# 기본 템플릿 함수
def default_df() -> pd.DataFrame:
    return pd.DataFrame({
        "use": [True, True, True, True],
        "activity": ["면접1", "면접2", "인성검사", "커피챗"],
        "mode": ["individual"] * 4,
        "duration_min": [10] * 4,
        "room_type": ["면접1실", "면접2실", "인성검사실", "커피챗실"],
        "min_cap": [1] * 4,
        "max_cap": [1] * 4,
    })

df = st.session_state["activities"].copy()

# 누락 컬럼 보강
for col, default_val in {
    "use": True,
    "mode": "individual",
    "duration_min": 10,
    "room_type": "",
    "min_cap": 1,
    "max_cap": 1,
}.items():
    if col not in df.columns:
        df[col] = default_val

# AG-Grid 설정
gb = GridOptionsBuilder.from_dataframe(df)

gb.configure_column(
    "use",
    header_name="사용",
    cellEditor="agCheckboxCellEditor",
    cellRenderer="agCheckboxCellRenderer",
    editable=True,
    singleClickEdit=True,
    width=80,
)

gb.configure_column("activity", header_name="활동 이름", editable=True)

# mode_values에 parallel과 batched 추가
mode_values = ["individual", "parallel", "batched"]
gb.configure_column(
    "mode",
    header_name="모드",
    editable=True,
    cellEditor="agSelectCellEditor",
    cellEditorParams={"values": mode_values},
    width=110,
)

for col, hdr in [("duration_min", "소요시간(분)"), ("min_cap", "최소 인원"), ("max_cap", "최대 인원")]:
    gb.configure_column(
        col,
        header_name=hdr,
        editable=True,
        type=["numericColumn", "numberColumnFilter"],
        width=120,
    )

gb.configure_column("room_type", header_name="면접실 이름", editable=True)

grid_opts = gb.build()

st.markdown("#### 활동 정의")

# 모드 설명 추가
with st.expander("ℹ️ 모드 설명", expanded=False):
    st.markdown("""
    - **individual**: 1명이 혼자 면접 (기존 방식)
    - **parallel**: 여러명이 같은 공간에서 각자 다른 활동 (예: 개별 작업)
    - **batched**: 여러명이 동시에 같은 활동 (예: 그룹토론, PT발표)
    
    **주의**: batched 모드를 사용하는 모든 활동의 min_cap, max_cap은 동일해야 합니다.
    """)

# 행 추가/삭제 기능 (위로 이동)
col_add, col_del = st.columns(2)

with col_add:
    if st.button("➕ 활동 행 추가", key="add_activity"):
        new_row = {
            "use": True,
            "activity": "NEW_ACT",
            "mode": "individual",
            "duration_min": 10,
            "room_type": "",
            "min_cap": 1,
            "max_cap": 1,
        }
        st.session_state["activities"] = pd.concat(
            [st.session_state["activities"], pd.DataFrame([new_row])],
            ignore_index=True,
        )
        st.rerun()

with col_del:
    act_df = st.session_state["activities"].copy()
    if not act_df.empty:
        # 인덱스와 활동명을 안전하게 결합
        delete_options = []
        valid_indices = []
        for idx, row in act_df.iterrows():
            activity_name = str(row.get('activity', 'Unknown'))
            if activity_name and activity_name != 'nan':
                delete_options.append(f"{idx}: {activity_name}")
                valid_indices.append(idx)
        
        to_delete = st.multiselect(
            "삭제할 활동 선택",
            options=delete_options,
            key="del_activity_select"
        )
        if st.button("❌ 선택된 활동 삭제", key="del_activity"):
            if to_delete:
                # 선택된 인덱스 추출
                selected_indices = [int(s.split(":")[0]) for s in to_delete]
                
                # 실제 DataFrame에 존재하는 인덱스만 필터링
                valid_to_drop = [idx for idx in selected_indices if idx in act_df.index]
                
                if valid_to_drop:
                    kept = st.session_state["activities"].drop(valid_to_drop).reset_index(drop=True)
                    st.session_state["activities"] = kept
                    st.success(f"선택된 {len(valid_to_drop)}개 활동이 삭제되었습니다.")
                    st.rerun()
                else:
                    st.error("삭제할 유효한 활동이 없습니다.")
    else:
        st.info("삭제할 활동이 없습니다.")

# AG-Grid 표시 (행 추가/삭제 기능 아래로 이동)
# 섹션별 새로고침을 위한 동적 key 생성
activities_refresh_count = st.session_state.get("section_refresh_counter", {}).get("activities", 0)

grid_ret = AgGrid(
    df,
    gridOptions=grid_opts,
    data_return_mode=DataReturnMode.AS_INPUT,
    update_mode=GridUpdateMode.VALUE_CHANGED,
    allow_unsafe_jscode=True,
    fit_columns_on_grid_load=True,
    theme="balham",
    key=f"activities_grid_{activities_refresh_count}",  # 동적 key로 강제 재렌더링
)

st.session_state["activities"] = grid_ret["data"]

# batched 모드 일관성 검증
batched_acts = st.session_state["activities"][st.session_state["activities"]["mode"] == "batched"]
if not batched_acts.empty:
    min_caps = batched_acts["min_cap"].unique()
    max_caps = batched_acts["max_cap"].unique()
    
    if len(min_caps) > 1 or len(max_caps) > 1:
        st.error("⚠️ 모든 batched 활동의 그룹 크기(min_cap, max_cap)는 동일해야 합니다!")
        st.info("현재 설정: " + 
                f"min_cap = {list(min_caps)}, " +
                f"max_cap = {list(max_caps)}")

st.divider()

# =============================================================================
# 섹션 2: 선후행 제약 설정 (면접 활동 정의 바로 다음으로 이동)
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("2️⃣ 선후행 제약 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_precedence", help="선후행 제약 UI가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "precedence" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["precedence"] = 0
        st.session_state["section_refresh_counter"]["precedence"] += 1
        st.rerun()

st.markdown("면접 활동 간의 순서 제약과 시간 간격을 설정합니다.")

# 공통 데이터 로드
acts_df = st.session_state.get("activities", pd.DataFrame())
jobs_df = st.session_state.get("job_acts_map", pd.DataFrame())

if not acts_df.empty:
    ACT_OPTS = acts_df.query("use == True")["activity"].tolist()
    
    # precedence 규칙 초기화 (빈 문자열도 허용)
    prec_df = st.session_state["precedence"].copy()
    valid_acts = set(ACT_OPTS) | {"__START__", "__END__", ""}
    prec_df = prec_df[prec_df["predecessor"].isin(valid_acts) & prec_df["successor"].isin(valid_acts)]
    st.session_state["precedence"] = prec_df
    
    # 동선 미리보기 함수
    def generate_flow_preview():
        """현재 선후행 제약 규칙을 바탕으로 가능한 동선을 시각화"""
        prec_rules = st.session_state["precedence"].copy()
        if prec_rules.empty:
            return "📋 설정된 제약 규칙이 없습니다. 모든 활동이 자유롭게 배치 가능합니다."
        
        # START와 END 규칙 분리
        start_rules = prec_rules[prec_rules["predecessor"] == "__START__"]
        end_rules = prec_rules[prec_rules["successor"] == "__END__"]
        middle_rules = prec_rules[
            (~prec_rules["predecessor"].isin(["__START__", "__END__"])) &
            (~prec_rules["successor"].isin(["__START__", "__END__"]))
        ]
        
        flow_text = "🔄 **예상 면접 동선:**\n\n"
        
        # START 규칙이 있는 경우
        if not start_rules.empty:
            first_acts = start_rules["successor"].tolist()
            flow_text += f"**시작:** 🏁 → {' / '.join(first_acts)}\n\n"
        else:
            flow_text += "**시작:** 🏁 → (모든 활동 가능)\n\n"
        
        # 중간 규칙들
        if not middle_rules.empty:
            flow_text += "**중간 연결:**\n"
            for _, rule in middle_rules.iterrows():
                gap_info = f" ({rule['gap_min']}분 간격)" if rule['gap_min'] > 0 else ""
                adj_info = " [인접]" if rule['adjacent'] else ""
                flow_text += f"• {rule['predecessor']} → {rule['successor']}{gap_info}{adj_info}\n"
            flow_text += "\n"
        
        # END 규칙이 있는 경우
        if not end_rules.empty:
            last_acts = end_rules["predecessor"].tolist()
            flow_text += f"**종료:** {' / '.join(last_acts)} → 🏁"
        else:
            flow_text += "**종료:** (모든 활동 가능) → 🏁"
        
        return flow_text
    
    # 가능한 모든 활동 순서 계산 함수 (과거 코드 복원)
    def render_dynamic_flows(prec_df: pd.DataFrame, base_nodes: list[str]) -> list[str]:
        """
        prec_df: ['predecessor','successor','gap_min','adjacent']
        base_nodes: 순서에 포함될 활동 리스트
        """
        import itertools
        
        # 1) 규칙(rules)에 adjacent까지 함께 읽어오기
        rules = [
            (row.predecessor, row.successor,
             int(row.gap_min),
             bool(getattr(row, "adjacent", False)))
            for row in prec_df.itertuples()
        ]
        n = len(base_nodes)
        valid_orders = []

        # 2) 인터뷰 느낌 이모지 풀(10개) + 동적 매핑
        emoji_pool = ["📝","🧑‍💼","🎤","💼","🗣️","🤝","🎯","🔎","📋","⏰"]
        icons = { act: emoji_pool[i % len(emoji_pool)]
                  for i, act in enumerate(base_nodes) }

        # 3) 모든 순열 검사
        for perm in itertools.permutations(base_nodes, n):
            ok = True
            for p, s, gap, adj in rules:
                # START → S
                if p == "__START__":
                    if perm[0] != s:
                        ok = False
                    if not ok: break
                    else: continue
                # P → END
                if s == "__END__":
                    if perm[-1] != p:
                        ok = False
                    if not ok: break
                    else: continue
                # 일반 활동 간 제약
                if p in perm and s in perm:
                    i_p, i_s = perm.index(p), perm.index(s)
                    # 붙이기(adjacent) 또는 gap>0 모두 "인접" 처리
                    if adj or gap > 0:
                        if i_s != i_p + 1:
                            ok = False
                            break
                    else:
                        # gap==0: 순서만 보장
                        if i_p >= i_s:
                            ok = False
                            break
            if ok:
                valid_orders.append(perm)

        # 4) 문자열로 변환 (아이콘 + 활동명)
        flow_strs = []
        for order in valid_orders:
            labels = [f"{icons[act]} {act}" for act in order]
            flow_strs.append(" ➔ ".join(labels))
        return flow_strs
    
    # 실시간 동선(활동 순서) 미리보기
    with st.expander("🔍 실시간 동선(활동 순서) 미리보기", expanded=True):
        prec_df_latest = st.session_state["precedence"]
        
        # 기본 규칙 표시
        st.markdown("**📋 현재 설정된 선후행 제약 규칙:**")
        if prec_df_latest.empty:
            st.info("설정된 제약 규칙이 없습니다. 모든 활동이 자유롭게 배치 가능합니다.")
        else:
            # 규칙을 사용자 친화적으로 표시
            for _, rule in prec_df_latest.iterrows():
                pred = rule['predecessor']
                succ = rule['successor']
                gap = rule['gap_min']
                adj = rule['adjacent']
                
                # 표시 형식 개선
                if pred == "__START__":
                    pred_display = "🏁 시작"
                elif pred == "":
                    pred_display = "🏁 시작"
                else:
                    pred_display = f"📝 {pred}"
                
                if succ == "__END__":
                    succ_display = "🏁 종료"
                elif succ == "":
                    succ_display = "🏁 종료"
                else:
                    succ_display = f"📝 {succ}"
                
                gap_info = f" ({gap}분 간격)" if gap > 0 else ""
                adj_info = " [인접 배치]" if adj else ""
                
                st.markdown(f"• {pred_display} → {succ_display}{gap_info}{adj_info}")
        
        st.markdown("---")
        
        # 가능한 활동 순서 계산 및 표시
        if ACT_OPTS:
            flows = render_dynamic_flows(prec_df_latest, ACT_OPTS)
            if not flows:
                st.warning("⚠️ 현재 제약을 만족하는 활동 순서가 없습니다. 제약 조건을 확인해주세요.")
            else:
                st.markdown("**🔄 가능한 모든 활동 순서:**")
                for i, f in enumerate(flows, 1):
                    st.markdown(f"{i}. {f}")
                
                if len(flows) == 1:
                    st.success("✅ 제약 조건에 따라 활동 순서가 고유하게 결정됩니다!")
                else:
                    st.info(f"💡 총 {len(flows)}가지 가능한 활동 순서가 있습니다.")
        else:
            st.warning("활성화된 활동이 없습니다. 먼저 활동을 정의해주세요.")

    
    # ═══════════════════════════════════════════
    # 🎯 면접 순서 설정 (단계별 가이드)
    # ═══════════════════════════════════════════

    # 현재 설정된 START/END 규칙 확인
    current_start = None
    current_end = None
    for _, rule in prec_df.iterrows():
        if rule['predecessor'] == "__START__":
            current_start = rule['successor']
        if rule['successor'] == "__END__":
            current_end = rule['predecessor']
    
    # 섹션별 새로고침을 위한 동적 key 생성
    precedence_refresh_count = st.session_state.get("section_refresh_counter", {}).get("precedence", 0)
    
    st.markdown("---")
    st.subheader("🎯 면접 순서 설정")
    st.markdown("면접 활동들의 순서와 제약을 설정합니다.")
    
    # 탭으로 기능 구분
    tab1, tab2 = st.tabs(["🏁 시작/끝 규칙", "🔗 순서 규칙"])
    
    with tab1:
        st.markdown("💡 **면접의 첫 번째와 마지막 활동을 지정하세요.** (선택사항)")
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        # 현재 값을 기본값으로 설정
        start_idx = 0
        end_idx = 0
        if current_start and current_start in ACT_OPTS:
            start_idx = ACT_OPTS.index(current_start) + 1
        if current_end and current_end in ACT_OPTS:
            end_idx = ACT_OPTS.index(current_end) + 1
        
        with col1:
            first = st.selectbox(
                "🏁 가장 먼저 할 활동", 
                ["(지정 안 함)"] + ACT_OPTS, 
                index=start_idx, 
                key=f"first_act_{precedence_refresh_count}",
                help="면접 프로세스의 첫 번째 활동을 선택하세요"
            )
        
        with col2:
            last = st.selectbox(
                "🏁 가장 마지막 활동", 
                ["(지정 안 함)"] + ACT_OPTS, 
                index=end_idx, 
                key=f"last_act_{precedence_refresh_count}",
                help="면접 프로세스의 마지막 활동을 선택하세요"
            )
        
        with col3:
            st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
            if st.button("✅ 적용", key="btn_add_start_end", type="primary", use_container_width=True):
                # 기존 __START__/__END__ 관련 행 제거
                tmp = prec_df[
                    (~prec_df["predecessor"].isin(["__START__", "__END__"])) &
                    (~prec_df["successor"].isin(["__START__", "__END__"]))
                ].copy()
                
                rows = []
                if first != "(지정 안 함)":
                    rows.append({"predecessor": "__START__", "successor": first, "gap_min": 0, "adjacent": True})
                if last != "(지정 안 함)":
                    rows.append({"predecessor": last, "successor": "__END__", "gap_min": 0, "adjacent": True})
                
                st.session_state["precedence"] = pd.concat([tmp, pd.DataFrame(rows)], ignore_index=True)
                st.success("✅ 시작/끝 활동이 설정되었습니다!")
                st.rerun()
    
    with tab2:
        st.markdown("💡 **특정 활동 다음에 반드시 와야 하는 활동을 연결하세요.** (선택사항)")
        st.markdown("📝 **예시:** 면접1 → 면접2 (면접1 후에 반드시 면접2가 와야 함)")
        
        # 충돌 처리 방식 설명
        with st.expander("❓ 전역 간격과 선후행 제약 충돌 시 처리 방식", expanded=False):
            st.markdown("""
            **🔄 충돌 해결 우선순위:**
            
            1. **연속 배치 (adjacent=True)**: 선후행 제약의 정확한 간격 적용
               - 전역 간격 무시
               - 예: 발표준비 → 발표면접 (0분 간격) ✅
            
            2. **일반 선후행 (adjacent=False)**: 더 큰 간격 적용
               - max(선후행 간격, 전역 간격) 사용
               - 예: 전역 5분 vs 선후행 0분 → 5분 적용
            
            3. **권장 설정**:
               - 붙여서 진행할 활동: 연속 배치 체크 ✅
               - 여유 시간 필요한 활동: 전역 간격보다 큰 값 설정
            """)
            
            st.info("💡 **발표준비-발표면접 0분 간격 설정 시**: '연속 배치' 옵션을 체크하면 정확히 0분으로 배치됩니다!")
        
        # 기본 연결 설정
        st.markdown("#### 📌 활동 연결")
        col_form1, col_form2, col_form3 = st.columns([2, 2, 1])
        
        with col_form1:
            p = st.selectbox("🔸 먼저 할 활동", ACT_OPTS, key=f"pred_select_{precedence_refresh_count}", help="선행 활동을 선택하세요")
        
        with col_form2:
            s = st.selectbox("🔹 다음에 할 활동", ACT_OPTS, key=f"succ_select_{precedence_refresh_count}", help="후행 활동을 선택하세요")
        
        with col_form3:
            st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
            add_rule_btn = st.button("✅ 적용", key="btn_add_sequence", type="primary", use_container_width=True)
        
        # 고급 옵션을 별도 영역으로 분리
        st.markdown("#### ⚙️ 고급 옵션")
        col_gap, col_adj = st.columns(2)
        
        # 전역 간격 정보 표시
        current_global_gap = st.session_state.get('global_gap_min', 5)
        
        with col_gap:
            g = st.number_input(
                "⏱️ 최소 간격 (분)", 
                0, 60, 5, 
                key=f"gap_input_{precedence_refresh_count}", 
                help=f"두 활동 사이의 최소 시간 간격 (현재 전역 간격: {current_global_gap}분)"
            )
            
            # 충돌 경고
            if g < current_global_gap:
                st.warning(f"⚠️ 전역 간격({current_global_gap}분)보다 작습니다!")
                
        with col_adj:
            adj = st.checkbox(
                "📌 연속 배치 (붙여서 진행)", 
                value=True, 
                key=f"adj_checkbox_{precedence_refresh_count}", 
                help="체크 시: 전역 간격 무시하고 정확히 지정된 간격으로 배치"
            )
            
            if adj and g < current_global_gap:
                st.success("✅ 연속 배치로 충돌 해결됩니다!")
        
        if add_rule_btn:
            df = st.session_state["precedence"]
            dup = ((df["predecessor"] == p) & (df["successor"] == s)).any()
            if p == s:
                st.error("❌ 같은 활동끼리는 연결할 수 없습니다.")
            elif dup:
                st.warning("⚠️ 이미 존재하는 규칙입니다.")
            else:
                st.session_state["precedence"] = pd.concat(
                    [df, pd.DataFrame([{"predecessor": p, "successor": s, "gap_min": g, "adjacent": adj}])],
                    ignore_index=True
                )
                st.success(f"✅ 규칙 추가: {p} → {s}")
                st.rerun()
    
    # 설정된 규칙 관리 및 삭제
    with st.expander("🗂️ 설정된 규칙 관리", expanded=True):
        prec_df = st.session_state["precedence"].copy()
        
        if not prec_df.empty:
            # 규칙을 보기 좋게 표시
            prec_df["규칙표시용"] = prec_df.apply(
                lambda r: f"{r.predecessor} → {r.successor}" + 
                         (f" (간격: {r.gap_min}분)" if r.gap_min > 0 else "") +
                         (" [연속배치]" if r.adjacent else ""), axis=1
            )
            
            # 2단 구조: 왼쪽은 규칙 목록, 오른쪽은 삭제 기능
            col_rules, col_actions = st.columns([3, 2])
            
            with col_rules:
                st.markdown("**📋 현재 설정된 규칙들**")
                
                # START/END 규칙과 일반 규칙 분리해서 표시
                start_end_rules = prec_df[
                    (prec_df["predecessor"] == "__START__") | (prec_df["successor"] == "__END__")
                ]
                normal_rules = prec_df[
                    (~prec_df["predecessor"].isin(["__START__", "__END__"])) &
                    (~prec_df["successor"].isin(["__START__", "__END__"]))
                ]
                
                if not start_end_rules.empty:
                    st.markdown("**🏁 시작/끝 규칙:**")
                    for rule in start_end_rules["규칙표시용"]:
                        st.markdown(f"• {rule}")
                
                if not normal_rules.empty:
                    st.markdown("**🔗 순서 연결 규칙:**")
                    for rule in normal_rules["규칙표시용"]:
                        st.markdown(f"• {rule}")
            
            with col_actions:
                st.markdown("**🗑️ 규칙 삭제**")
                
                # 삭제할 규칙 선택
                delete_options = prec_df["규칙표시용"].tolist()
                to_delete = st.multiselect(
                    "삭제할 규칙 선택",
                    options=delete_options,
                    key=f"del_prec_select_{precedence_refresh_count}",
                    help="여러 규칙을 한 번에 선택 가능"
                )
                
                # 삭제 버튼들
                if st.button("❌ 선택 삭제", key="del_prec", disabled=not to_delete, use_container_width=True):
                    if to_delete:
                        new_prec = prec_df[~prec_df["규칙표시용"].isin(to_delete)].drop(
                            columns="규칙표시용"
                        ).reset_index(drop=True)
                        st.session_state["precedence"] = new_prec.copy()
                        st.success(f"✅ {len(to_delete)}개 규칙 삭제!")
                        st.rerun()
                
                if st.button("🗑️ 전체 삭제", key="clear_all_prec", type="secondary", use_container_width=True):
                    st.session_state["precedence"] = pd.DataFrame(columns=["predecessor", "successor", "gap_min", "adjacent"])
                    st.success("✅ 모든 규칙 삭제!")
                    st.rerun()
        else:
            st.info("📋 설정된 선후행 제약 규칙이 없습니다. 위에서 규칙을 추가해보세요.")

st.divider()

# =============================================================================
# 섹션 3: 면접 날짜 및 직무 설정
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("3️⃣ 면접 날짜 및 직무 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🔄 섹션 새로고침", key="refresh_date_settings"):
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "date_settings" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["date_settings"] = 0
        st.session_state["section_refresh_counter"]["date_settings"] += 1
        st.rerun()

st.markdown("면접을 진행할 날짜와 각 날짜별 직무 및 인원을 설정합니다.")

# 기본값: 내일 날짜
from datetime import date, timedelta
default_date = date.today() + timedelta(days=1)

# 날짜 모드 선택 (여러 날짜를 기본값으로)
date_mode = st.radio(
    "📅 날짜 설정 모드",
    options=["single", "multiple"],
    format_func=lambda x: "단일 날짜" if x == "single" else "여러 날짜",
    index=1,  # 여러 날짜를 기본값으로 설정
    horizontal=True,
    help="단일 날짜: 하루만 면접 진행 | 여러 날짜: 여러 날에 걸쳐 면접 진행"
)

if date_mode == "single":
    # 단일 날짜 입력
    interview_date = st.date_input(
        "📅 면접 날짜",
        value=st.session_state.get("interview_date", default_date),
        min_value=date.today(),
        help="면접을 진행할 날짜를 선택하세요"
    )
    
    # 세션 상태에 저장 (단일 날짜를 리스트로 저장)
    st.session_state["interview_dates"] = [interview_date]
    st.session_state["interview_date"] = interview_date  # 하위 호환성
    
    # 확인 메시지
    st.success(f"✅ 면접 날짜: **{interview_date.strftime('%Y년 %m월 %d일 (%A)')}**")
    
    # 단일 날짜 모드에서 직무별 인원 설정
    st.markdown("### 👥 직무별 응시 인원")
    
    # 기존 데이터 로드 또는 기본값 설정
    if "single_date_jobs" not in st.session_state:
        st.session_state["single_date_jobs"] = [{"code": "JOB01", "count": 20}]
    
    jobs = st.session_state["single_date_jobs"]
    
    # 직무 추가/제거 버튼
    col_add_job, col_remove_job = st.columns(2)
    
    with col_add_job:
        if st.button("➕ 직무 추가", key="add_single_job", help="새로운 직무를 추가합니다"):
            job_num = len(jobs) + 1
            jobs.append({"code": f"JOB{job_num:02d}", "count": 10})
            st.rerun()
    
    with col_remove_job:
        if len(jobs) > 1 and st.button("➖ 마지막 직무 제거", key="remove_single_job", help="마지막 직무를 제거합니다"):
            jobs.pop()
            st.rerun()
    
    # 직무 선택 삭제 기능 추가 (단일 날짜 모드)
    if len(jobs) > 1:
        st.markdown("---")
        st.markdown("**🗑️ 직무 선택 삭제**")
        
        # 삭제할 직무 선택 옵션 생성
        delete_job_options = []
        for i, job in enumerate(jobs):
            job_code = job.get("code", f"JOB{i+1:02d}")
            job_count = job.get("count", 0)
            delete_job_options.append(f"{i}: {job_code} ({job_count}명)")
        
        # 직무 선택 삭제 UI
        col_select, col_delete = st.columns([3, 1])
        
        with col_select:
            jobs_to_delete = st.multiselect(
                "삭제할 직무 선택",
                options=delete_job_options,
                key="del_single_job_select",
                help="여러 직무를 한 번에 선택하여 삭제할 수 있습니다"
            )
        
        with col_delete:
            st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
            if st.button("❌ 선택 삭제", key="del_single_job", 
                       disabled=not jobs_to_delete, 
                       type="secondary", 
                       use_container_width=True):
                if jobs_to_delete:
                    # 선택된 인덱스 추출
                    selected_indices = [int(s.split(":")[0]) for s in jobs_to_delete]
                    
                    # 실제 jobs 리스트에 존재하는 인덱스만 필터링 (내림차순으로 정렬하여 뒤에서부터 삭제)
                    valid_to_delete = sorted([idx for idx in selected_indices if 0 <= idx < len(jobs)], reverse=True)
                    
                    if valid_to_delete:
                        # 뒤에서부터 삭제하여 인덱스 변화 방지
                        for idx in valid_to_delete:
                            jobs.pop(idx)
                        
                        st.success(f"✅ 선택된 {len(valid_to_delete)}개 직무가 삭제되었습니다.")
                        st.rerun()
                    else:
                        st.error("삭제할 유효한 직무가 없습니다.")
    
    # 직무별 설정 UI
    for i, job in enumerate(jobs):
        col_code, col_count = st.columns(2)
        
        with col_code:
            job["code"] = st.text_input(
                f"직무 코드 {i+1}",
                value=job.get("code", f"JOB{i+1:02d}"),
                key=f"single_job_code_{i}",
                help="직무를 구분하는 고유 코드"
            )
        
        with col_count:
            job["count"] = st.number_input(
                f"응시 인원 {i+1}",
                min_value=1,
                max_value=500,
                value=job.get("count", 20),
                key=f"single_job_count_{i}",
                help="해당 직무의 응시 인원 수"
            )
    
    # 총 인원 계산
    total_applicants = sum(job["count"] for job in jobs)
    st.info(f"📊 **총 응시 인원**: {total_applicants}명")

else:
    # 멀티 날짜 입력
    st.markdown("### 📅 여러 날짜 설정")
    
    # 기존 저장된 날짜별 설정 가져오기
    if "multidate_plans" not in st.session_state:
        st.session_state["multidate_plans"] = {}
    
    multidate_plans = st.session_state["multidate_plans"]
    
    # 날짜 추가/제거 버튼
    col_add, col_remove = st.columns(2)
    
    with col_add:
        if st.button("➕ 날짜 추가", help="새로운 면접 날짜를 추가합니다"):
            new_date = default_date
            while new_date.isoformat() in multidate_plans:
                new_date += timedelta(days=1)
            
            multidate_plans[new_date.isoformat()] = {
                "date": new_date,
                "enabled": True,
                "jobs": [{"code": "JOB01", "count": 20}]  # 기본값
            }
            st.rerun()
    
    with col_remove:
        if multidate_plans and st.button("➖ 마지막 날짜 제거", help="마지막 추가된 날짜를 제거합니다"):
            if multidate_plans:
                last_key = max(multidate_plans.keys())
                del multidate_plans[last_key]
                st.rerun()
    
    # 날짜가 없으면 기본 날짜 하나 추가
    if not multidate_plans:
        multidate_plans[default_date.isoformat()] = {
            "date": default_date,
            "enabled": True,
            "jobs": [{"code": "JOB01", "count": 20}]
        }
    
    # 날짜별 설정 UI
    st.markdown("### 📋 날짜별 상세 설정")
    
    selected_dates = []
    total_applicants = 0
    
    for date_key in sorted(multidate_plans.keys()):
        plan = multidate_plans[date_key]
        plan_date = plan["date"]
        
        with st.expander(f"📅 {plan_date.strftime('%Y년 %m월 %d일 (%A)')} 설정", expanded=True):
            col1, col2 = st.columns([1, 4])
            
            with col1:
                # 날짜 활성화/비활성화
                enabled = st.checkbox(
                    "사용",
                    value=plan.get("enabled", True),
                    key=f"date_enabled_{date_key}",
                    help="이 날짜에 면접을 진행할지 선택"
                )
                plan["enabled"] = enabled
            
            with col2:
                # 날짜 수정
                new_date = st.date_input(
                    "날짜",
                    value=plan_date,
                    min_value=date.today(),
                    key=f"date_picker_{date_key}",
                    help="면접 날짜를 변경할 수 있습니다"
                )
                
                if new_date != plan_date:
                    # 날짜가 변경되면 키를 업데이트
                    new_key = new_date.isoformat()
                    if new_key not in multidate_plans:
                        plan["date"] = new_date
                        multidate_plans[new_key] = plan
                        del multidate_plans[date_key]
                        st.rerun()
                    else:
                        st.error("❌ 이미 존재하는 날짜입니다.")
            
            if enabled:
                # 직무별 인원 설정
                st.markdown("**👥 직무별 응시 인원**")
                
                jobs = plan.get("jobs", [{"code": "JOB01", "count": 20}])
                
                # 직무 추가/제거 버튼
                col_add_job, col_remove_job = st.columns(2)
                
                with col_add_job:
                    if st.button("➕ 직무 추가", key=f"add_job_{date_key}", help="새로운 직무를 추가합니다"):
                        job_num = len(jobs) + 1
                        jobs.append({"code": f"JOB{job_num:02d}", "count": 10})
                        plan["jobs"] = jobs
                        st.rerun()
                
                with col_remove_job:
                    if len(jobs) > 1 and st.button("➖ 마지막 직무 제거", key=f"remove_job_{date_key}", help="마지막 직무를 제거합니다"):
                        jobs.pop()
                        plan["jobs"] = jobs
                        st.rerun()
                
                # 직무 선택 삭제 기능 추가
                if len(jobs) > 1:
                    st.markdown("---")
                    st.markdown("**🗑️ 직무 선택 삭제**")
                    
                    # 삭제할 직무 선택 옵션 생성
                    delete_job_options = []
                    for i, job in enumerate(jobs):
                        job_code = job.get("code", f"JOB{i+1:02d}")
                        job_count = job.get("count", 0)
                        delete_job_options.append(f"{i}: {job_code} ({job_count}명)")
                    
                    # 직무 선택 삭제 UI
                    col_select, col_delete = st.columns([3, 1])
                    
                    with col_select:
                        jobs_to_delete = st.multiselect(
                            "삭제할 직무 선택",
                            options=delete_job_options,
                            key=f"del_job_select_{date_key}",
                            help="여러 직무를 한 번에 선택하여 삭제할 수 있습니다"
                        )
                    
                    with col_delete:
                        st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
                        if st.button("❌ 선택 삭제", key=f"del_job_{date_key}", 
                                   disabled=not jobs_to_delete, 
                                   type="secondary", 
                                   use_container_width=True):
                            if jobs_to_delete:
                                # 선택된 인덱스 추출
                                selected_indices = [int(s.split(":")[0]) for s in jobs_to_delete]
                                
                                # 실제 jobs 리스트에 존재하는 인덱스만 필터링 (내림차순으로 정렬하여 뒤에서부터 삭제)
                                valid_to_delete = sorted([idx for idx in selected_indices if 0 <= idx < len(jobs)], reverse=True)
                                
                                if valid_to_delete:
                                    # 뒤에서부터 삭제하여 인덱스 변화 방지
                                    for idx in valid_to_delete:
                                        jobs.pop(idx)
                                    
                                    plan["jobs"] = jobs
                                    st.success(f"✅ 선택된 {len(valid_to_delete)}개 직무가 삭제되었습니다.")
                                    st.rerun()
                                else:
                                    st.error("삭제할 유효한 직무가 없습니다.")
                
                # 직무별 설정 UI
                for i, job in enumerate(jobs):
                    col_code, col_count = st.columns(2)
                    
                    with col_code:
                        job["code"] = st.text_input(
                            f"직무 코드",
                            value=job.get("code", f"JOB{i+1:02d}"),
                            key=f"job_code_{date_key}_{i}",
                            help="직무를 구분하는 고유 코드"
                        )
                    
                    with col_count:
                        job["count"] = st.number_input(
                            f"응시 인원",
                            min_value=1,
                            max_value=500,
                            value=job.get("count", 20),
                            key=f"job_count_{date_key}_{i}",
                            help="해당 직무의 응시 인원 수"
                        )
                
                # 이 날짜의 총 인원 계산
                date_total = sum(job["count"] for job in jobs)
                st.info(f"📊 **{plan_date.strftime('%m/%d')} 총 인원**: {date_total}명")
                
                # 활성화된 날짜만 선택 목록에 추가
                selected_dates.append(plan_date)
                total_applicants += date_total
    
    # 전체 요약
    if selected_dates:
        st.session_state["interview_dates"] = sorted(selected_dates)
        st.session_state["interview_date"] = selected_dates[0]  # 첫 번째 날짜를 기본값으로
        
        # 확인 메시지
        if len(selected_dates) == 1:
            st.success(f"✅ 면접 날짜: **{selected_dates[0].strftime('%Y년 %m월 %d일 (%A)')}** (총 {total_applicants}명)")
        else:
            date_list = [d.strftime('%m/%d') for d in selected_dates]
            st.success(f"✅ 면접 날짜: **{len(selected_dates)}일간** ({', '.join(date_list)}) - **총 {total_applicants}명**")
    else:
        st.warning("⚠️ 최소 하나의 날짜를 활성화해주세요.")

st.divider()

# =============================================================================
# 섹션 4: 직무별 면접활동 정의 (현황판)
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("4️⃣ 직무별 면접활동 정의")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🔄 섹션 새로고침", key="refresh_job_activities", help="직무별 면접활동 AG-Grid가 먹통일 때 새로고침"):
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "job_activities" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["job_activities"] = 0
        st.session_state["section_refresh_counter"]["job_activities"] += 1
        st.rerun()

st.markdown("위에서 설정한 날짜와 직무 코드를 기반으로 각 직무가 어떤 면접활동을 진행할지 설정합니다.")

# 활동 목록 확보
acts_df = st.session_state.get("activities")
if acts_df is None or acts_df.empty:
    st.error("먼저 면접활동을 정의해주세요.")
else:
    act_list = acts_df.query("use == True")["activity"].tolist()
    
    if not act_list:
        st.error("활동을 최소 1개 '사용'으로 체크해야 합니다.")
    else:
        # 날짜와 직무 설정에서 직무 코드 목록 추출
        all_job_codes = set()
        
        # 단일 날짜 모드에서 직무 코드 추출
        if "single_date_jobs" in st.session_state:
            for job in st.session_state["single_date_jobs"]:
                if job.get("code"):
                    all_job_codes.add(job["code"])
        
        # 멀티 날짜 모드에서 직무 코드 추출
        if "multidate_plans" in st.session_state:
            for plan in st.session_state["multidate_plans"].values():
                if plan.get("enabled", False):
                    for job in plan.get("jobs", []):
                        if job.get("code"):
                            all_job_codes.add(job["code"])
        
        if not all_job_codes:
            st.warning("⚠️ 먼저 위에서 면접 날짜와 직무를 설정해주세요.")
        else:
            # 현황판 표시
            st.subheader("📊 직무별 면접활동 현황")
            
            # 기존 job_acts_map 데이터 로드 또는 생성
            if "job_acts_map" in st.session_state:
                job_df = st.session_state["job_acts_map"].copy()
            else:
                job_df = pd.DataFrame()
            
            # 새로운 직무 코드에 대해 기본 설정 추가
            for job_code in sorted(all_job_codes):
                if job_code not in job_df["code"].values if not job_df.empty else True:
                    # 날짜별 직무 설정에서 해당 직무의 총 인원수 계산
                    total_count = 0
                    
                    # 단일 날짜 모드에서 인원수 계산
                    if "single_date_jobs" in st.session_state:
                        for job in st.session_state["single_date_jobs"]:
                            if job.get("code") == job_code:
                                total_count += job.get("count", 0)
                    
                    # 멀티 날짜 모드에서 인원수 계산
                    if "multidate_plans" in st.session_state:
                        for plan in st.session_state["multidate_plans"].values():
                            if plan.get("enabled", False):
                                for job in plan.get("jobs", []):
                                    if job.get("code") == job_code:
                                        total_count += job.get("count", 0)
                    
                    new_row = {"code": job_code, "count": max(total_count, 1)}
                    for act in act_list:
                        new_row[act] = True
                    
                    if job_df.empty:
                        job_df = pd.DataFrame([new_row])
                    else:
                        job_df = pd.concat([job_df, pd.DataFrame([new_row])], ignore_index=True)
            
            # 더 이상 사용하지 않는 직무 코드 제거
            if not job_df.empty:
                job_df = job_df[job_df["code"].isin(all_job_codes)].reset_index(drop=True)
            
            # 컬럼 동기화
            for act in act_list:
                if act not in job_df.columns:
                    job_df[act] = True
            
            if "count" not in job_df.columns:
                job_df["count"] = 1
            
            # 열 순서 정리
            cols = ["code", "count"] + act_list
            job_df = job_df.reindex(columns=cols, fill_value=True)
            
            st.session_state["job_acts_map"] = job_df
            
            # 현황 정보 표시
            col1, col2 = st.columns(2)
            with col1:
                st.metric("감지된 직무 수", len(all_job_codes))
            with col2:
                st.metric("활용 가능한 활동 수", len(act_list))
            
            # 직무별 활동 설정을 위한 간단한 편집 인터페이스
            st.markdown("### ✏️ 직무별 활동 설정")
            
            # 편집용 AG-Grid
            df_to_display = st.session_state["job_acts_map"].copy()
            
            gb2 = GridOptionsBuilder.from_dataframe(df_to_display)
            gb2.configure_selection(selection_mode="none")
            gb2.configure_default_column(resizable=True, editable=True)
            
            gb2.configure_column("code", header_name="직무 코드", width=120, editable=False)  # 직무 코드는 읽기 전용
            gb2.configure_column(
                "count", header_name="인원수", type=["numericColumn"], width=90, editable=True
            )
            
            for act in act_list:
                gb2.configure_column(
                    act,
                    header_name=act,
                    cellRenderer="agCheckboxCellRenderer",
                    cellEditor="agCheckboxCellEditor",
                    editable=True,
                    singleClickEdit=True,
                    width=110,
                )
            
            grid_opts2 = gb2.build()
            
            # 섹션별 새로고침을 위한 동적 key 생성
            job_activities_refresh_count = st.session_state.get("section_refresh_counter", {}).get("job_activities", 0)
            
            grid_ret2 = AgGrid(
                df_to_display,
                gridOptions=grid_opts2,
                update_mode=GridUpdateMode.VALUE_CHANGED,
                data_return_mode=DataReturnMode.AS_INPUT,
                fit_columns_on_grid_load=True,
                theme="balham",
                key=f"job_grid_display_{job_activities_refresh_count}",  # 동적 key로 강제 재렌더링
            )
            
            edited_df = pd.DataFrame(grid_ret2["data"])
            
            # 데이터 검증
            def validate_job_data(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
                msgs: list[str] = []
                df = df.copy()
                df["code"] = df["code"].astype(str).str.strip()
                df["count"] = pd.to_numeric(df["count"], errors="coerce").fillna(0).astype(int)
                
                # count ≤ 0
                zero_cnt = df[df["count"] <= 0]["code"].tolist()
                if zero_cnt:
                    msgs.append(f"0 이하 인원수: {', '.join(map(str, zero_cnt))}")
                
                # 활동이 하나도 선택되지 않은 행
                no_act = [
                    row.code
                    for row in df.itertuples()
                    if not any(getattr(row, a) for a in act_list)
                ]
                if no_act:
                    msgs.append(f"모든 활동이 False인 코드: {', '.join(no_act)}")
                
                return df, msgs
            
            clean_df, errors = validate_job_data(edited_df)
            
            if errors:
                for msg in errors:
                    st.error(msg)
            else:
                st.success("✅ 모든 설정이 유효합니다!")
            
            st.info(f"📊 총 설정된 인원수: **{clean_df['count'].sum()}** 명")
            st.session_state["job_acts_map"] = clean_df
            
            # 요약 정보 표시
            with st.expander("📋 직무별 활동 요약", expanded=False):
                for _, row in clean_df.iterrows():
                    job_code = row["code"]
                    selected_acts = [act for act in act_list if row[act]]
                    st.write(f"**{job_code}**: {', '.join(selected_acts)} ({row['count']}명)")

st.divider()

# =============================================================================
# 섹션 5: 운영 공간 설정
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("5️⃣ 운영 공간 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🔄 섹션 새로고침", key="refresh_room_settings", help="운영 공간 설정 UI가 먹통일 때 새로고침"):
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "room_settings" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["room_settings"] = 0
        st.session_state["section_refresh_counter"]["room_settings"] += 1
        st.rerun()

st.markdown("면접을 운영할 경우, 하루에 동원 가능한 모든 공간의 종류와 수, 그리고 최대 수용 인원을 설정합니다.")

# 활동 DF에서 room_types 확보
acts_df = st.session_state.get("activities")
if acts_df is not None and not acts_df.empty:
    room_types = sorted(
        acts_df.query("use == True and room_type != '' and room_type.notna()")["room_type"].unique()
    )
    
    if room_types:
        min_cap_req = acts_df.set_index("room_type")["min_cap"].to_dict()
        max_cap_req = acts_df.set_index("room_type")["max_cap"].to_dict()
        
        # 공간 템플릿 설정
        tpl_dict = st.session_state.get("room_template", {})
        
        # room_types 동기화
        for rt in room_types:
            tpl_dict.setdefault(rt, {"count": 1, "cap": max_cap_req.get(rt, 1)})
        for rt in list(tpl_dict):
            if rt not in room_types:
                tpl_dict.pop(rt)
        
        st.subheader("하루 기준 운영 공간 설정")
        
        # 섹션별 새로고침을 위한 동적 key 생성
        room_settings_refresh_count = st.session_state.get("section_refresh_counter", {}).get("room_settings", 0)
        
        col_cnt, col_cap = st.columns(2, gap="large")
        
        with col_cnt:
            st.markdown("#### 방 개수")
            for rt in room_types:
                tpl_dict[rt]["count"] = st.number_input(
                    f"{rt} 개수", 
                    min_value=0, 
                    max_value=50, 
                    value=tpl_dict[rt].get("count", 1), 
                    key=f"tpl_{rt}_cnt_{room_settings_refresh_count}"  # 동적 key로 강제 재렌더링
                )
        
        with col_cap:
            st.markdown("#### 최대 동시 수용 인원")
            for rt in room_types:
                min_val = min_cap_req.get(rt, 1)
                max_val = max_cap_req.get(rt, 50)
                current_val = tpl_dict[rt].get("cap", max_val)
                safe_val = max(min_val, min(current_val, max_val))
                
                tpl_dict[rt]["cap"] = st.number_input(
                    f"{rt} 최대 동시 수용 인원",
                    min_value=min_val,
                    max_value=max_val,
                    value=safe_val,
                    key=f"tpl_{rt}_cap_{room_settings_refresh_count}",  # 동적 key로 강제 재렌더링
                )
        
        # 변경된 템플릿 정보를 세션에 저장
        st.session_state['room_template'] = tpl_dict
        
        # room_plan 생성 및 저장
        final_plan_dict = {}
        for rt, values in tpl_dict.items():
            final_plan_dict[f"{rt}_count"] = values['count']
            final_plan_dict[f"{rt}_cap"] = values['cap']
        
        st.session_state['room_plan'] = pd.DataFrame([final_plan_dict])
        
        # room_template도 동기화
        st.session_state['room_template'] = tpl_dict
        
        with st.expander("🗂 저장된 room_plan 데이터 미리보기"):
            st.dataframe(st.session_state.get('room_plan', pd.DataFrame()), use_container_width=True)
    else:
        st.error("사용(use=True)하도록 설정된 활동 중, 'room_type'이 지정된 활동이 없습니다.")
else:
    st.error("먼저 면접활동을 정의해주세요.")

st.divider()

# =============================================================================
# 섹션 6: 운영 시간 설정 (기존 섹션 5에서 번호 변경)
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("6️⃣ 운영 시간 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_time_settings", help="운영 시간 설정 UI가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "time_settings" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["time_settings"] = 0
        st.session_state["section_refresh_counter"]["time_settings"] += 1
        st.rerun()

st.markdown("면접을 운영할 경우의 하루 기준 운영 시작 및 종료 시간을 설정합니다.")

# 기존 값 불러오기
init_start = st.session_state.get("oper_start_time", time(9, 0))
init_end = st.session_state.get("oper_end_time", time(18, 0))

st.subheader("하루 기준 공통 운영 시간")

# 섹션별 새로고침을 위한 동적 key 생성
time_settings_refresh_count = st.session_state.get("section_refresh_counter", {}).get("time_settings", 0)

col_start, col_end = st.columns(2)

with col_start:
    t_start = st.time_input("운영 시작 시간", value=init_start, key=f"oper_start_{time_settings_refresh_count}")

with col_end:
    t_end = st.time_input("운영 종료 시간", value=init_end, key=f"oper_end_{time_settings_refresh_count}")

if t_start >= t_end:
    st.error("오류: 운영 시작 시간은 종료 시간보다 빨라야 합니다.")
else:
    # 설정된 시간을 세션 상태에 저장
    st.session_state["oper_start_time"] = t_start
    st.session_state["oper_end_time"] = t_end
    
    # oper_window DataFrame 생성 및 저장
    oper_window_dict = {
        "start_time": t_start.strftime("%H:%M"),
        "end_time": t_end.strftime("%H:%M")
    }
    st.session_state['oper_window'] = pd.DataFrame([oper_window_dict])
    
    with st.expander("🗂 저장된 oper_window 데이터 미리보기"):
        st.dataframe(st.session_state.get('oper_window', pd.DataFrame()), use_container_width=True)
    
    st.success(f"운영 시간이 {t_start.strftime('%H:%M')}부터 {t_end.strftime('%H:%M')}까지로 설정되었습니다.")

st.divider()

# 위로 가기 기능
st.markdown("---")

# 중앙 링크 방식
st.markdown("""
<div style="text-align: center; margin: 20px 0;">
    <a href="#top" style="
        display: inline-block;
        background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
        color: white;
        text-decoration: none;
        padding: 12px 24px;
        border-radius: 25px;
        font-weight: bold;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        transition: all 0.3s ease;
    ">
        ⬆️ 빠른 이동: 맨 위로
    </a>
</div>
""", unsafe_allow_html=True)

# 우하단 고정 버튼 (개선된 버전)
st.markdown("""
<style>
    .floating-top-button {
        position: fixed;
        bottom: 30px;
        right: 30px;
        z-index: 999;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 50%;
        width: 60px;
        height: 60px;
        font-size: 20px;
        cursor: pointer;
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
        transition: all 0.3s ease;
        display: flex;
        align-items: center;
        justify-content: center;
        text-decoration: none;
    }
    
    .floating-top-button:hover {
        transform: translateY(-3px) scale(1.1);
        box-shadow: 0 6px 25px rgba(0,0,0,0.4);
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
    
    .floating-top-button:active {
        transform: translateY(-1px) scale(1.05);
    }
</style>

<a href="#top" class="floating-top-button" title="맨 위로 이동">
    ⬆️
</a>
""", unsafe_allow_html=True)

# =============================================================================
# 섹션 7: 집단면접 설정 (새로 추가)
# =============================================================================
# batched 모드가 있을 때만 표시
acts_df = st.session_state.get("activities", pd.DataFrame())
has_batched = any(acts_df["mode"] == "batched") if not acts_df.empty and "mode" in acts_df.columns else False

if has_batched:
    col_header, col_refresh = st.columns([3, 2])
    with col_header:
        st.header("7️⃣ 집단면접 설정")
    with col_refresh:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 섹션 새로고침", key="refresh_group_settings"):
            if "section_refresh_counter" not in st.session_state:
                st.session_state["section_refresh_counter"] = {}
            if "group_settings" not in st.session_state["section_refresh_counter"]:
                st.session_state["section_refresh_counter"]["group_settings"] = 0
            st.session_state["section_refresh_counter"]["group_settings"] += 1
            st.rerun()
    
    st.markdown("집단면접(batched) 활동에 대한 전역 설정을 관리합니다.")
    
    # batched 활동 목록 표시
    batched_activities = acts_df[acts_df["mode"] == "batched"]["activity"].tolist()
    st.info(f"집단면접 활동: {', '.join(batched_activities)}")
    
    # 그룹 크기 설정
    col1, col2 = st.columns(2)
    
    # batched 활동 필터링 추가
    batched_acts = acts_df[acts_df["mode"] == "batched"]
    
    with col1:
        # 기존 값 가져오기
        current_min = st.session_state.get('group_min_size', 4)
        # batched 활동이 있으면 그 값 사용
        if not batched_acts.empty:
            current_min = int(batched_acts.iloc[0]['min_cap'])
        
        group_min = st.number_input(
            "그룹 최소 인원",
            min_value=2,
            max_value=20,
            value=current_min,
            key="group_min_input",
            help="모든 집단면접 활동에 적용됩니다"
        )
    
    with col2:
        # 기존 값 가져오기
        current_max = st.session_state.get('group_max_size', 6)
        # batched 활동이 있으면 그 값 사용
        if not batched_acts.empty:
            current_max = int(batched_acts.iloc[0]['max_cap'])
        
        group_max = st.number_input(
            "그룹 최대 인원",
            min_value=group_min,
            max_value=30,
            value=max(current_max, group_min),
            key="group_max_input",
            help="모든 집단면접 활동에 적용됩니다"
        )
    
    # 값이 변경되면 모든 batched 활동에 적용
    if group_min != st.session_state.get('group_min_size') or group_max != st.session_state.get('group_max_size'):
        st.session_state['group_min_size'] = group_min
        st.session_state['group_max_size'] = group_max
        
        # activities DataFrame 업데이트
        acts_df = st.session_state["activities"]
        acts_df.loc[acts_df["mode"] == "batched", "min_cap"] = group_min
        acts_df.loc[acts_df["mode"] == "batched", "max_cap"] = group_max
        st.session_state["activities"] = acts_df
        
        st.success(f"✅ 모든 집단면접 활동의 그룹 크기가 {group_min}~{group_max}명으로 설정되었습니다.")
    
    # 추가 설정들
    st.markdown("### 고급 설정")
    
    col3, col4 = st.columns(2)
    
    with col3:
        global_gap = st.number_input(
            "전역 활동 간격(분)",
            min_value=0,
            max_value=60,
            value=st.session_state.get('global_gap_min', 5),
            key="global_gap_input",
            help="모든 활동 간 기본 간격입니다. Precedence에서 개별 설정 가능합니다."
        )
        
        # 충돌 경고 표시
        prec_df = st.session_state.get("precedence", pd.DataFrame())
        if not prec_df.empty:
            conflicts = []
            for _, rule in prec_df.iterrows():
                if rule["gap_min"] < global_gap and not rule.get("adjacent", False):
                    conflicts.append(f"{rule['predecessor']} → {rule['successor']} ({rule['gap_min']}분)")
            
            if conflicts:
                st.warning(f"⚠️ **충돌 감지**: 다음 선후행 제약이 전역 간격({global_gap}분)보다 작습니다:\n" + 
                          "\n".join(f"• {c}" for c in conflicts))
                st.info("💡 **해결 방법**: 선후행 제약의 '연속배치' 옵션을 체크하거나, 간격을 늘리세요.")
        st.session_state['global_gap_min'] = global_gap
    
    with col4:
        max_stay = st.number_input(
            "최대 체류시간(시간)",
            min_value=1,
            max_value=12,
            value=st.session_state.get('max_stay_hours', 8),
            key="max_stay_input",
            help="지원자가 면접장에 머무를 수 있는 최대 시간입니다."
        )
        st.session_state['max_stay_hours'] = max_stay
    
    # 직무별 분리 원칙 표시
    st.markdown("### 그룹 구성 원칙")
    st.info("""
    ✅ **직무별 분리**: 같은 직무끼리만 그룹 구성
    ✅ **그룹 일관성**: 한 번 구성된 그룹은 모든 batched 활동에서 유지
    ✅ **동일 직무 동일 방**: 같은 직무는 모든 활동에서 동일한 접미사(A,B,C...) 사용
    ✅ **그룹 수 최소화**: 더미 지원자를 활용하여 그룹 수 최소화
    """)
    
    st.divider()