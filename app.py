# pages/1_면접운영스케줄링.py
import streamlit as st
import pandas as pd
import re
import json
from datetime import time, datetime, timedelta
from st_aggrid import (
    AgGrid,
    GridOptionsBuilder,
    GridUpdateMode,
    DataReturnMode,
)
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import core
from solver.solver import solve_for_days

st.set_page_config(
    page_title="면접운영스케줄링",
    layout="wide"
)

# 사이드바에서 app 페이지 숨기기
st.markdown("""
<style>
    /* 사이드바에서 첫 번째 항목(app) 숨기기 */
    .css-1d391kg .css-1rs6os .css-17eq0hr:first-child {
        display: none !important;
    }
    
    /* 다른 방법으로 app 항목 숨기기 */
    [data-testid="stSidebar"] [data-testid="stSidebarNav"] ul li:first-child {
        display: none !important;
    }
    
    /* 또 다른 방법 */
    .css-1rs6os .css-17eq0hr:first-child {
        display: none !important;
    }
    
    /* 사이드바 네비게이션에서 첫 번째 링크 숨기기 */
    .css-1rs6os a[href="/"]:first-child {
        display: none !important;
    }
    
    /* 최신 Streamlit 버전용 */
    [data-testid="stSidebarNav"] > ul > li:first-child {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

# 페이지 상단 앵커 포인트
st.markdown('<div id="top"></div>', unsafe_allow_html=True)

st.title("🎯 면접운영스케줄링")
st.markdown("""
**올인원 면접 스케줄링 솔루션**  
이 페이지에서 면접 스케줄링에 필요한 모든 설정과 일정 생성을 한 번에 진행할 수 있습니다.
운영일정 추정을 먼저 확인한 후, 필요한 설정들을 순차적으로 진행해보세요.
""")

# 세션 상태 초기화
def init_session_states():
    # 기본 활동 템플릿
    default_activities = pd.DataFrame({
        "use": [True, True, True],
        "activity": ["토론면접", "발표준비", "발표면접"],
        "mode": ["batched", "parallel", "individual"],
        "duration_min": [30, 5, 15],
        "room_type": ["토론면접실", "발표준비실", "발표면접실"],
        "min_cap": [4, 1, 1],
        "max_cap": [6, 2, 1],
    })
    st.session_state.setdefault("activities", default_activities)
    
    # 스마트 직무 매핑 (모든 기본 활동 활성화 + 실제 인원수)
    if "job_acts_map" not in st.session_state:
        act_list = default_activities.query("use == True")["activity"].tolist()
        job_codes = ["JOB01", "JOB02", "JOB03", "JOB04", "JOB05", "JOB06", "JOB07", "JOB08", "JOB09", "JOB10", "JOB11"]
        job_counts = [23, 23, 20, 20, 12, 15, 6, 6, 6, 3, 3]
        
        job_data = {"code": job_codes, "count": job_counts}
        for act in act_list:
            job_data[act] = [True] * len(job_codes)
        st.session_state["job_acts_map"] = pd.DataFrame(job_data)
    
    # 기본 선후행 제약 (발표준비 → 발표면접)
    default_precedence = pd.DataFrame([
        {"predecessor": "발표준비", "successor": "발표면접", "gap_min": 0, "adjacent": True}  # 발표준비 → 발표면접 연속
    ])
    st.session_state.setdefault("precedence", default_precedence)
    
    # 기본 운영 시간
    st.session_state.setdefault("oper_start_time", time(9, 0))
    st.session_state.setdefault("oper_end_time", time(18, 0))
    
    # 스마트 방 템플릿 (기본 활동에 맞춰 자동 생성)
    if "room_template" not in st.session_state:
        room_template = {
            "토론면접실": {"count": 2, "cap": 6},
            "발표준비실": {"count": 1, "cap": 2},
            "발표면접실": {"count": 2, "cap": 1}
        }
        st.session_state["room_template"] = room_template
    
    # 스마트 운영 공간 계획 (room_template 기반으로 자동 생성)
    if "room_plan" not in st.session_state:
        room_template = st.session_state.get("room_template", {})
        if room_template:
            final_plan_dict = {}
            for rt, values in room_template.items():
                final_plan_dict[f"{rt}_count"] = values['count']
                final_plan_dict[f"{rt}_cap"] = values['cap']
            st.session_state["room_plan"] = pd.DataFrame([final_plan_dict])
        else:
            st.session_state["room_plan"] = pd.DataFrame()
    
    # 스마트 운영 시간 (자동 생성)
    if "oper_window" not in st.session_state:
        t_start = st.session_state["oper_start_time"]
        t_end = st.session_state["oper_end_time"]
        oper_window_dict = {
            "start_time": t_start.strftime("%H:%M"),
            "end_time": t_end.strftime("%H:%M")
        }
        st.session_state["oper_window"] = pd.DataFrame([oper_window_dict])
    
    # 시뮬레이션 결과
    st.session_state.setdefault('final_schedule', None)
    st.session_state.setdefault('last_solve_logs', "")
    st.session_state.setdefault('solver_status', "미실행")
    st.session_state.setdefault('daily_limit', 0)
    
    # 집단면접 설정 초기화 (토론면접에 맞춰 설정)
    st.session_state.setdefault('group_min_size', 4)
    st.session_state.setdefault('group_max_size', 6)
    st.session_state.setdefault('global_gap_min', 5)
    st.session_state.setdefault('max_stay_hours', 5)  # 5시간으로 단축

init_session_states()

# =============================================================================
# 섹션 0: 운영일정 추정 (메인 섹션)
# =============================================================================
st.header("🚀 운영일정 추정")
st.markdown("현재 설정을 바탕으로 최적의 운영일정을 추정합니다.")

# 첫 방문자를 위한 안내
if st.session_state.get('solver_status', '미실행') == '미실행':
    st.info("👋 **처음 방문하셨나요?** 바로 아래 '운영일정추정 시작' 버튼을 눌러보세요! 기본 설정으로 데모를 체험할 수 있습니다.")
    st.markdown("💡 **팁:** 추정 후 아래 섹션들에서 세부 설정을 조정하여 더 정확한 결과를 얻을 수 있습니다.")

# Excel 출력 함수 (타임슬롯 기능 통합)
def df_to_excel(df: pd.DataFrame, stream=None, group_info: dict = None) -> None:
    wb = Workbook()
    
    # 기본 팔레트
    PALETTE = ['E3F2FD', 'FFF3E0', 'E8F5E9', 'FCE4EC', 'E1F5FE', 'F3E5F5', 'FFFDE7', 'E0F2F1', 'EFEBE9', 'ECEFF1']
    
    # 활동별 모드 정보 가져오기 (batched 활동 확인용)
    activities_df = st.session_state.get("activities", pd.DataFrame())
    activity_modes = {}
    if not activities_df.empty:
        for _, act in activities_df.iterrows():
            if act["use"]:
                activity_modes[act["activity"]] = act.get("mode", "individual")
    
    # ===== 1) 기본 스케줄 시트 =====
    ws1 = wb.active
    ws1.title = 'Schedule'
    df_copy = df.copy()
    
    # 체류시간 계산 함수
    def calculate_stay_time(row):
        """각 지원자의 체류시간을 계산 (분 단위)"""
        start_times = []
        end_times = []
        
        for col in row.index:
            if col.startswith('start_') and pd.notna(row[col]):
                try:
                    time_val = pd.to_datetime(row[col])
                    start_times.append(time_val)
                except:
                    pass
            elif col.startswith('end_') and pd.notna(row[col]):
                try:
                    time_val = pd.to_datetime(row[col])
                    end_times.append(time_val)
                except:
                    pass
        
        if start_times and end_times:
            first_start = min(start_times)
            last_end = max(end_times)
            stay_minutes = (last_end - first_start).total_seconds() / 60
            return int(stay_minutes)
        return 0
    
    # 체류시간 칼럼 추가
    df_copy['체류시간(분)'] = df_copy.apply(calculate_stay_time, axis=1)
    df_copy['체류시간(시:분)'] = df_copy['체류시간(분)'].apply(
        lambda x: f"{int(x//60)}:{int(x%60):02d}" if x > 0 else ""
    )
    
    # 그룹 정보 추가 (group_info가 제공된 경우)
    if group_info:
        df_copy['그룹번호'] = df_copy['id'].map(group_info.get('member_to_group', {}))
        df_copy['그룹크기'] = df_copy['그룹번호'].map(group_info.get('group_sizes', {}))
    else:
        df_copy['그룹번호'] = ''
        df_copy['그룹크기'] = ''
    
    # 칼럼 순서 재정렬 - 기본 정보 다음에 체류시간과 그룹 정보를 배치
    base_cols = ['id', 'interview_date', 'code']
    extra_cols = ['그룹번호', '그룹크기', '체류시간(분)', '체류시간(시:분)']
    activity_cols = [col for col in df_copy.columns if col not in base_cols + extra_cols]
    
    # 새로운 칼럼 순서: 기본 정보 → 그룹 정보 → 체류시간 → 활동 정보
    new_column_order = base_cols + [col for col in ['그룹번호', '그룹크기'] if col in df_copy.columns] + \
                      [col for col in ['체류시간(분)', '체류시간(시:분)'] if col in df_copy.columns] + \
                      activity_cols
    df_copy = df_copy[new_column_order]
    
    # 날짜별로 색상 지정
    unique_dates = df_copy['interview_date'].dt.date.unique()
    date_color_map = {date: PALETTE[i % len(PALETTE)] for i, date in enumerate(unique_dates)}
    
    df_copy = df_copy.astype(object).where(pd.notna(df_copy), None)
    for r in dataframe_to_rows(df_copy, index=False, header=True):
        ws1.append(r)
    
    header_fill = PatternFill('solid', fgColor='D9D9D9')
    special_header_fill = PatternFill('solid', fgColor='B8B8B8')  # 체류시간/그룹 정보는 진한 회색
    
    for cell in ws1[1]:
        if cell.value in ['체류시간(분)', '체류시간(시:분)', '그룹번호', '그룹크기']:
            cell.fill = special_header_fill
            cell.font = Font(bold=True)
        else:
            cell.fill = header_fill
    
    # 날짜 열 찾기
    date_col_idx = -1
    for j, col_name in enumerate(df_copy.columns, 1):
        if col_name == 'interview_date':
            date_col_idx = j
            break
    
    for i, row in enumerate(ws1.iter_rows(min_row=2, max_row=ws1.max_row), 2):
        if date_col_idx != -1:
            date_val = row[date_col_idx - 1].value
            if date_val and hasattr(date_val, 'date'):
                row_color = date_color_map.get(date_val.date())
                if row_color:
                    row_fill = PatternFill('solid', fgColor=row_color)
                    for cell in row:
                        cell.fill = row_fill
    
    # 시간 형식 지정
    for j, col_name in enumerate(df_copy.columns, 1):
        if 'start' in col_name or 'end' in col_name:
            for i in range(2, ws1.max_row + 1):
                ws1.cell(i, j).number_format = 'hh:mm'
    
    # ===== 2) 타임슬롯 시트들 추가 =====
    def _color_picker():
        """활동명 → 고정 색상 매핑"""
        mapping = {}
        def _pick(act: str) -> str:
            if act not in mapping:
                mapping[act] = PALETTE[len(mapping) % len(PALETTE)]
            return mapping[act]
        return _pick
    
    def _build_timeslot_sheet(ws, df_day: pd.DataFrame, pick_color, group_info_param=None, activity_modes_param=None):
        """단일 날짜 스케줄 → 타임슬롯 매트릭스"""
        loc_cols = [c for c in df_day.columns if c.startswith("loc_")]
        start_cols = [c for c in df_day.columns if c.startswith("start_")]
        end_cols = [c for c in df_day.columns if c.startswith("end_")]
        
        # group_info 파라미터 사용
        group_info = group_info_param
        
        # 활동별 모드 정보는 파라미터로 전달받음
        activity_modes = activity_modes_param if activity_modes_param is not None else {}
        
        # 공간 목록
        locs = sorted(set(df_day[loc_cols].stack().dropna().unique()))
        if not locs:
            return
        
        # 시간 범위 계산
        t_min = t_max = None
        for col in start_cols + end_cols:
            ts = pd.to_datetime(df_day[col], errors="coerce").dropna()
            if ts.empty:
                continue
            t_min = ts.min() if t_min is None else min(t_min, ts.min())
            t_max = ts.max() if t_max is None else max(t_max, ts.max())
        if t_min is None or t_max is None:
            return
        
        TIME_STEP_MIN = 5
        t_min = t_min.floor(f"{TIME_STEP_MIN}min")
        t_max = (t_max.ceil(f"{TIME_STEP_MIN}min") + timedelta(minutes=TIME_STEP_MIN))
        times = pd.date_range(t_min, t_max, freq=f"{TIME_STEP_MIN}min")
        
        # 헤더 작성
        ws.cell(1, 1, "Time")
        for j, loc in enumerate(locs, start=2):
            cell = ws.cell(1, j, loc)
            cell.alignment = Alignment(horizontal="center")
        
        for i, t in enumerate(times, start=2):
            ws.cell(i, 1, t.strftime("%H:%M"))
            ws.cell(i, 1).alignment = Alignment(horizontal="right")
        
        # 셀 채우기
        for _, row in df_day.iterrows():
            for st_col in start_cols:
                suffix = st_col[len("start_"):]
                end_col = f"end_{suffix}"
                loc_col = f"loc_{suffix}"
                if end_col not in df_day.columns or loc_col not in df_day.columns:
                    continue
                st = row[st_col]
                ed = row[end_col]
                loc = row[loc_col]
                if pd.isna(st) or pd.isna(ed) or loc in ("", None):
                    continue
                st_dt = pd.to_datetime(st, errors="coerce")
                ed_dt = pd.to_datetime(ed, errors="coerce")
                if pd.isna(st_dt) or pd.isna(ed_dt):
                    continue
                base_act = suffix
                color = pick_color(base_act)
                if loc not in locs:
                    continue
                col_idx = locs.index(loc) + 2
                
                # 표시할 내용 결정: batched 활동이면 그룹 번호, 아니면 ID
                display_value = str(row["id"])
                if group_info and base_act in activity_modes and activity_modes[base_act] == "batched":
                    # batched 활동인 경우 그룹 번호 표시
                    member_id = row["id"]
                    if member_id in group_info.get('member_to_group', {}):
                        group_num = group_info['member_to_group'][member_id]
                        display_value = f"G{group_num}"
                
                cur = st_dt.floor(f"{TIME_STEP_MIN}min")
                while cur < ed_dt:
                    if cur < t_min or cur > t_max:
                        cur += timedelta(minutes=TIME_STEP_MIN)
                        continue
                    row_idx = times.get_loc(cur) + 2
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value in (None, ""):
                        cell.value = display_value
                        cell.fill = PatternFill("solid", fgColor=color)
                    else:
                        existing = str(cell.value)
                        if display_value not in existing.split("\n"):
                            cell.value = existing + "\n" + display_value
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cur += timedelta(minutes=TIME_STEP_MIN)
        
        # 열 너비·행 높이 자동 조정
        for j, loc in enumerate(locs, start=2):
            max_len = len(str(loc))
            for i in range(2, ws.max_row + 1):
                val = ws.cell(i, j).value
                if val is None:
                    continue
                for part in str(val).split("\n"):
                    max_len = max(max_len, len(part))
            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = max(10, min(1.2 * max_len, 30))
        
        default_ht = 15
        for i in range(2, ws.max_row + 1):
            max_lines = 1
            for j in range(2, ws.max_column + 1):
                val = ws.cell(i, j).value
                if val is None:
                    continue
                lines = str(val).count("\n") + 1
                max_lines = max(max_lines, lines)
            ws.row_dimensions[i].height = default_ht * max_lines
    
    # 날짜별 타임슬롯 시트 생성
    pick_color = _color_picker()
    for the_date, df_day in df.groupby("interview_date"):
        ws_name = f"TS_{pd.to_datetime(the_date).strftime('%m%d')}"
        ws_ts = wb.create_sheet(ws_name)
        _build_timeslot_sheet(ws_ts, df_day.copy(), pick_color, group_info, activity_modes)
    
    wb.save(stream or "interview_schedule.xlsx")

def reset_run_state():
    st.session_state['final_schedule'] = None
    st.session_state['last_solve_logs'] = ""
    st.session_state['solver_status'] = "미실행"
    st.session_state['daily_limit'] = 0

# 기본 파라미터 설정 (하드코딩)
params = {
    "min_gap_min": st.session_state.get('global_gap_min', 5),
    "time_limit_sec": 120,
    "max_stay_hours": st.session_state.get('max_stay_hours', 8)
}

# batched 모드가 있는지 확인
acts_df = st.session_state.get("activities", pd.DataFrame())
has_batched = any(acts_df["mode"] == "batched") if not acts_df.empty and "mode" in acts_df.columns else False

# 집단면접 설정이 있으면 params에 추가
if has_batched:
    params["group_min_size"] = st.session_state.get('group_min_size', 4)
    params["group_max_size"] = st.session_state.get('group_max_size', 6)

# 데이터 검증
job_df = st.session_state.get("job_acts_map", pd.DataFrame())
room_plan = st.session_state.get("room_plan", pd.DataFrame())
oper_df = st.session_state.get("oper_window", pd.DataFrame())
prec_df = st.session_state.get("precedence", pd.DataFrame())

# 검증 결과 표시
validation_errors = []

if acts_df.empty or not (acts_df["use"] == True).any():
    validation_errors.append("활동을 하나 이상 정의하고 'use=True'로 설정해주세요.")

if job_df.empty or (job_df["count"].sum() == 0):
    validation_errors.append("직무 코드를 추가하고 인원수를 1명 이상 입력해주세요.")

if job_df["code"].duplicated().any():
    validation_errors.append("중복된 직무 코드가 있습니다.")

# batched 모드가 있으면 그룹 크기 일관성 검증
if has_batched:
    batched_activities = acts_df[acts_df["mode"] == "batched"]
    min_caps = batched_activities["min_cap"].unique()
    max_caps = batched_activities["max_cap"].unique()
    
    if len(min_caps) > 1 or len(max_caps) > 1:
        validation_errors.append("모든 batched 활동의 그룹 크기(min_cap, max_cap)는 동일해야 합니다.")
    
    # 방 용량 vs 그룹 크기 검증
    for _, act in batched_activities.iterrows():
        room_type = act["room_type"]
        max_cap = act["max_cap"]
        room_cap_col = f"{room_type}_cap"
        
        if room_cap_col in room_plan.columns:
            room_cap = room_plan[room_cap_col].iloc[0] if not room_plan.empty else 0
            if room_cap < max_cap:
                validation_errors.append(f"{room_type}의 용량({room_cap})이 {act['activity']}의 최대 그룹 크기({max_cap})보다 작습니다.")

# room_plan 검증: room_template이 있으면 유효한 것으로 간주
room_template = st.session_state.get("room_template", {})
if room_plan.empty and not room_template:
    validation_errors.append("운영 공간을 설정해주세요.")

if validation_errors:
    st.error("다음 항목을 설정해주세요:")
    for error in validation_errors:
        st.error(f"• {error}")
    st.info("⬇️ 아래 섹션들에서 필요한 설정을 완료한 후 다시 추정해보세요.")
else:
    st.success("✅ 입력 데이터 검증 통과 – 운영일정추정을 시작할 수 있습니다!")

# 운영일정 추정 실행
if st.button("🚀 운영일정추정 시작", type="primary", use_container_width=True, on_click=reset_run_state):
    if not validation_errors:
        with st.spinner("최적의 운영 일정을 계산 중입니다..."):
            try:
                cfg = core.build_config(st.session_state)
                
                # solve_for_days가 내부적으로 batched 모드를 처리하도록 함
                # 디버그 모드 활성화 (문제 추적을 위해)
                status, final_wide, logs, limit = solve_for_days(cfg, params, debug=True)
                
                st.session_state['last_solve_logs'] = logs
                st.session_state['solver_status'] = status
                st.session_state['daily_limit'] = limit
                
                if status == "OK" and final_wide is not None and not final_wide.empty:
                    st.session_state['final_schedule'] = final_wide
                    st.balloons()
                else:
                    st.session_state['final_schedule'] = None
            except Exception as e:
                st.error(f"계산 중 오류가 발생했습니다: {str(e)}")
                st.session_state['solver_status'] = "ERROR"

# 결과 표시
st.markdown("---")
status = st.session_state.get('solver_status', '미실행')
daily_limit = st.session_state.get('daily_limit', 0)

col1, col2 = st.columns(2)
with col1:
    st.info(f"Solver Status: `{status}`")
with col2:
    if daily_limit > 0:
        st.info(f"계산된 일일 최대 처리 인원: **{daily_limit}명**")

# 디버그 정보 생성 함수
def generate_debug_info():
    """문제 진단을 위한 종합 디버그 정보 생성"""
    import json
    from datetime import datetime
    
    debug_info = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "solver_status": st.session_state.get('solver_status', '미실행'),
        "daily_limit": st.session_state.get('daily_limit', 0),
        "total_candidates": 0,
        "settings": {}
    }
    
    # 1. 활동 정의
    activities_df = st.session_state.get("activities", pd.DataFrame())
    if not activities_df.empty:
        debug_info["settings"]["activities"] = activities_df.to_dict('records')
        
        # 활동별 통계
        active_activities = activities_df[activities_df["use"] == True]
        debug_info["settings"]["active_activities_count"] = len(active_activities)
        debug_info["settings"]["activity_modes"] = active_activities["mode"].value_counts().to_dict()
    
    # 2. 직무별 인원수와 활동
    job_acts_df = st.session_state.get("job_acts_map", pd.DataFrame())
    if not job_acts_df.empty:
        debug_info["settings"]["job_activities"] = job_acts_df.to_dict('records')
        debug_info["total_candidates"] = int(job_acts_df["count"].sum())
        debug_info["settings"]["job_count"] = len(job_acts_df)
    
    # 3. 선후행 제약
    precedence_df = st.session_state.get("precedence", pd.DataFrame())
    if not precedence_df.empty:
        debug_info["settings"]["precedence"] = precedence_df.to_dict('records')
    else:
        debug_info["settings"]["precedence"] = []
    
    # 4. 방 설정
    room_template = st.session_state.get("room_template", {})
    room_plan_df = st.session_state.get("room_plan", pd.DataFrame())
    
    debug_info["settings"]["room_template"] = room_template
    if not room_plan_df.empty:
        debug_info["settings"]["room_plan"] = room_plan_df.to_dict('records')
    
    # 방별 총 용량 계산
    total_room_capacity = {}
    for room_type, values in room_template.items():
        total_room_capacity[room_type] = {
            "count": values.get("count", 0),
            "capacity_per_room": values.get("cap", 0),
            "total_capacity": values.get("count", 0) * values.get("cap", 0)
        }
    debug_info["settings"]["room_capacity_summary"] = total_room_capacity
    
    # 5. 운영 시간
    oper_window_df = st.session_state.get("oper_window", pd.DataFrame())
    if not oper_window_df.empty:
        debug_info["settings"]["operating_hours"] = oper_window_df.to_dict('records')[0]
    else:
        debug_info["settings"]["operating_hours"] = {
            "start_time": st.session_state.get("oper_start_time", time(9, 0)).strftime("%H:%M"),
            "end_time": st.session_state.get("oper_end_time", time(18, 0)).strftime("%H:%M")
        }
    
    # 운영 시간 계산
    start_time = st.session_state.get("oper_start_time", time(9, 0))
    end_time = st.session_state.get("oper_end_time", time(18, 0))
    operating_minutes = (end_time.hour * 60 + end_time.minute) - (start_time.hour * 60 + start_time.minute)
    debug_info["settings"]["operating_minutes_per_day"] = operating_minutes
    
    # 6. 집단면접 설정 (batched 모드가 있을 때만)
    if "batched" in debug_info["settings"].get("activity_modes", {}):
        debug_info["settings"]["batched_settings"] = {
            "group_min_size": st.session_state.get('group_min_size', 4),
            "group_max_size": st.session_state.get('group_max_size', 6),
            "global_gap_min": st.session_state.get('global_gap_min', 5),
            "max_stay_hours": st.session_state.get('max_stay_hours', 8)
        }
    
    # 7. 예상 처리량 분석
    if activities_df.empty or room_template == {}:
        debug_info["analysis"] = {"error": "활동 또는 방 설정이 없음"}
    else:
        analysis = {}
        active_activities = activities_df[activities_df["use"] == True]
        
        # 각 활동별 병목 분석
        for _, activity in active_activities.iterrows():
            room_type = activity["room_type"]
            if room_type in room_template:
                room_count = room_template[room_type]["count"]
                room_cap = room_template[room_type]["cap"]
                duration = activity["duration_min"]
                
                # 하루 동안 한 방에서 처리 가능한 최대 인원
                slots_per_day = operating_minutes // duration
                max_per_room = slots_per_day * room_cap
                max_total = max_per_room * room_count
                
                analysis[activity["activity"]] = {
                    "duration_min": duration,
                    "room_type": room_type,
                    "room_count": room_count,
                    "room_capacity": room_cap,
                    "max_candidates_per_day": max_total,
                    "slots_per_day_per_room": slots_per_day
                }
        
        debug_info["analysis"] = analysis
        
        # 전체 병목 활동 찾기
        if analysis:
            bottleneck = min(analysis.items(), key=lambda x: x[1]["max_candidates_per_day"])
            debug_info["analysis"]["bottleneck_activity"] = bottleneck[0]
            debug_info["analysis"]["bottleneck_capacity"] = bottleneck[1]["max_candidates_per_day"]
    
    # 8. 마지막 실행 로그 (처음 100줄만)
    last_logs = st.session_state.get('last_solve_logs', "")
    if last_logs:
        log_lines = last_logs.split('\n')[:100]
        debug_info["last_logs_preview"] = '\n'.join(log_lines)
        debug_info["total_log_lines"] = len(last_logs.split('\n'))
    
    return debug_info

# 디버그 정보 표시 (NO_SOLUTION이거나 사용자가 원할 때)
# NO_SOLUTION이나 ERROR일 때는 무조건 표시, 그 외에는 체크박스로 선택
show_debug = False
if status in ["NO_SOLUTION", "ERROR"]:
    show_debug = True
else:
    show_debug = st.checkbox("🔍 디버그 정보 보기", value=False)

if show_debug:
    if status == "ERROR":
        st.error("❌ 스케줄링 중 오류가 발생했습니다")
    elif status == "NO_SOLUTION":
        st.warning("⚠️ 문제 진단을 위한 디버그 정보")
    
    debug_info = generate_debug_info()
    
    # 주요 정보 요약
    st.markdown("### 📊 주요 정보 요약")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("총 지원자", f"{debug_info['total_candidates']}명")
    with col2:
        st.metric("활성 활동", f"{debug_info['settings'].get('active_activities_count', 0)}개")
    with col3:
        st.metric("운영 시간", f"{debug_info['settings'].get('operating_minutes_per_day', 0)}분")
    with col4:
        st.metric("선후행 제약", f"{len(debug_info['settings'].get('precedence', []))}개")
    
    # 병목 분석
    if "bottleneck_activity" in debug_info.get("analysis", {}):
        st.error(f"🚨 병목 활동: **{debug_info['analysis']['bottleneck_activity']}** "
                f"(일일 최대 {debug_info['analysis']['bottleneck_capacity']}명)")
    
    # 구조적 문제 확인
    # Batched/Parallel → Individual adjacent 제약 검사
    activities_df = st.session_state.get("activities", pd.DataFrame())
    precedence_df = st.session_state.get("precedence", pd.DataFrame())
    
    if not activities_df.empty and not precedence_df.empty and 'mode' in activities_df.columns:
        structural_problems = []
        
        for _, rule in precedence_df.iterrows():
            if not rule.get('adjacent', False):
                continue
                
            pred = rule['predecessor']
            succ = rule['successor']
            
            pred_act = activities_df[activities_df['activity'] == pred]
            succ_act = activities_df[activities_df['activity'] == succ]
            
            if pred_act.empty or succ_act.empty:
                continue
                
            pred_mode = pred_act.iloc[0].get('mode', 'individual')
            succ_mode = succ_act.iloc[0].get('mode', 'individual')
            pred_cap = int(pred_act.iloc[0].get('max_cap', 1))
            succ_cap = int(succ_act.iloc[0].get('max_cap', 1))
            
            # Batched → Individual/Parallel adjacent 문제
            if pred_mode == 'batched' and succ_mode in ['individual', 'parallel']:
                # 그룹 크기 확인
                group_size = debug_info['settings'].get('batched_settings', {}).get('group_max_size', 6)
                succ_room_type = succ_act.iloc[0]['room_type']
                succ_room_count = debug_info['settings']['room_template'].get(succ_room_type, {}).get('count', 1)
                
                if group_size > succ_room_count * succ_cap:
                    structural_problems.append({
                        'type': 'batched_to_individual_adjacent',
                        'pred': pred,
                        'succ': succ,
                        'group_size': group_size,
                        'succ_capacity': succ_room_count * succ_cap,
                        'message': f"{pred}(그룹 {group_size}명) → {succ}(동시 {succ_room_count * succ_cap}명) adjacent 불가능"
                    })
            
            # Parallel → Individual adjacent 문제
            elif pred_mode == 'parallel' and succ_mode == 'individual':
                succ_room_type = succ_act.iloc[0]['room_type']
                succ_room_count = debug_info['settings']['room_template'].get(succ_room_type, {}).get('count', 1)
                
                if pred_cap > succ_room_count:
                    structural_problems.append({
                        'type': 'parallel_to_individual_adjacent',
                        'pred': pred,
                        'succ': succ,
                        'pred_capacity': pred_cap,
                        'succ_capacity': succ_room_count,
                        'message': f"{pred}(동시 {pred_cap}명) → {succ}(동시 {succ_room_count}명) adjacent 불가능"
                    })
        
        if structural_problems:
            st.error("### 🚫 구조적 문제 발견")
            for problem in structural_problems:
                st.error(f"**{problem['message']}**")
                
                if problem['type'] == 'batched_to_individual_adjacent':
                    st.markdown(f"""
                    **문제 상세:**
                    - {problem['pred']}은 **{problem['group_size']}명이 한 그룹**으로 활동
                    - {problem['succ']}은 **동시에 {problem['succ_capacity']}명만** 수용 가능
                    - Adjacent 제약으로 인해 {problem['pred']} 직후 {problem['succ']}를 해야 함
                    - **{problem['group_size'] - problem['succ_capacity']}명이 대기**해야 하므로 adjacent 제약 위반
                    
                    **해결 방법:**
                    1. **Adjacent 제약 제거**: {problem['pred']} → {problem['succ']}의 'adjacent'를 false로 변경
                    2. **방 증설**: {problem['succ']} 방을 {problem['group_size'] // int(succ_cap) + (1 if problem['group_size'] % int(succ_cap) else 0)}개로 증가
                    3. **그룹 크기 축소**: 집단면접 그룹을 {problem['succ_capacity']}명 이하로 설정
                    """)
                elif problem['type'] == 'parallel_to_individual_adjacent':
                    st.markdown(f"""
                    **문제 상세:**
                    - {problem['pred']}은 **동시에 {problem['pred_capacity']}명** 진행
                    - {problem['succ']}은 **동시에 {problem['succ_capacity']}명만** 수용 가능
                    - Adjacent 제약으로 인해 나머지 **{problem['pred_capacity'] - problem['succ_capacity']}명이 대기**
                    
                    **해결 방법:**
                    1. **Adjacent 제약 제거**: 더 유연한 스케줄링 허용
                    2. **방 증설**: {problem['succ']} 방을 {problem['pred_capacity']}개로 증가
                    3. **용량 조정**: {problem['pred']}의 max_cap을 {problem['succ_capacity']}로 축소
                    """)
            
            st.info("💡 **권장사항**: Adjacent 제약은 물리적으로 연속된 공간이나 즉각적인 이동이 필요한 경우에만 사용하세요.")
    
    # 디버그 정보 JSON 표시
    with st.expander("🔧 전체 디버그 정보 (복사해서 공유 가능)", expanded=True):
        # JSON 형태로 예쁘게 출력
        debug_json = json.dumps(debug_info, indent=2, ensure_ascii=False)
        st.code(debug_json, language="json")
        
        # 복사 버튼
        st.download_button(
            label="📥 디버그 정보 다운로드",
            data=debug_json,
            file_name=f"debug_info_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )
    


if "솔버 시간 초과" in st.session_state.get('last_solve_logs', ''):
    st.warning("⚠️ 연산 시간이 2분(120초)을 초과하여, 현재까지 찾은 최적의 스케줄을 반환했습니다. 결과는 최상이 아닐 수 있습니다.")

# 결과 출력
final_schedule = st.session_state.get('final_schedule')
if final_schedule is not None and not final_schedule.empty:
    st.success("🎉 운영일정 추정이 완료되었습니다!")
    
    # 요약 정보
    total_candidates = len(final_schedule)
    total_days = final_schedule['interview_date'].nunique()
    st.info(f"총 {total_candidates}명의 지원자를 {total_days}일에 걸쳐 면접 진행")
    
    # 체류시간 분석 추가
    st.subheader("⏱️ 직무별 체류시간 분석")
    
    def calculate_stay_duration_stats(schedule_df):
        """각 지원자의 체류시간을 계산하고 직무별 통계를 반환"""
        stats_data = []
        
        # 컬럼명 매핑 (실제 데이터에 맞게 조정)
        id_col = 'id' if 'id' in schedule_df.columns else 'candidate_id'
        job_col = 'code' if 'code' in schedule_df.columns else 'job_code'
        
        # 시간 컬럼들 찾기 (start_활동명, end_활동명 형태)
        start_cols = [col for col in schedule_df.columns if col.startswith('start_')]
        end_cols = [col for col in schedule_df.columns if col.startswith('end_')]
        
        if not start_cols or not end_cols:
            st.error("시간 정보 컬럼을 찾을 수 없습니다. start_* 또는 end_* 형태의 컬럼이 필요합니다.")
            return pd.DataFrame(), pd.DataFrame()
        
        # 지원자별로 그룹화하여 체류시간 계산
        for candidate_id, candidate_data in schedule_df.groupby(id_col):
            # 해당 지원자의 모든 활동 시간 정보 수집
            all_start_times = []
            all_end_times = []
            
            for _, row in candidate_data.iterrows():
                for start_col in start_cols:
                    if pd.notna(row[start_col]) and row[start_col] != '':
                        try:
                            # 시간 문자열을 datetime으로 변환
                            time_str = str(row[start_col])
                            if ':' in time_str:
                                # HH:MM:SS 또는 HH:MM 형태
                                time_obj = pd.to_datetime(time_str, format='%H:%M:%S', errors='coerce')
                                if pd.isna(time_obj):
                                    time_obj = pd.to_datetime(time_str, format='%H:%M', errors='coerce')
                                if not pd.isna(time_obj):
                                    all_start_times.append(time_obj)
                        except:
                            continue
                
                for end_col in end_cols:
                    if pd.notna(row[end_col]) and row[end_col] != '':
                        try:
                            # 시간 문자열을 datetime으로 변환
                            time_str = str(row[end_col])
                            if ':' in time_str:
                                # HH:MM:SS 또는 HH:MM 형태
                                time_obj = pd.to_datetime(time_str, format='%H:%M:%S', errors='coerce')
                                if pd.isna(time_obj):
                                    time_obj = pd.to_datetime(time_str, format='%H:%M', errors='coerce')
                                if not pd.isna(time_obj):
                                    all_end_times.append(time_obj)
                        except:
                            continue
            
            if all_start_times and all_end_times:
                # 전체 체류시간 = 첫 번째 활동 시작 ~ 마지막 활동 종료
                total_start = min(all_start_times)
                total_end = max(all_end_times)
                stay_duration_minutes = (total_end - total_start).total_seconds() / 60
                
                # 직무 코드 (첫 번째 행에서 가져오기)
                job_code = candidate_data.iloc[0].get(job_col, 'Unknown')
                
                stats_data.append({
                    'candidate_id': candidate_id,
                    'job_code': job_code,
                    'stay_duration_minutes': stay_duration_minutes,
                    'start_time': total_start,
                    'end_time': total_end
                })
        
        if not stats_data:
            st.warning("체류시간을 계산할 수 있는 유효한 데이터가 없습니다.")
            return pd.DataFrame(), pd.DataFrame()
        
        stats_df = pd.DataFrame(stats_data)
        
        # 직무별 통계 계산
        job_stats = []
        for job_code, job_data in stats_df.groupby('job_code'):
            durations = job_data['stay_duration_minutes']
            job_stats.append({
                'job_code': job_code,
                'count': len(job_data),
                'min_duration': durations.min(),
                'max_duration': durations.max(),
                'avg_duration': durations.mean(),
                'median_duration': durations.median()
            })
        
        return pd.DataFrame(job_stats), stats_df
    
    try:
        job_stats_df, individual_stats_df = calculate_stay_duration_stats(final_schedule)
        
        if not job_stats_df.empty:
            # 직무별 통계 표시
            st.markdown("**📊 직무별 체류시간 통계 (분 단위)**")
            
            # 표시용 데이터프레임 생성
            display_stats = job_stats_df.copy()
            display_stats['min_duration'] = display_stats['min_duration'].round(1)
            display_stats['max_duration'] = display_stats['max_duration'].round(1)
            display_stats['avg_duration'] = display_stats['avg_duration'].round(1)
            display_stats['median_duration'] = display_stats['median_duration'].round(1)
            
            # 컬럼명 한글화
            display_stats.columns = ['직무코드', '인원수', '최소시간(분)', '최대시간(분)', '평균시간(분)', '중간값(분)']
            
            st.dataframe(display_stats, use_container_width=True)
            
            # 시각적 요약
            col1, col2, col3 = st.columns(3)
            with col1:
                overall_min = job_stats_df['min_duration'].min()
                st.metric("전체 최소 체류시간", f"{overall_min:.1f}분")
            with col2:
                overall_max = job_stats_df['max_duration'].max()
                st.metric("전체 최대 체류시간", f"{overall_max:.1f}분")
            with col3:
                overall_avg = (job_stats_df['avg_duration'] * job_stats_df['count']).sum() / job_stats_df['count'].sum()
                st.metric("전체 평균 체류시간", f"{overall_avg:.1f}분")
            
            # 체류시간 제한 확인
            max_stay_minutes = params.get('max_stay_hours', 8) * 60
            if overall_max > max_stay_minutes:
                st.warning(f"⚠️ 일부 지원자의 체류시간이 설정된 제한({params.get('max_stay_hours', 8)}시간)을 초과했습니다.")
    
    except Exception as e:
        st.error(f"체류시간 분석 중 오류 발생: {str(e)}")
    
    # 스케줄 표시 옵션
    col1, col2 = st.columns([3, 1])
    with col1:
        # 날짜별 요약 정보
        date_summary = final_schedule.groupby('interview_date').size().reset_index(name='인원수')
        date_summary['interview_date'] = pd.to_datetime(date_summary['interview_date']).dt.strftime('%Y-%m-%d')
        date_summary.columns = ['날짜', '인원수']
        
        st.markdown("**📅 날짜별 면접 인원**")
        st.dataframe(date_summary, use_container_width=True, hide_index=True)
    
    with col2:
        st.markdown("**💾 결과 다운로드**")
        excel_buffer = BytesIO()
        
        # 그룹 정보 가져오기 (있다면)
        group_info_data = st.session_state.get('last_group_info', None)
        
        df_to_excel(final_schedule, excel_buffer, group_info_data)
        excel_buffer.seek(0)
        
        st.download_button(
            label="📥 Excel 다운로드",
            data=excel_buffer,
            file_name=f"interview_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    # 상세 스케줄 표시
    with st.expander("📋 상세 스케줄 보기", expanded=False):
        # 날짜별로 탭 생성
        dates = sorted(final_schedule['interview_date'].unique())
        tabs = st.tabs([pd.to_datetime(d).strftime('%Y-%m-%d') for d in dates])
        
        for i, (tab, date) in enumerate(zip(tabs, dates)):
            with tab:
                day_schedule = final_schedule[final_schedule['interview_date'] == date].copy()
                
                # 시간 컬럼들을 더 읽기 쉽게 표시
                display_cols = ['id', 'code']
                for col in day_schedule.columns:
                    if col.startswith(('start_', 'end_', 'loc_')) and col not in display_cols:
                        display_cols.append(col)
                
                day_schedule = day_schedule[display_cols]
                
                # 인덱스 숨기고 표시
                st.dataframe(day_schedule, use_container_width=True, hide_index=True)
                
                # batched 모드가 있으면 그룹 정보도 표시
                if has_batched:
                    # TODO: 그룹 정보 표시 구현
                    pass

st.divider()

# =============================================================================
# 섹션 1: 면접활동 정의
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("1️⃣ 면접활동 정의")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_activities", help="활동 정의 AG-Grid가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "activities" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["activities"] = 0
        st.session_state["section_refresh_counter"]["activities"] += 1
        st.rerun()

st.markdown("면접 프로세스에 포함될 활동들을 정의합니다. 각 활동의 모드, 소요시간, 필요한 공간 등을 설정하세요.")

# 기본 템플릿 함수
def default_df() -> pd.DataFrame:
    return pd.DataFrame({
        "use": [True, True, True, True],
        "activity": ["면접1", "면접2", "인성검사", "커피챗"],
        "mode": ["individual"] * 4,
        "duration_min": [10] * 4,
        "room_type": ["면접1실", "면접2실", "인성검사실", "커피챗실"],
        "min_cap": [1] * 4,
        "max_cap": [1] * 4,
    })

df = st.session_state["activities"].copy()

# 누락 컬럼 보강
for col, default_val in {
    "use": True,
    "mode": "individual",
    "duration_min": 10,
    "room_type": "",
    "min_cap": 1,
    "max_cap": 1,
}.items():
    if col not in df.columns:
        df[col] = default_val

# AG-Grid 설정
gb = GridOptionsBuilder.from_dataframe(df)

gb.configure_column(
    "use",
    header_name="사용",
    cellEditor="agCheckboxCellEditor",
    cellRenderer="agCheckboxCellRenderer",
    editable=True,
    singleClickEdit=True,
    width=80,
)

gb.configure_column("activity", header_name="활동 이름", editable=True)

# mode_values에 parallel과 batched 추가
mode_values = ["individual", "parallel", "batched"]
gb.configure_column(
    "mode",
    header_name="모드",
    editable=True,
    cellEditor="agSelectCellEditor",
    cellEditorParams={"values": mode_values},
    width=110,
)

# duration_min은 항상 편집 가능
gb.configure_column(
    "duration_min",
    header_name="소요시간(분)",
    editable=True,
    type=["numericColumn", "numberColumnFilter"],
    width=120,
)

# min_cap, max_cap은 조건부 편집 가능 (individual 모드에서는 비활성화)
for col, hdr in [("min_cap", "최소 인원"), ("max_cap", "최대 인원")]:
    gb.configure_column(
        col,
        header_name=hdr,
        editable=True,
        type=["numericColumn", "numberColumnFilter"],
        width=120,
        cellEditor="agNumberCellEditor",
        cellEditorParams={
            "min": 1,
            "max": 50
        },
        # individual 모드일 때 편집 불가능하게 하는 조건
        cellClassRules={
            "ag-cell-not-editable": "data.mode === 'individual'"
        },
        # individual 모드일 때 회색으로 표시
        cellStyle={
            "backgroundColor": "params.data.mode === 'individual' ? '#f0f0f0' : 'white'",
            "color": "params.data.mode === 'individual' ? '#888' : 'black'"
        }
    )

gb.configure_column("room_type", header_name="면접실 이름", editable=True)

grid_opts = gb.build()

st.markdown("#### 활동 정의")

# AG-Grid에서 비활성화된 셀 스타일링을 위한 CSS 추가
st.markdown("""
<style>
.ag-cell-not-editable {
    background-color: #f5f5f5 !important;
    color: #999 !important;
    cursor: not-allowed !important;
}
.ag-cell-not-editable:hover {
    background-color: #f0f0f0 !important;
}
</style>
""", unsafe_allow_html=True)

# 모드 설명 추가
with st.expander("ℹ️ 모드 설명", expanded=False):
    st.markdown("""
    - **individual**: 1명이 혼자 면접 (기존 방식)
      - 최소/최대 인원이 자동으로 1명으로 고정됩니다 (수정 불가)
    - **parallel**: 여러명이 같은 공간에서 각자 다른 활동 (예: 개별 작업)
      - 최소/최대 인원을 자유롭게 설정 가능합니다
    - **batched**: 여러명이 동시에 같은 활동 (예: 그룹토론, PT발표)
      - 최소/최대 인원을 자유롭게 설정 가능합니다
    
    **주의**: 
    - **individual 모드**: 인원수 수정 불가 (1명 고정)
    - **batched 모드**: 모든 batched 활동의 min_cap, max_cap은 동일해야 합니다
    """)

# 행 추가/삭제 기능 (위로 이동)
col_add, col_del = st.columns(2)

with col_add:
    if st.button("➕ 활동 행 추가", key="add_activity"):
        new_row = {
            "use": True,
            "activity": "NEW_ACT",
            "mode": "individual",
            "duration_min": 10,
            "room_type": "",
            "min_cap": 1,
            "max_cap": 1,
        }
        st.session_state["activities"] = pd.concat(
            [st.session_state["activities"], pd.DataFrame([new_row])],
            ignore_index=True,
        )
        st.rerun()

with col_del:
    act_df = st.session_state["activities"].copy()
    if not act_df.empty:
        # 인덱스와 활동명을 안전하게 결합
        delete_options = []
        valid_indices = []
        for idx, row in act_df.iterrows():
            activity_name = str(row.get('activity', 'Unknown'))
            if activity_name and activity_name != 'nan':
                delete_options.append(f"{idx}: {activity_name}")
                valid_indices.append(idx)
        
        to_delete = st.multiselect(
            "삭제할 활동 선택",
            options=delete_options,
            key="del_activity_select"
        )
        if st.button("❌ 선택된 활동 삭제", key="del_activity"):
            if to_delete:
                # 선택된 인덱스 추출
                selected_indices = [int(s.split(":")[0]) for s in to_delete]
                
                # 실제 DataFrame에 존재하는 인덱스만 필터링
                valid_to_drop = [idx for idx in selected_indices if idx in act_df.index]
                
                if valid_to_drop:
                    kept = st.session_state["activities"].drop(valid_to_drop).reset_index(drop=True)
                    st.session_state["activities"] = kept
                    st.success(f"선택된 {len(valid_to_drop)}개 활동이 삭제되었습니다.")
                    st.rerun()
                else:
                    st.error("삭제할 유효한 활동이 없습니다.")
    else:
        st.info("삭제할 활동이 없습니다.")

# AG-Grid 표시 (행 추가/삭제 기능 아래로 이동)
# 섹션별 새로고침을 위한 동적 key 생성
activities_refresh_count = st.session_state.get("section_refresh_counter", {}).get("activities", 0)

grid_ret = AgGrid(
    df,
    gridOptions=grid_opts,
    data_return_mode=DataReturnMode.AS_INPUT,
    update_mode=GridUpdateMode.VALUE_CHANGED,
    allow_unsafe_jscode=True,
    fit_columns_on_grid_load=True,
    theme="balham",
    key=f"activities_grid_{activities_refresh_count}",  # 동적 key로 강제 재렌더링
)

# 그리드 데이터 처리 및 individual 모드 강제 조정
activities_data = pd.DataFrame(grid_ret["data"])

# individual 모드인 행들의 min_cap, max_cap을 1로 고정
individual_mask = activities_data["mode"] == "individual"
activities_data.loc[individual_mask, "min_cap"] = 1
activities_data.loc[individual_mask, "max_cap"] = 1

st.session_state["activities"] = activities_data

# Activities → Room Plan 자동 연동
activities_df = st.session_state["activities"]
room_plan_df = st.session_state.get("room_plan", pd.DataFrame())
room_template = st.session_state.get("room_template", {})

if not activities_df.empty and (not room_plan_df.empty or room_template):
    # 각 room_type별 최대 max_cap 계산
    room_capacity_updates = {}
    
    for room_type in activities_df["room_type"].unique():
        if pd.notna(room_type) and room_type != "":
            # 해당 room_type을 사용하는 활동들의 최대 max_cap
            max_cap_for_room = activities_df[activities_df["room_type"] == room_type]["max_cap"].max()
            
            if pd.notna(max_cap_for_room):
                room_capacity_updates[room_type] = int(max_cap_for_room)
    
    # room_template 업데이트
    template_updated = False
    for room_type, new_capacity in room_capacity_updates.items():
        if room_type in room_template:
            if room_template[room_type]["cap"] < new_capacity:
                room_template[room_type]["cap"] = new_capacity
                template_updated = True
    
    # room_plan DataFrame 업데이트
    plan_updated = False
    if not room_plan_df.empty:
        for room_type, new_capacity in room_capacity_updates.items():
            cap_col = f"{room_type}_cap"
            if cap_col in room_plan_df.columns:
                for idx in room_plan_df.index:
                    current_cap = room_plan_df.at[idx, cap_col]
                    if pd.notna(current_cap) and current_cap < new_capacity:
                        room_plan_df.at[idx, cap_col] = new_capacity
                        plan_updated = True
    
    # 변경사항 저장 및 알림
    if template_updated:
        st.session_state["room_template"] = room_template
    if plan_updated:
        st.session_state["room_plan"] = room_plan_df
    
    if template_updated or plan_updated:
        updated_rooms = [f"{rt}({cap}명)" for rt, cap in room_capacity_updates.items()]
        st.info(f"📝 활동 그룹 크기 변경에 따라 방 수용 인원이 자동 조정되었습니다: {', '.join(updated_rooms)}")

# batched 모드 일관성 검증
batched_acts = st.session_state["activities"][st.session_state["activities"]["mode"] == "batched"]
if not batched_acts.empty:
    min_caps = batched_acts["min_cap"].unique()
    max_caps = batched_acts["max_cap"].unique()
    
    if len(min_caps) > 1 or len(max_caps) > 1:
        st.error("⚠️ 모든 batched 활동의 그룹 크기(min_cap, max_cap)는 동일해야 합니다!")
        st.info("현재 설정: " + 
                f"min_cap = {list(min_caps)}, " +
                f"max_cap = {list(max_caps)}")

st.divider()

# =============================================================================
# 집단면접 설정 (batched 모드가 있을 때만 표시)
# =============================================================================
acts_df = st.session_state.get("activities", pd.DataFrame())
has_batched = any(acts_df["mode"] == "batched") if not acts_df.empty and "mode" in acts_df.columns else False

if has_batched:
    with st.expander("🎯 집단면접 설정", expanded=True):
        st.markdown("집단면접(batched) 활동에 대한 전역 설정을 관리합니다.")
        
        # batched 활동 목록 표시
        batched_activities = acts_df[acts_df["mode"] == "batched"]["activity"].tolist()
        st.info(f"집단면접 활동: {', '.join(batched_activities)}")
        
        # 그룹 크기 설정
        col1, col2 = st.columns(2)
        
        # batched 활동 필터링 추가
        batched_acts = acts_df[acts_df["mode"] == "batched"]
        
        with col1:
            # 기존 값 가져오기
            current_min = st.session_state.get('group_min_size', 4)
            # batched 활동이 있으면 그 값 사용
            if not batched_acts.empty:
                current_min = int(batched_acts.iloc[0]['min_cap'])
            
            # min_value보다 작으면 기본값 사용
            if current_min < 2:
                current_min = 4
            
            group_min = st.number_input(
                "그룹 최소 인원",
                min_value=2,
                max_value=20,
                value=current_min,
                key="group_min_input",
                help="모든 집단면접 활동에 적용됩니다"
            )
        
        with col2:
            # 기존 값 가져오기
            current_max = st.session_state.get('group_max_size', 6)
            # batched 활동이 있으면 그 값 사용
            if not batched_acts.empty:
                current_max = int(batched_acts.iloc[0]['max_cap'])
            
            # min_value보다 작으면 기본값 사용
            if current_max < group_min:
                current_max = max(6, group_min)
            
            group_max = st.number_input(
                "그룹 최대 인원",
                min_value=group_min,
                max_value=30,
                value=current_max,
                key="group_max_input",
                help="모든 집단면접 활동에 적용됩니다"
            )
        
        # 값이 변경되면 모든 batched 활동에 적용
        if group_min != st.session_state.get('group_min_size') or group_max != st.session_state.get('group_max_size'):
            st.session_state['group_min_size'] = group_min
            st.session_state['group_max_size'] = group_max
            
            # activities DataFrame 업데이트
            acts_df = st.session_state["activities"]
            acts_df.loc[acts_df["mode"] == "batched", "min_cap"] = group_min
            acts_df.loc[acts_df["mode"] == "batched", "max_cap"] = group_max
            st.session_state["activities"] = acts_df
            
            # 집단면접 설정 → Room Plan 자동 연동
            room_plan_df = st.session_state.get("room_plan", pd.DataFrame())
            room_template = st.session_state.get("room_template", {})
            
            if not room_plan_df.empty and room_template:
                # batched 활동이 사용하는 room_type들의 capacity 업데이트
                batched_room_types = acts_df[acts_df["mode"] == "batched"]["room_type"].unique()
                
                room_updated = False
                for room_type in batched_room_types:
                    if pd.notna(room_type) and room_type != "":
                        # room_template 업데이트
                        if room_type in room_template:
                            if room_template[room_type]["cap"] < group_max:
                                room_template[room_type]["cap"] = group_max
                                room_updated = True
                        
                        # room_plan DataFrame 업데이트
                        cap_col = f"{room_type}_cap"
                        if cap_col in room_plan_df.columns:
                            for idx in room_plan_df.index:
                                current_cap = room_plan_df.at[idx, cap_col]
                                if pd.notna(current_cap) and current_cap < group_max:
                                    room_plan_df.at[idx, cap_col] = group_max
                                    room_updated = True
                
                if room_updated:
                    st.session_state["room_template"] = room_template
                    st.session_state["room_plan"] = room_plan_df
                    st.info(f"📝 집단면접 그룹 크기 변경에 따라 관련 방의 수용 인원이 {group_max}명으로 자동 조정되었습니다.")
            
            st.success(f"✅ 모든 집단면접 활동의 그룹 크기가 {group_min}~{group_max}명으로 설정되었습니다.")
        
        # 추가 설정들
        st.markdown("### 고급 설정")
        
        col3, col4 = st.columns(2)
        
        with col3:
            global_gap = st.number_input(
                "전역 활동 간격(분)",
                min_value=0,
                max_value=60,
                value=st.session_state.get('global_gap_min', 5),
                key="global_gap_input",
                help="모든 활동 간 기본 간격입니다. Precedence에서 개별 설정 가능합니다."
            )
            st.session_state['global_gap_min'] = global_gap
        
        with col4:
            max_stay = st.number_input(
                "최대 체류시간(시간)",
                min_value=1,
                max_value=12,
                value=st.session_state.get('max_stay_hours', 8),
                key="max_stay_input",
                help="지원자가 면접장에 머무를 수 있는 최대 시간입니다."
            )
            st.session_state['max_stay_hours'] = max_stay
        
        # 직무별 분리 원칙 표시
        st.markdown("### 그룹 구성 원칙")
        st.info("""
        ✅ **직무별 분리**: 같은 직무끼리만 그룹 구성
        ✅ **그룹 일관성**: 한 번 구성된 그룹은 모든 batched 활동에서 유지
        ✅ **동일 직무 동일 방**: 같은 직무는 모든 활동에서 동일한 접미사(A,B,C...) 사용
        ✅ **그룹 수 최소화**: 더미 지원자를 활용하여 그룹 수 최소화
        """)

    st.divider()

# =============================================================================
# 섹션 2: 선후행 제약 설정
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("2️⃣ 선후행 제약 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_precedence", help="선후행 제약 UI가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "precedence" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["precedence"] = 0
        st.session_state["section_refresh_counter"]["precedence"] += 1
        st.rerun()

st.markdown("면접 활동 간의 순서 제약과 시간 간격을 설정합니다.")

# 공통 데이터 로드
acts_df = st.session_state.get("activities", pd.DataFrame())
jobs_df = st.session_state.get("job_acts_map", pd.DataFrame())

if not acts_df.empty:
    ACT_OPTS = acts_df.query("use == True")["activity"].tolist()
    
    # precedence 규칙 초기화 (빈 문자열도 허용)
    prec_df = st.session_state["precedence"].copy()
    valid_acts = set(ACT_OPTS) | {"__START__", "__END__", ""}
    prec_df = prec_df[prec_df["predecessor"].isin(valid_acts) & prec_df["successor"].isin(valid_acts)]
    st.session_state["precedence"] = prec_df
    
    # 동선 미리보기 함수
    def generate_flow_preview():
        """현재 선후행 제약 규칙을 바탕으로 가능한 동선을 시각화"""
        prec_rules = st.session_state["precedence"].copy()
        if prec_rules.empty:
            return "📋 설정된 제약 규칙이 없습니다. 모든 활동이 자유롭게 배치 가능합니다."
        
        # START와 END 규칙 분리
        start_rules = prec_rules[prec_rules["predecessor"] == "__START__"]
        end_rules = prec_rules[prec_rules["successor"] == "__END__"]
        middle_rules = prec_rules[
            (~prec_rules["predecessor"].isin(["__START__", "__END__"])) &
            (~prec_rules["successor"].isin(["__START__", "__END__"]))
        ]
        
        flow_text = "🔄 **예상 면접 동선:**\n\n"
        
        # START 규칙이 있는 경우
        if not start_rules.empty:
            first_acts = start_rules["successor"].tolist()
            flow_text += f"**시작:** 🏁 → {' / '.join(first_acts)}\n\n"
        else:
            flow_text += "**시작:** 🏁 → (모든 활동 가능)\n\n"
        
        # 중간 규칙들
        if not middle_rules.empty:
            flow_text += "**중간 연결:**\n"
            for _, rule in middle_rules.iterrows():
                gap_info = f" ({rule['gap_min']}분 간격)" if rule['gap_min'] > 0 else ""
                adj_info = " [인접]" if rule['adjacent'] else ""
                flow_text += f"• {rule['predecessor']} → {rule['successor']}{gap_info}{adj_info}\n"
            flow_text += "\n"
        
        # END 규칙이 있는 경우
        if not end_rules.empty:
            last_acts = end_rules["predecessor"].tolist()
            flow_text += f"**종료:** {' / '.join(last_acts)} → 🏁"
        else:
            flow_text += "**종료:** (모든 활동 가능) → 🏁"
        
        return flow_text
    
    # 가능한 모든 활동 순서 계산 함수 (과거 코드 복원)
    def render_dynamic_flows(prec_df: pd.DataFrame, base_nodes: list[str]) -> list[str]:
        """
        prec_df: ['predecessor','successor','gap_min','adjacent']
        base_nodes: 순서에 포함될 활동 리스트
        """
        import itertools
        
        # 1) 규칙(rules)에 adjacent까지 함께 읽어오기
        rules = [
            (row.predecessor, row.successor,
             int(row.gap_min),
             bool(getattr(row, "adjacent", False)))
            for row in prec_df.itertuples()
        ]
        n = len(base_nodes)
        valid_orders = []

        # 2) 인터뷰 느낌 이모지 풀(10개) + 동적 매핑
        emoji_pool = ["📝","🧑‍💼","🎤","💼","🗣️","🤝","🎯","🔎","📋","⏰"]
        icons = { act: emoji_pool[i % len(emoji_pool)]
                  for i, act in enumerate(base_nodes) }

        # 3) 모든 순열 검사
        for perm in itertools.permutations(base_nodes, n):
            ok = True
            for p, s, gap, adj in rules:
                # START → S
                if p == "__START__":
                    if perm[0] != s:
                        ok = False
                    if not ok: break
                    else: continue
                # P → END
                if s == "__END__":
                    if perm[-1] != p:
                        ok = False
                    if not ok: break
                    else: continue
                # 일반 활동 간 제약
                if p in perm and s in perm:
                    i_p, i_s = perm.index(p), perm.index(s)
                    # 붙이기(adjacent) 또는 gap>0 모두 "인접" 처리
                    if adj or gap > 0:
                        if i_s != i_p + 1:
                            ok = False
                            break
                    else:
                        # gap==0: 순서만 보장
                        if i_p >= i_s:
                            ok = False
                            break
            if ok:
                valid_orders.append(perm)

        # 4) 문자열로 변환 (아이콘 + 활동명)
        flow_strs = []
        for order in valid_orders:
            labels = [f"{icons[act]} {act}" for act in order]
            flow_strs.append(" ➔ ".join(labels))
        return flow_strs
    
    # 실시간 동선(활동 순서) 미리보기
    with st.expander("🔍 실시간 동선(활동 순서) 미리보기", expanded=True):
        prec_df_latest = st.session_state["precedence"]
        
        # 기본 규칙 표시
        st.markdown("**📋 현재 설정된 선후행 제약 규칙:**")
        if prec_df_latest.empty:
            st.info("설정된 제약 규칙이 없습니다. 모든 활동이 자유롭게 배치 가능합니다.")
        else:
            # 규칙을 사용자 친화적으로 표시
            for _, rule in prec_df_latest.iterrows():
                pred = rule['predecessor']
                succ = rule['successor']
                gap = rule['gap_min']
                adj = rule['adjacent']
                
                # 표시 형식 개선
                if pred == "__START__":
                    pred_display = "🏁 시작"
                elif pred == "":
                    pred_display = "🏁 시작"
                else:
                    pred_display = f"📝 {pred}"
                
                if succ == "__END__":
                    succ_display = "🏁 종료"
                elif succ == "":
                    succ_display = "🏁 종료"
                else:
                    succ_display = f"📝 {succ}"
                
                gap_info = f" ({gap}분 간격)" if gap > 0 else ""
                adj_info = " [인접 배치]" if adj else ""
                
                st.markdown(f"• {pred_display} → {succ_display}{gap_info}{adj_info}")
        
        st.markdown("---")
        
        # 가능한 활동 순서 계산 및 표시
        if ACT_OPTS:
            flows = render_dynamic_flows(prec_df_latest, ACT_OPTS)
            if not flows:
                st.warning("⚠️ 현재 제약을 만족하는 활동 순서가 없습니다. 제약 조건을 확인해주세요.")
            else:
                st.markdown("**🔄 가능한 모든 활동 순서:**")
                for i, f in enumerate(flows, 1):
                    st.markdown(f"{i}. {f}")
                
                if len(flows) == 1:
                    st.success("✅ 제약 조건에 따라 활동 순서가 고유하게 결정됩니다!")
                else:
                    st.info(f"💡 총 {len(flows)}가지 가능한 활동 순서가 있습니다.")
        else:
            st.warning("활성화된 활동이 없습니다. 먼저 활동을 정의해주세요.")

    
    # ═══════════════════════════════════════════
    # 🎯 면접 순서 설정 (단계별 가이드)
    # ═══════════════════════════════════════════

    # 현재 설정된 START/END 규칙 확인
    current_start = None
    current_end = None
    for _, rule in prec_df.iterrows():
        if rule['predecessor'] == "__START__":
            current_start = rule['successor']
        if rule['successor'] == "__END__":
            current_end = rule['predecessor']
    
    # 섹션별 새로고침을 위한 동적 key 생성
    precedence_refresh_count = st.session_state.get("section_refresh_counter", {}).get("precedence", 0)
    
    st.markdown("---")
    st.subheader("🎯 면접 순서 설정")
    st.markdown("면접 활동들의 순서와 제약을 설정합니다.")
    
    # 탭으로 기능 구분
    tab1, tab2 = st.tabs(["🏁 시작/끝 규칙", "🔗 순서 규칙"])
    
    with tab1:
        st.markdown("💡 **면접의 첫 번째와 마지막 활동을 지정하세요.** (선택사항)")
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        # 현재 값을 기본값으로 설정
        start_idx = 0
        end_idx = 0
        if current_start and current_start in ACT_OPTS:
            start_idx = ACT_OPTS.index(current_start) + 1
        if current_end and current_end in ACT_OPTS:
            end_idx = ACT_OPTS.index(current_end) + 1
        
        with col1:
            first = st.selectbox(
                "🏁 가장 먼저 할 활동", 
                ["(지정 안 함)"] + ACT_OPTS, 
                index=start_idx, 
                key=f"first_act_{precedence_refresh_count}",
                help="면접 프로세스의 첫 번째 활동을 선택하세요"
            )
        
        with col2:
            last = st.selectbox(
                "🏁 가장 마지막 활동", 
                ["(지정 안 함)"] + ACT_OPTS, 
                index=end_idx, 
                key=f"last_act_{precedence_refresh_count}",
                help="면접 프로세스의 마지막 활동을 선택하세요"
            )
        
        with col3:
            st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
            if st.button("✅ 적용", key="btn_add_start_end", type="primary", use_container_width=True):
                # 기존 __START__/__END__ 관련 행 제거
                tmp = prec_df[
                    (~prec_df["predecessor"].isin(["__START__", "__END__"])) &
                    (~prec_df["successor"].isin(["__START__", "__END__"]))
                ].copy()
                
                rows = []
                if first != "(지정 안 함)":
                    rows.append({"predecessor": "__START__", "successor": first, "gap_min": 0, "adjacent": True})
                if last != "(지정 안 함)":
                    rows.append({"predecessor": last, "successor": "__END__", "gap_min": 0, "adjacent": True})
                
                st.session_state["precedence"] = pd.concat([tmp, pd.DataFrame(rows)], ignore_index=True)
                st.success("✅ 시작/끝 활동이 설정되었습니다!")
                st.rerun()
    
    with tab2:
        st.markdown("💡 **특정 활동 다음에 반드시 와야 하는 활동을 연결하세요.** (선택사항)")
        st.markdown("📝 **예시:** 면접1 → 면접2 (면접1 후에 반드시 면접2가 와야 함)")
        
        # 기본 연결 설정
        st.markdown("#### 📌 활동 연결")
        col_form1, col_form2, col_form3 = st.columns([2, 2, 1])
        
        with col_form1:
            p = st.selectbox("🔸 먼저 할 활동", ACT_OPTS, key=f"pred_select_{precedence_refresh_count}", help="선행 활동을 선택하세요")
        
        with col_form2:
            s = st.selectbox("🔹 다음에 할 활동", ACT_OPTS, key=f"succ_select_{precedence_refresh_count}", help="후행 활동을 선택하세요")
        
        with col_form3:
            st.markdown("<br>", unsafe_allow_html=True)  # 버튼 높이 맞추기
            add_rule_btn = st.button("✅ 적용", key="btn_add_sequence", type="primary", use_container_width=True)
        
        # 고급 옵션을 별도 영역으로 분리
        st.markdown("#### ⚙️ 고급 옵션")
        col_gap, col_adj = st.columns(2)
        with col_gap:
            g = st.number_input("⏱️ 최소 간격 (분)", 0, 60, 5, key=f"gap_input_{precedence_refresh_count}", help="두 활동 사이의 최소 시간 간격")
        with col_adj:
            adj = st.checkbox("📌 연속 배치 (붙여서 진행)", value=True, key=f"adj_checkbox_{precedence_refresh_count}", help="두 활동을 시간적으로 연속해서 배치")
        
        if add_rule_btn:
            df = st.session_state["precedence"]
            dup = ((df["predecessor"] == p) & (df["successor"] == s)).any()
            if p == s:
                st.error("❌ 같은 활동끼리는 연결할 수 없습니다.")
            elif dup:
                st.warning("⚠️ 이미 존재하는 규칙입니다.")
            else:
                st.session_state["precedence"] = pd.concat(
                    [df, pd.DataFrame([{"predecessor": p, "successor": s, "gap_min": g, "adjacent": adj}])],
                    ignore_index=True
                )
                st.success(f"✅ 규칙 추가: {p} → {s}")
                st.rerun()
    
    # 설정된 규칙 관리 및 삭제
    with st.expander("🗂️ 설정된 규칙 관리", expanded=True):
        prec_df = st.session_state["precedence"].copy()
        
        if not prec_df.empty:
            # 규칙을 보기 좋게 표시
            prec_df["규칙표시용"] = prec_df.apply(
                lambda r: f"{r.predecessor} → {r.successor}" + 
                         (f" (간격: {r.gap_min}분)" if r.gap_min > 0 else "") +
                         (" [연속배치]" if r.adjacent else ""), axis=1
            )
            
            # 2단 구조: 왼쪽은 규칙 목록, 오른쪽은 삭제 기능
            col_rules, col_actions = st.columns([3, 2])
            
            with col_rules:
                st.markdown("**📋 현재 설정된 규칙들**")
                
                # START/END 규칙과 일반 규칙 분리해서 표시
                start_end_rules = prec_df[
                    (prec_df["predecessor"] == "__START__") | (prec_df["successor"] == "__END__")
                ]
                normal_rules = prec_df[
                    (~prec_df["predecessor"].isin(["__START__", "__END__"])) &
                    (~prec_df["successor"].isin(["__START__", "__END__"]))
                ]
                
                if not start_end_rules.empty:
                    st.markdown("**🏁 시작/끝 규칙:**")
                    for rule in start_end_rules["규칙표시용"]:
                        st.markdown(f"• {rule}")
                
                if not normal_rules.empty:
                    st.markdown("**🔗 순서 연결 규칙:**")
                    for rule in normal_rules["규칙표시용"]:
                        st.markdown(f"• {rule}")
            
            with col_actions:
                st.markdown("**🗑️ 규칙 삭제**")
                
                # 삭제할 규칙 선택
                delete_options = prec_df["규칙표시용"].tolist()
                to_delete = st.multiselect(
                    "삭제할 규칙 선택",
                    options=delete_options,
                    key=f"del_prec_select_{precedence_refresh_count}",
                    help="여러 규칙을 한 번에 선택 가능"
                )
                
                # 삭제 버튼들
                if st.button("❌ 선택 삭제", key="del_prec", disabled=not to_delete, use_container_width=True):
                    if to_delete:
                        new_prec = prec_df[~prec_df["규칙표시용"].isin(to_delete)].drop(
                            columns="규칙표시용"
                        ).reset_index(drop=True)
                        st.session_state["precedence"] = new_prec.copy()
                        st.success(f"✅ {len(to_delete)}개 규칙 삭제!")
                        st.rerun()
                
                if st.button("🗑️ 전체 삭제", key="clear_all_prec", type="secondary", use_container_width=True):
                    st.session_state["precedence"] = pd.DataFrame(columns=["predecessor", "successor", "gap_min", "adjacent"])
                    st.success("✅ 모든 규칙 삭제!")
                    st.rerun()
        else:
            st.info("📋 설정된 선후행 제약 규칙이 없습니다. 위에서 규칙을 추가해보세요.")

st.divider()

# =============================================================================
# 섹션 3: 직무별 면접활동 정의 (선후행 제약 설정 다음으로 이동)
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("3️⃣ 직무별 면접활동 정의")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_job_activities", help="직무별 면접활동 AG-Grid가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "job_activities" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["job_activities"] = 0
        st.session_state["section_refresh_counter"]["job_activities"] += 1
        st.rerun()

st.markdown("각 직무 코드별로 어떤 면접활동을 진행할지 설정하고 인원수를 지정합니다.")

# 활동 목록 확보
acts_df = st.session_state.get("activities")
if acts_df is None or acts_df.empty:
    st.error("먼저 면접활동을 정의해주세요.")
else:
    act_list = acts_df.query("use == True")["activity"].tolist()
    
    if not act_list:
        st.error("활동을 최소 1개 '사용'으로 체크해야 합니다.")
    else:
        # 직무 매핑 데이터 로드
        if "job_acts_map" in st.session_state:
            job_df = st.session_state["job_acts_map"].copy()
        else:
            job_df = pd.DataFrame({"code": ["JOB01"], "count": [1]})
        
        # 컬럼 동기화
        for act in act_list:
            if act not in job_df.columns:
                job_df[act] = True
        
        if "count" not in job_df.columns:
            job_df["count"] = 1
        
        # 열 순서 정리
        cols = ["code", "count"] + act_list
        job_df = job_df.reindex(columns=cols, fill_value=True)
        
        st.session_state["job_acts_map"] = job_df
        
        # 행 추가/삭제 기능
        col_add2, col_del2 = st.columns(2)
        
        with col_add2:
            if st.button("➕ 직무 코드 행 추가", key="add_job"):
                current = st.session_state["job_acts_map"].copy()
                
                # 새 코드 자동 생성
                pattern = re.compile(r"^([A-Za-z]+)(\d+)$")
                prefixes, numbers = [], []
                for c in current["code"]:
                    m = pattern.match(str(c))
                    if m:
                        prefixes.append(m.group(1))
                        numbers.append(int(m.group(2)))
                
                if prefixes:
                    pref = pd.Series(prefixes).mode()[0]
                    max_num = max(n for p, n in zip(prefixes, numbers) if p == pref)
                    next_num = max_num + 1
                else:
                    pref, next_num = "JOB", 1
                
                new_code = f"{pref}{next_num:02d}"
                new_row = {"code": new_code, "count": 1, **{act: True for act in act_list}}
                
                current = pd.concat([current, pd.DataFrame([new_row])], ignore_index=True)
                st.session_state["job_acts_map"] = current
                st.success(f"'{new_code}' 행이 추가되었습니다.")
                st.rerun()
        
        with col_del2:
            codes = job_df["code"].tolist()
            to_delete2 = st.multiselect(
                "삭제할 코드 선택", 
                options=codes, 
                key="del_job_select"
            )
            
            if st.button("❌ 선택 코드 삭제", key="del_job"):
                if to_delete2:
                    kept = job_df[~job_df["code"].isin(to_delete2)].reset_index(drop=True)
                    st.session_state["job_acts_map"] = kept
                    st.success("삭제 완료!")
                    st.rerun()
        
        # 편집용 AG-Grid
        df_to_display = st.session_state["job_acts_map"].copy()
        
        gb2 = GridOptionsBuilder.from_dataframe(df_to_display)
        gb2.configure_selection(selection_mode="none")
        gb2.configure_default_column(resizable=True, editable=True)
        
        gb2.configure_column("code", header_name="직무 코드", width=120, editable=True)
        gb2.configure_column(
            "count", header_name="인원수", type=["numericColumn"], width=90, editable=True
        )
        
        for act in act_list:
            gb2.configure_column(
                act,
                header_name=act,
                cellRenderer="agCheckboxCellRenderer",
                cellEditor="agCheckboxCellEditor",
                editable=True,
                singleClickEdit=True,
                width=110,
            )
        
        grid_opts2 = gb2.build()
        
        # 섹션별 새로고침을 위한 동적 key 생성
        job_activities_refresh_count = st.session_state.get("section_refresh_counter", {}).get("job_activities", 0)
        
        grid_ret2 = AgGrid(
            df_to_display,
            gridOptions=grid_opts2,
            update_mode=GridUpdateMode.VALUE_CHANGED,
            data_return_mode=DataReturnMode.AS_INPUT,
            fit_columns_on_grid_load=True,
            theme="balham",
            key=f"job_grid_display_{job_activities_refresh_count}",  # 동적 key로 강제 재렌더링
        )
        
        edited_df = pd.DataFrame(grid_ret2["data"])
        
        # 데이터 검증
        def validate_job_data(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
            msgs: list[str] = []
            df = df.copy()
            df["code"] = df["code"].astype(str).str.strip()
            df["count"] = pd.to_numeric(df["count"], errors="coerce").fillna(0).astype(int)
            
            # 빈 코드
            if (df["code"] == "").any():
                msgs.append("빈 코드가 있습니다.")
                df = df[df["code"] != ""].reset_index(drop=True)
            
            # 중복 코드
            dup_codes = df[df["code"].duplicated()]["code"].unique().tolist()
            if dup_codes:
                msgs.append(f"중복 코드: {', '.join(dup_codes)}")
            
            # count ≤ 0
            zero_cnt = df[df["count"] <= 0]["code"].tolist()
            if zero_cnt:
                msgs.append(f"0 이하 인원수: {', '.join(map(str, zero_cnt))}")
            
            # 활동이 하나도 선택되지 않은 행
            no_act = [
                row.code
                for row in df.itertuples()
                if not any(getattr(row, a) for a in act_list)
            ]
            if no_act:
                msgs.append(f"모든 활동이 False인 코드: {', '.join(no_act)}")
            
            return df, msgs
        
        clean_df, errors = validate_job_data(edited_df)
        
        if errors:
            for msg in errors:
                st.error(msg)
        else:
            st.success("모든 입력이 유효합니다!")
        
        st.info(f"총 인원수: **{clean_df['count'].sum()}** 명")
        st.session_state["job_acts_map"] = clean_df

st.divider()

# =============================================================================
# 섹션 4: 운영 공간 설정
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("4️⃣ 운영 공간 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_room_settings", help="운영 공간 설정 UI가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "room_settings" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["room_settings"] = 0
        st.session_state["section_refresh_counter"]["room_settings"] += 1
        st.rerun()

st.markdown("면접을 운영할 경우, 하루에 동원 가능한 모든 공간의 종류와 수, 그리고 최대 수용 인원을 설정합니다.")

# 활동 DF에서 room_types 확보
acts_df = st.session_state.get("activities")
if acts_df is not None and not acts_df.empty:
    room_types = sorted(
        acts_df.query("use == True and room_type != '' and room_type.notna()")["room_type"].unique()
    )
    
    if room_types:
        min_cap_req = acts_df.set_index("room_type")["min_cap"].to_dict()
        max_cap_req = acts_df.set_index("room_type")["max_cap"].to_dict()
        
        # 공간 템플릿 설정
        tpl_dict = st.session_state.get("room_template", {})
        
        # room_types 동기화
        for rt in room_types:
            tpl_dict.setdefault(rt, {"count": 1, "cap": max_cap_req.get(rt, 1)})
        for rt in list(tpl_dict):
            if rt not in room_types:
                tpl_dict.pop(rt)
        
        st.subheader("하루 기준 운영 공간 설정")
        
        # 섹션별 새로고침을 위한 동적 key 생성
        room_settings_refresh_count = st.session_state.get("section_refresh_counter", {}).get("room_settings", 0)
        
        col_cnt, col_cap = st.columns(2, gap="large")
        
        with col_cnt:
            st.markdown("#### 방 개수")
            for rt in room_types:
                tpl_dict[rt]["count"] = st.number_input(
                    f"{rt} 개수", 
                    min_value=0, 
                    max_value=50, 
                    value=tpl_dict[rt].get("count", 1), 
                    key=f"tpl_{rt}_cnt_{room_settings_refresh_count}"  # 동적 key로 강제 재렌더링
                )
        
        with col_cap:
            st.markdown("#### 최대 동시 수용 인원")
            for rt in room_types:
                min_val = min_cap_req.get(rt, 1)
                max_val = max_cap_req.get(rt, 50)
                current_val = tpl_dict[rt].get("cap", max_val)
                safe_val = max(min_val, min(current_val, max_val))
                
                tpl_dict[rt]["cap"] = st.number_input(
                    f"{rt} 최대 동시 수용 인원",
                    min_value=min_val,
                    max_value=max_val,
                    value=safe_val,
                    key=f"tpl_{rt}_cap_{room_settings_refresh_count}",  # 동적 key로 강제 재렌더링
                )
        
        # 변경된 템플릿 정보를 세션에 저장
        st.session_state['room_template'] = tpl_dict
        
        # room_plan 생성 및 저장
        final_plan_dict = {}
        for rt, values in tpl_dict.items():
            final_plan_dict[f"{rt}_count"] = values['count']
            final_plan_dict[f"{rt}_cap"] = values['cap']
        
        st.session_state['room_plan'] = pd.DataFrame([final_plan_dict])
        
        # Room Plan → Activities 역방향 자동 연동
        activities_df = st.session_state.get("activities", pd.DataFrame())
        if not activities_df.empty:
            capacity_changed = False
            for rt, values in tpl_dict.items():
                new_capacity = values['cap']
                
                # 해당 room_type을 사용하는 활동들의 max_cap 업데이트
                room_activities = activities_df[activities_df["room_type"] == rt]
                for idx in room_activities.index:
                    current_max_cap = activities_df.at[idx, 'max_cap']
                    if current_max_cap != new_capacity:
                        activities_df.at[idx, 'max_cap'] = new_capacity
                        # min_cap도 조정 (max_cap보다 크면 안됨)
                        if activities_df.at[idx, 'min_cap'] > new_capacity:
                            activities_df.at[idx, 'min_cap'] = new_capacity
                        capacity_changed = True
            
            if capacity_changed:
                st.session_state["activities"] = activities_df
                st.info("📝 방 수용 인원 변경에 따라 활동의 그룹 크기가 자동 조정되었습니다.")
        
        with st.expander("🗂 저장된 room_plan 데이터 미리보기"):
            st.dataframe(st.session_state.get('room_plan', pd.DataFrame()), use_container_width=True)
    else:
        st.error("사용(use=True)하도록 설정된 활동 중, 'room_type'이 지정된 활동이 없습니다.")

st.divider()

# =============================================================================
# 섹션 5: 운영 시간 설정
# =============================================================================
col_header, col_refresh = st.columns([3, 2])
with col_header:
    st.header("5️⃣ 운영 시간 설정")
with col_refresh:
    st.markdown("<br>", unsafe_allow_html=True)  # 헤더와 높이 맞추기
    if st.button("🔄 섹션 새로고침", key="refresh_time_settings", help="운영 시간 설정 UI가 먹통일 때 새로고침"):
        # 섹션별 새로고침
        if "section_refresh_counter" not in st.session_state:
            st.session_state["section_refresh_counter"] = {}
        if "time_settings" not in st.session_state["section_refresh_counter"]:
            st.session_state["section_refresh_counter"]["time_settings"] = 0
        st.session_state["section_refresh_counter"]["time_settings"] += 1
        st.rerun()

st.markdown("면접을 운영할 경우의 하루 기준 운영 시작 및 종료 시간을 설정합니다.")

# 기존 값 불러오기
init_start = st.session_state.get("oper_start_time", time(9, 0))
init_end = st.session_state.get("oper_end_time", time(18, 0))

st.subheader("하루 기준 공통 운영 시간")

# 섹션별 새로고침을 위한 동적 key 생성
time_settings_refresh_count = st.session_state.get("section_refresh_counter", {}).get("time_settings", 0)

col_start, col_end = st.columns(2)

with col_start:
    t_start = st.time_input("운영 시작 시간", value=init_start, key=f"oper_start_{time_settings_refresh_count}")

with col_end:
    t_end = st.time_input("운영 종료 시간", value=init_end, key=f"oper_end_{time_settings_refresh_count}")

if t_start >= t_end:
    st.error("오류: 운영 시작 시간은 종료 시간보다 빨라야 합니다.")
else:
    # 설정된 시간을 세션 상태에 저장
    st.session_state["oper_start_time"] = t_start
    st.session_state["oper_end_time"] = t_end
    
    # oper_window DataFrame 생성 및 저장
    oper_window_dict = {
        "start_time": t_start.strftime("%H:%M"),
        "end_time": t_end.strftime("%H:%M")
    }
    st.session_state['oper_window'] = pd.DataFrame([oper_window_dict])
    
    with st.expander("🗂 저장된 oper_window 데이터 미리보기"):
        st.dataframe(st.session_state.get('oper_window', pd.DataFrame()), use_container_width=True)
    
    st.success(f"운영 시간이 {t_start.strftime('%H:%M')}부터 {t_end.strftime('%H:%M')}까지로 설정되었습니다.")

st.divider()

# 위로 가기 기능
st.markdown("---")

# 중앙 링크 방식
st.markdown("""
<div style="text-align: center; margin: 20px 0;">
    <a href="#top" style="
        display: inline-block;
        background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
        color: white;
        text-decoration: none;
        padding: 12px 24px;
        border-radius: 25px;
        font-weight: bold;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        transition: all 0.3s ease;
    ">
        ⬆️ 빠른 이동: 맨 위로
    </a>
</div>
""", unsafe_allow_html=True)

# 우하단 고정 버튼 (개선된 버전)
st.markdown("""
<style>
    .floating-top-button {
        position: fixed;
        bottom: 30px;
        right: 30px;
        z-index: 999;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 50%;
        width: 60px;
        height: 60px;
        font-size: 20px;
        cursor: pointer;
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
        transition: all 0.3s ease;
        display: flex;
        align-items: center;
        justify-content: center;
        text-decoration: none;
    }
    
    .floating-top-button:hover {
        transform: translateY(-3px) scale(1.1);
        box-shadow: 0 6px 25px rgba(0,0,0,0.4);
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
    
    .floating-top-button:active {
        transform: translateY(-1px) scale(1.05);
    }
</style>

<a href="#top" class="floating-top-button" title="맨 위로 이동">
    ⬆️
</a>
""", unsafe_allow_html=True)

# 섹션 6은 섹션 1 아래로 이동됨